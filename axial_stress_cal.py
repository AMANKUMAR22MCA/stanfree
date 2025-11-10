# import pandas as pd

# # Input CSV file
# INPUT_CSV = "column_actions_per_combo_allmembers.csv"
# OUTPUT_XLSX = "max_fx_per_member_xm.xlsx"

# # Read CSV
# df = pd.read_csv(INPUT_CSV)

# # Clean column names
# df.columns = [c.strip() for c in df.columns]

# # Ensure correct numeric types
# df["member"] = df["member"].astype(int)
# df["x_m"] = df["x_m"].astype(float)
# df["Fx (kN)"] = df["Fx (kN)"].astype(float)

# # Group by member and x_m → take max Fx
# df_max = (
#     df.groupby(["member", "x_m"], as_index=False)["Fx (kN)"]
#     .max()
#     .rename(columns={
#         "member": "member id",
#         "x_m": "xm",
#         "Fx (kN)": "max fx (kN)"
#     })
# )

# # Optional: add axial stress, limit, and check columns (empty placeholders)
# df_max["axial stress"] = ""
# df_max["limit"] = ""
# df_max["check (YES/NO)"] = ""

# # Save to Excel
# df_max.to_excel(OUTPUT_XLSX, index=False)
# print(f"✅ Output saved to {OUTPUT_XLSX}")



# import pandas as pd
# import openpyxl
# from openpyxl.styles import PatternFill
# from typing import Optional
# import tkinter as tk

# # ================================
# # SETTINGS
# # ================================
# INPUT_CSV = "column_actions_per_combo_allmembers.csv"
# COLUMN_FILE = "columns_with_beam_clear_height.xlsx"
# OUTPUT_XLSX = "max_fx_per_member_xm.xlsx"
# FCK = 30
# LIMIT = 0.4 * FCK

# # ================================
# # POPUP FUNCTION
# # ================================
# def _centered_popup(title: str, message: str, stay_seconds: Optional[int] = None) -> bool:
#     try:
#         root = tk.Tk()
#         root.title(title)
#         root.attributes("-topmost", True)
#         root.resizable(False, False)
#         root.configure(bg="#ffffff")

#         frm = tk.Frame(root, padx=25, pady=20, bg="#ffffff")
#         frm.pack(fill="both", expand=True)

#         lbl_title = tk.Label(frm, text="✅ Axial Stress Check Completed",
#                              font=("Segoe UI", 14, "bold"),
#                              bg="#ffffff", fg="#2C3E50")
#         lbl_title.pack(pady=(0, 10))

#         lbl_msg = tk.Label(frm,
#                            text=message + "\n\nKindly review the Excel file.",
#                            font=("Segoe UI", 11),
#                            bg="#ffffff", fg="#333333", justify="left")
#         lbl_msg.pack()

#         btn = tk.Button(frm, text="Close", width=12,
#                         font=("Segoe UI", 10, "bold"),
#                         command=root.destroy,
#                         bg="#3498DB", fg="white")
#         btn.pack(pady=(15, 0))

#         root.update_idletasks()
#         w = root.winfo_width()
#         h = root.winfo_height()
#         sw = root.winfo_screenwidth()
#         sh = root.winfo_screenheight()
#         x = (sw // 2) - (w // 2)
#         y = (sh // 3) - (h // 3)
#         root.geometry(f"{w}x{h}+{x}+{y}")

#         if stay_seconds is not None:
#             root.after(int(stay_seconds * 1000), root.destroy)

#         root.mainloop()
#         return True
#     except Exception as e:
#         print("Popup error:", e)
#         return False


# # ================================
# # MAIN PROCESS
# # ================================

# # Read CSV input
# df = pd.read_csv(INPUT_CSV)
# df.columns = [c.strip() for c in df.columns]

# # Ensure correct numeric types
# df["member"] = df["member"].astype(int)
# df["x_m"] = df["x_m"].astype(float)
# df["Fx (kN)"] = df["Fx (kN)"].astype(float)

# # Find max Fx per member and x_m
# df_max = (
#     df.groupby(["member", "x_m"], as_index=False)["Fx (kN)"]
#     .max()
#     .rename(columns={
#         "member": "member id",
#         "x_m": "xm",
#         "Fx (kN)": "max fx (kN)"
#     })
# )

# # Read column size data (b = depth, D = width)
# cols_df = pd.read_excel(COLUMN_FILE)
# cols_df.columns = [c.strip() for c in cols_df.columns]

# # Check required columns
# required_cols = ["Column ID", "Column Depth (mm)", "Column Width (mm)"]
# for col in required_cols:
#     if col not in cols_df.columns:
#         raise KeyError(f"❌ Missing required column in Excel: {col}")

# # Merge based on Column ID == member id
# merged = pd.merge(
#     df_max,
#     cols_df[["Column ID", "Column Depth (mm)", "Column Width (mm)"]],
#     left_on="member id",
#     right_on="Column ID",
#     how="left"
# )

# # Rename for clarity
# merged.rename(columns={
#     "Column Depth (mm)": "b (mm)",
#     "Column Width (mm)": "D (mm)"
# }, inplace=True)

# # Compute axial stress and checks
# merged["axial stress (N/mm²)"] = (merged["max fx (kN)"] * 1000) / (merged["b (mm)"] * merged["D (mm)"])
# merged["limit (N/mm²)"] = LIMIT
# merged["check (YES/NO)"] = merged["axial stress (N/mm²)"].apply(lambda x: "YES" if x < LIMIT else "NO")

# # Save to Excel
# merged.to_excel(OUTPUT_XLSX, index=False)

# # ================================
# # COLOR FAILED CELLS (RED)
# # ================================
# wb = openpyxl.load_workbook(OUTPUT_XLSX)
# ws = wb.active
# red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

# # Apply red fill where check = "NO"
# check_col = None
# for i, cell in enumerate(ws[1], start=1):
#     if cell.value == "check (YES/NO)":
#         check_col = i
#         break

# if check_col:
#     for row in range(2, ws.max_row + 1):
#         if ws.cell(row, check_col).value == "NO":
#             for col in range(1, ws.max_column + 1):
#                 ws.cell(row, col).fill = red_fill

# wb.save(OUTPUT_XLSX)
# wb.close()

# # Popup summary
# n_total = len(merged)
# n_fail = (merged["check (YES/NO)"] == "NO").sum()
# n_pass = n_total - n_fail
# _centered_popup("Axial Stress Check Summary", f"Total = {n_total}\n✓ Pass = {n_pass}\n✖ Fail = {n_fail}\nSaved to:\n{OUTPUT_XLSX}")

# print(f"✅ Done: Total={n_total}, Pass={n_pass}, Fail={n_fail} → {OUTPUT_XLSX}")

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from typing import Optional
import tkinter as tk

# ================================
# SETTINGS
# ================================
INPUT_CSV = "column_actions_per_combo_allmembers.csv"
COLUMN_FILE = "columns_with_beam_clear_height.xlsx"
OUTPUT_XLSX = "max_fx_per_member_xm_with_station_label.xlsx"
FCK = 30
LIMIT = 0.4 * FCK

# ================================
# POPUP FUNCTION
# ================================
def _centered_popup(title: str, message: str, stay_seconds: Optional[int] = None) -> bool:
    try:
        root = tk.Tk()
        root.title(title)
        root.attributes("-topmost", True)
        root.resizable(False, False)
        root.configure(bg="#ffffff")

        frm = tk.Frame(root, padx=25, pady=20, bg="#ffffff")
        frm.pack(fill="both", expand=True)

        lbl_title = tk.Label(frm, text=title,
                             font=("Segoe UI", 14, "bold"),
                             bg="#ffffff", fg="#2C3E50")
        lbl_title.pack(pady=(0, 10))

        lbl_msg = tk.Label(frm,
                           text=message + "\n\nKindly review the Excel file.",
                           font=("Segoe UI", 11),
                           bg="#ffffff", fg="#333333", justify="left")
        lbl_msg.pack()

        btn = tk.Button(frm, text="Close", width=12,
                        font=("Segoe UI", 10, "bold"),
                        command=root.destroy,
                        bg="#3498DB", fg="white")
        btn.pack(pady=(15, 0))

        root.update_idletasks()
        w = root.winfo_width()
        h = root.winfo_height()
        sw = root.winfo_screenwidth()
        sh = root.winfo_screenheight()
        x = (sw // 2) - (w // 2)
        y = (sh // 3) - (h // 3)
        root.geometry(f"{w}x{h}+{x}+{y}")

        if stay_seconds is not None:
            root.after(int(stay_seconds * 1000), root.destroy)

        root.mainloop()
        return True
    except Exception as e:
        print("Popup error:", e)
        return False

# ================================
# MAIN PROCESS
# ================================
# Read CSV input
df = pd.read_csv(INPUT_CSV)
df.columns = [c.strip() for c in df.columns]

# Ensure correct numeric types
df["member"] = df["member"].astype(int)
df["x_m"] = df["x_m"].astype(float)
df["Fx (kN)"] = df["Fx (kN)"].astype(float)

# ✅ Keep station_label for each (member, x_m)
# Take first station_label found for that combination
station_labels = (
    df.groupby(["member", "x_m"], as_index=False)["station_label"]
    .first()
)

# Find max Fx per member and x_m
df_max = (
    df.groupby(["member", "x_m"], as_index=False)["Fx (kN)"]
    .max()
    .rename(columns={
        "member": "member id",
        "x_m": "xm",
        "Fx (kN)": "max fx (kN)"
    })
)

# ✅ Merge station_label info
df_max = pd.merge(
    df_max,
    station_labels.rename(columns={"station_label": "Station Label"}),
    left_on=["member id", "xm"],
    right_on=["member", "x_m"],
    how="left"
).drop(columns=["member", "x_m"], errors="ignore")

# Read column size data (b = depth, D = width)
cols_df = pd.read_excel(COLUMN_FILE)
cols_df.columns = [c.strip() for c in cols_df.columns]

# Check required columns
required_cols = ["Column ID", "Column Depth (mm)", "Column Width (mm)"]
for col in required_cols:
    if col not in cols_df.columns:
        raise KeyError(f"❌ Missing required column in Excel: {col}")

# Merge based on Column ID == member id
merged = pd.merge(
    df_max,
    cols_df[["Column ID", "Column Depth (mm)", "Column Width (mm)"]],
    left_on="member id",
    right_on="Column ID",
    how="left"
)

# Rename for clarity
merged.rename(columns={
    "Column Depth (mm)": "b (mm)",
    "Column Width (mm)": "D (mm)"
}, inplace=True)

# Compute axial stress and checks
def compute_axial(fx_kN, b_mm, D_mm):
    try:
        if pd.isna(b_mm) or pd.isna(D_mm) or b_mm <= 0 or D_mm <= 0:
            return None
        return (fx_kN * 1000.0) / (b_mm * D_mm)
    except Exception:
        return None

merged["axial stress (N/mm²)"] = merged.apply(
    lambda r: compute_axial(r["max fx (kN)"], r["b (mm)"], r["D (mm)"]),
    axis=1
)

merged["limit (N/mm²)"] = LIMIT
merged["check (YES/NO)"] = merged["axial stress (N/mm²)"].apply(
    lambda x: "NO DATA" if x is None else ("YES" if x < LIMIT else "NO")
)

# --- remove Column ID ---
if "Column ID" in merged.columns:
    merged.drop(columns=["Column ID"], inplace=True)

# ✅ Arrange columns (place Station Label after xm)
cols_order = ["member id", "xm", "Station Label", "max fx (kN)", "b (mm)", "D (mm)", 
              "axial stress (N/mm²)", "limit (N/mm²)", "check (YES/NO)"]
cols_order = [c for c in cols_order if c in merged.columns]
merged = merged[cols_order + [c for c in merged.columns if c not in cols_order]]

# Save to Excel
merged.to_excel(OUTPUT_XLSX, index=False)

# ================================
# COLOR FAILED CELLS (RED)
# ================================
wb = openpyxl.load_workbook(OUTPUT_XLSX)
ws = wb.active
red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

# Find check column
check_col = None
for i, cell in enumerate(ws[1], start=1):
    if isinstance(cell.value, str) and cell.value.strip().lower() == "check (yes/no)":
        check_col = i
        break

if check_col:
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=check_col).value
        if val == "NO":
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = red_fill

wb.save(OUTPUT_XLSX)
wb.close()

# ================================
# POPUP SUMMARY
# ================================
n_total = len(merged)
n_fail = (merged["check (YES/NO)"] == "NO").sum()
n_pass = n_total - n_fail
_centered_popup("Axial Stress Check Summary", f"Total = {n_total}\n✓ Pass = {n_pass}\n✖ Fail = {n_fail}\nSaved to:\n{OUTPUT_XLSX}")

print(f"✅ Done: Total={n_total}, Pass={n_pass}, Fail={n_fail} → {OUTPUT_XLSX}")
