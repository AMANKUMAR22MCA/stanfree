
import re
import math
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# ================================
# CONFIG
# ================================
INPUT_FILE   = "MR. DINESH KHATRI.ANL"      # path to your .ANL file
OUTPUT_XLSX  = "column_design_data_checked.xlsx"
TOLERANCE_MM2 = 4                           # ± tolerance for area match

POPUP_MODE: str = "popup"
POPUP_SECONDS: Optional[int] = None         # keep popup open until close button

# ================================
# ALWAYS-WORKING POPUP
# ================================
def _centered_popup(title: str, message: str, stay_seconds: Optional[int] = None) -> bool:
    """
    GUARANTEED WORKING POPUP (Tkinter main-thread)
    """
    try:
        import tkinter as tk

        root = tk.Tk()
        root.title(title)
        root.attributes("-topmost", True)
        root.resizable(False, False)
        root.configure(bg="#ffffff")

        # Frame
        frm = tk.Frame(root, padx=25, pady=20, bg="#ffffff")
        frm.pack(fill="both", expand=True)

        # Title
        lbl_title = tk.Label(
            frm,
            text="✅ Please Check Sheet",
            font=("Segoe UI", 14, "bold"),
            bg="#ffffff",
            fg="#2C3E50"
        )
        lbl_title.pack(pady=(0, 10))

        # Main message
        lbl_msg = tk.Label(
            frm,
            text=message + "\n\nWe have failure cases — kindly review quickly.",
            font=("Segoe UI", 11),
            bg="#ffffff",
            fg="#333333",
            justify="left"
        )
        lbl_msg.pack()

        # Close button
        btn = tk.Button(
            frm,
            text="Close",
            width=12,
            font=("Segoe UI", 10, "bold"),
            command=root.destroy,
            bg="#3498DB",
            fg="white"
        )
        btn.pack(pady=(15, 0))

        # Centering
        root.update_idletasks()
        w = root.winfo_width()
        h = root.winfo_height()
        sw = root.winfo_screenwidth()
        sh = root.winfo_screenheight()
        x = (sw // 2) - (w // 2)
        y = (sh // 2) - (h // 3)
        root.geometry(f"{w}x{h}+{x}+{y}")

        # Auto-close
        if stay_seconds is not None:
            root.after(int(stay_seconds * 1000), root.destroy)

        root.mainloop()
        return True

    except Exception as e:
        print("Popup error:", e)
        return False


# ================================
# SUMMARY NOTIFICATION
# ================================
def notify_summary(total: int, n_pass: int, n_fail: int, path: str,
                   mode: str = POPUP_MODE, popup_seconds: Optional[int] = POPUP_SECONDS) -> None:

    msg = f"Columns: {total}  ✓ Pass: {n_pass}  ✖ Fail: {n_fail}\nSaved: {path}"

    shown = False
    if mode.lower() == "popup":
        shown = _centered_popup("Column Design Check", msg, stay_seconds=popup_seconds)

    print(f"INFO: Columns={total}, Pass={n_pass}, Fail={n_fail} → {path}"
          + ("" if shown else "  (popup failed → console only)"))


# ================================
# READ .ANL
# ================================
with open(INPUT_FILE, "r", encoding="utf-8", errors="ignore") as f:
    text = f.read()

# Split into blocks
blocks = re.split(r"C\s*O\s*L\s*U\s*M\s*N\s+N\s*O\.\s*(\d+)", text)
columns = {}

for i in range(1, len(blocks), 2):
    col_no = blocks[i].strip()
    block  = blocks[i + 1]

    def find(pattern, flags=0, group=1):
        m = re.search(pattern, block, flags)
        if not m:
            return None
        val = m.group(group)
        return val.replace(",", "") if isinstance(val, str) else val

    guiding_load = find(r"GUIDING\s*LOAD\s*CASE:\s*(\d+)")
    guiding_joint = find(r"END\s*JOINT:\s*(\d+)")
    req_steel = find(r"REQD\.\s*STEEL\s*AREA\s*[:=]?\s*([\d.]+)")
    req_conc = find(r"REQD\.\s*CONCRETE\s*AREA\s*[:=]?\s*([\d.]+)")

    prov_steel = find(r"MAIN\s*REINFORCEMENT\s*:\s*Provide.*?\((?:\s*[\d.]+%[, ]+\s*)?([\d.]+)",
                      flags=re.DOTALL)
    ratio = find(r"MAIN\s*REINFORCEMENT\s*:\s*Provide.*?\(\s*([\d.]+)%",
                 flags=re.DOTALL)

    no_bars = find(r"Provide\s+(\d+)\s*-\s*\d+\s*dia")
    bar_dia = find(r"Provide\s+\d+\s*-\s*(\d+)\s*dia")

    conf_dia = find(r"CONFINING\s*REINFORCEMENT\s*:\s*Provide\s*(\d+)")
    conf_space = find(r"rectangular\s*ties\s*@\s*(\d+)")
    conf_len = find(r"over\s+a\s+length\s*([\d.]+)\s*mm")
    tie_dia = find(r"TIE\s*REINFORCEMENT\s*.*?Provide\s*(\d+)", flags=re.DOTALL)
    tie_space = find(r"TIE\s*REINFORCEMENT\s*.*?@\s*(\d+)", flags=re.DOTALL)

    bar_area = total_area = None
    if bar_dia and no_bars:
        try:
            dia = float(bar_dia)
            count = float(no_bars)
            bar_area = math.pi * (dia ** 2) / 4.0
            total_area = bar_area * count
        except:
            pass

    area_check = ""
    if prov_steel and total_area is not None:
        try:
            prov_val = float(prov_steel)
            area_check = "YES" if abs(prov_val - total_area) <= TOLERANCE_MM2 else "NO"
        except:
            pass

    min_ratio_check = max_ratio_check = min_bar_check = min_dia_check = ""
    try:
        ratio_val = float(ratio) if ratio else None
        bars_val = int(no_bars) if no_bars else None
        dia_val = float(bar_dia) if bar_dia else None

        if ratio_val is not None:
            min_ratio_check = "YES" if ratio_val > 0.8 else "NO"
            max_ratio_check = "YES" if ratio_val < 4 else "NO"

        if bars_val is not None:
            min_bar_check = "YES" if bars_val >= 4 else "NO"

        if dia_val is not None:
            min_dia_check = "YES" if dia_val >= 12 else "NO"
    except:
        pass

    columns[col_no] = {
        "Column No": col_no,
        "Guiding Load Case": guiding_load,
        "Guiding Joint": guiding_joint,
        "Req Steel Area (Sq.mm)": req_steel,
        "Req Concrete Area (Sq.mm)": req_conc,
        "Provided Steel Area (Sq.mm)": prov_steel,
        "Steel Ratio (%)": ratio,
        "No of Bars": no_bars,
        "Bar Dia (mm)": bar_dia,
        "Area 1 Bar (mm²)": round(bar_area, 2) if bar_area else "",
        "Total Bar Area (mm²)": round(total_area, 2) if total_area else "",
        "Area Check": area_check,
        "Confinement Dia (mm)": conf_dia,
        "Confinement Spacing (mm c/c)": conf_space,
        "Confinement Length (mm)": conf_len,
        "Tie Dia (mm)": tie_dia,
        "Tie Spacing (mm c/c)": tie_space,
        "Min Steel Ratio (>0.8)": min_ratio_check,
        "Max Steel Ratio (<4)": max_ratio_check,
        "Min 4 Bars (>=4)": min_bar_check,
        "Min Bar Dia (>=12)": min_dia_check,
    }

# ================================
# WRITE EXCEL
# ================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Column_Design_Data"

if not columns:
    print("❌ No columns found!")
    wb.save(OUTPUT_XLSX)
else:
    headers = list(next(iter(columns.values())).keys())
    ws.append(headers)

    for c in columns.values():
        ws.append([c.get(h, "") for h in headers])

    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(str(header)) + 4, 16), 40)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    bold_red = Font(color="9C0006", bold=True)

    check_cols = [
        "Area Check",
        "Min Steel Ratio (>0.8)",
        "Max Steel Ratio (<4)",
        "Min 4 Bars (>=4)",
        "Min Bar Dia (>=12)",
    ]
    header_to_index = {h: i for i, h in enumerate(headers)}

    fail_rows = []

    for r in range(2, ws.max_row + 1):
        row_failed = False
        for col_name in check_cols:
            idx = header_to_index.get(col_name)
            if idx is None:
                continue
            cell = ws.cell(row=r, column=idx + 1)
            if str(cell.value).strip().upper() == "NO":
                cell.fill = red_fill
                cell.font = bold_red
                row_failed = True
        if row_failed:
            fail_rows.append(r)

    total = len(columns)
    n_fail = len(fail_rows)
    n_pass = total - n_fail

    ws.append([])
    ws.append(["Summary", f"Total = {total}", f"Pass = {n_pass}", f"Fail = {n_fail}"])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)

    ws_fail = wb.create_sheet("Failures_Only")
    ws_fail.append(headers + ["Fail Flags"])

    for row_idx in fail_rows:
        values = [ws.cell(row=row_idx, column=j + 1).value for j in range(len(headers))]
        flags = []
        for col_name in check_cols:
            idx = header_to_index[col_name]
            v = ws.cell(row=row_idx, column=idx + 1).value
            if str(v).strip().upper() == "NO":
                flags.append(col_name)
        ws_fail.append(values + [", ".join(flags)])

    for cell in ws_fail[1]:
        cell.fill = header_fill
        cell.font = header_font

    wb.save(OUTPUT_XLSX)

    print(f"✅ Extracted {total} columns → {OUTPUT_XLSX}  |  ✓ Pass: {n_pass}  ✖ Fail: {n_fail}")

    notify_summary(total, n_pass, n_fail, OUTPUT_XLSX, mode=POPUP_MODE, popup_seconds=POPUP_SECONDS)




# import re
# import math
# import openpyxl

# # ========== SETTINGS ==========
# INPUT_FILE = "MR. DINESH KHATRI.ANL"        # path to your .ANL file
# OUTPUT_XLSX = "column_design_data.xlsx"

# # ========== READ FILE ==========
# with open(INPUT_FILE, "r", encoding="utf-8", errors="ignore") as f:
#     text = f.read()

# # Split file into column blocks
# blocks = re.split(r"C\s*O\s*L\s*U\s*M\s*N\s+N\s*O\.\s*(\d+)", text)
# # pattern returns [before_text, col_no1, block1, col_no2, block2, ...]
# columns = []

# for i in range(1, len(blocks), 2):
#     col_no = blocks[i].strip()
#     block = blocks[i + 1]

#     def find(pattern, flags=0):
#         m = re.search(pattern, block, flags)
#         return m.group(1).replace(",", "") if m else None

#     # Basic values
#     guiding_load = find(r"GUIDING\s*LOAD\s*CASE:\s*(\d+)")
#     guiding_joint = find(r"END\s*JOINT:\s*(\d+)")
#     req_steel = find(r"REQD\.\s*STEEL\s*AREA\s*[:=]?\s*([\d.]+)")
#     req_conc = find(r"REQD\.\s*CONCRETE\s*AREA\s*[:=]?\s*([\d.]+)")
#     prov_steel = find(r"MAIN\s*REINFORCEMENT\s*:\s*Provide.*?\((?:\s*[\d.]+%[, ]+\s*)?([\d.]+)", flags=re.DOTALL)
#     ratio = find(r"MAIN\s*REINFORCEMENT\s*:\s*Provide.*?\(\s*([\d.]+)%", flags=re.DOTALL)
#     no_bars = find(r"Provide\s+(\d+)\s*-\s*\d+\s*dia")
#     bar_dia = find(r"Provide\s+\d+\s*-\s*(\d+)\s*dia")

#     # Confinement
#     conf_dia = find(r"CONFINING\s*REINFORCEMENT\s*:\s*Provide\s*(\d+)")
#     conf_space = find(r"rectangular\s*ties\s*@\s*(\d+)")
#     conf_len = find(r"over\s+a\s+length\s*([\d.]+)\s*mm")

#     # Tie
#     tie_dia = find(r"TIE\s*REINFORCEMENT\s*.*?Provide\s*(\d+)", flags=re.DOTALL)
#     tie_space = find(r"TIE\s*REINFORCEMENT\s*.*?@\s*(\d+)", flags=re.DOTALL)

#     # Calculations
#     if bar_dia and no_bars:
#         bar_area = math.pi * (float(bar_dia) ** 2) / 4
#         total_area = bar_area * float(no_bars)
#     else:
#         bar_area = total_area = None

#     columns.append({
#         "Column No": col_no,
#         "Guiding Load Case": guiding_load,
#         "Guiding Joint": guiding_joint,
#         "Req Steel Area (Sq.mm)": req_steel,
#         "Req Concrete Area (Sq.mm)": req_conc,
#         "Provided Steel Area (Sq.mm)": prov_steel,
#         "Steel Ratio (%)": ratio,
#         "No of Bars": no_bars,
#         "Bar Dia (mm)": bar_dia,
#         "Area 1 Bar (mm²)": round(bar_area, 2) if bar_area else "",
#         "Total Bar Area (mm²)": round(total_area, 2) if total_area else "",
#         "Confinement Dia (mm)": conf_dia,
#         "Confinement Spacing (mm c/c)": conf_space,
#         "Confinement Length (mm)": conf_len,
#         "Tie Dia (mm)": tie_dia,
#         "Tie Spacing (mm c/c)": tie_space
#     })

# # ========== WRITE TO EXCEL ==========
# wb = openpyxl.Workbook()
# ws = wb.active
# ws.title = "Column_Design_Data"

# headers = list(columns[0].keys())
# ws.append(headers)

# for c in columns:
#     ws.append([c.get(h, "") for h in headers])

# wb.save(OUTPUT_XLSX)
# print(f"✅ Extracted {len(columns)} column(s) → {OUTPUT_XLSX}")



# import re
# import math
# import openpyxl

# # ========== SETTINGS ==========
# INPUT_FILE = "MR. DINESH KHATRI.ANL"        # path to your .ANL file
# OUTPUT_XLSX = "column_design_data_checked.xlsx"

# # ========== READ FILE ==========
# with open(INPUT_FILE, "r", encoding="utf-8", errors="ignore") as f:
#     text = f.read()

# # Split file into column blocks
# blocks = re.split(r"C\s*O\s*L\s*U\s*M\s*N\s+N\s*O\.\s*(\d+)", text)
# columns = {}

# for i in range(1, len(blocks), 2):
#     col_no = blocks[i].strip()
#     block = blocks[i + 1]

#     def find(pattern, flags=0):
#         m = re.search(pattern, block, flags)
#         return m.group(1).replace(",", "") if m else None

#     # Basic values
#     guiding_load = find(r"GUIDING\s*LOAD\s*CASE:\s*(\d+)")
#     guiding_joint = find(r"END\s*JOINT:\s*(\d+)")
#     req_steel = find(r"REQD\.\s*STEEL\s*AREA\s*[:=]?\s*([\d.]+)")
#     req_conc = find(r"REQD\.\s*CONCRETE\s*AREA\s*[:=]?\s*([\d.]+)")
#     prov_steel = find(r"MAIN\s*REINFORCEMENT\s*:\s*Provide.*?\((?:\s*[\d.]+%[, ]+\s*)?([\d.]+)", flags=re.DOTALL)
#     ratio = find(r"MAIN\s*REINFORCEMENT\s*:\s*Provide.*?\(\s*([\d.]+)%", flags=re.DOTALL)
#     no_bars = find(r"Provide\s+(\d+)\s*-\s*\d+\s*dia")
#     bar_dia = find(r"Provide\s+\d+\s*-\s*(\d+)\s*dia")

#     # Confinement
#     conf_dia = find(r"CONFINING\s*REINFORCEMENT\s*:\s*Provide\s*(\d+)")
#     conf_space = find(r"rectangular\s*ties\s*@\s*(\d+)")
#     conf_len = find(r"over\s+a\s+length\s*([\d.]+)\s*mm")

#     # Tie
#     tie_dia = find(r"TIE\s*REINFORCEMENT\s*.*?Provide\s*(\d+)", flags=re.DOTALL)
#     tie_space = find(r"TIE\s*REINFORCEMENT\s*.*?@\s*(\d+)", flags=re.DOTALL)

#     # Calculations
#     bar_area = total_area = None
#     if bar_dia and no_bars:
#         bar_area = math.pi * (float(bar_dia) ** 2) / 4
#         total_area = bar_area * float(no_bars)

#     # Area check logic (within ±10 tolerance)
#     area_check = ""
#     if prov_steel and total_area:
#         try:
#             prov_val = float(prov_steel)
#             if abs(prov_val - total_area) <= 10:
#                 area_check = "YES"
#             else:
#                 area_check = "NO"
#         except ValueError:
#             pass

#     # Add / overwrite unique by column number
#     columns[col_no] = {
#         "Column No": col_no,
#         "Guiding Load Case": guiding_load,
#         "Guiding Joint": guiding_joint,
#         "Req Steel Area (Sq.mm)": req_steel,
#         "Req Concrete Area (Sq.mm)": req_conc,
#         "Provided Steel Area (Sq.mm)": prov_steel,
#         "Steel Ratio (%)": ratio,
#         "No of Bars": no_bars,
#         "Bar Dia (mm)": bar_dia,
#         "Area 1 Bar (mm²)": round(bar_area, 2) if bar_area else "",
#         "Total Bar Area (mm²)": round(total_area, 2) if total_area else "",
#         "Area Check": area_check,
#         "Confinement Dia (mm)": conf_dia,
#         "Confinement Spacing (mm c/c)": conf_space,
#         "Confinement Length (mm)": conf_len,
#         "Tie Dia (mm)": tie_dia,
#         "Tie Spacing (mm c/c)": tie_space,
#     }

# # ========== WRITE TO EXCEL ==========
# wb = openpyxl.Workbook()
# ws = wb.active
# ws.title = "Column_Design_Data"

# if not columns:
#     print("❌ No columns found in file!")
# else:
#     headers = list(next(iter(columns.values())).keys())
#     ws.append(headers)
#     for c in columns.values():
#         ws.append([c.get(h, "") for h in headers])

#     wb.save(OUTPUT_XLSX)
#     print(f"✅ Extracted {len(columns)} unique columns → {OUTPUT_XLSX}")



# import re
# import math
# import openpyxl

# # ========== SETTINGS ==========
# INPUT_FILE = "MR. DINESH KHATRI.ANL"        # path to your .ANL file
# OUTPUT_XLSX = "column_design_data_checked.xlsx"
# TOLERANCE_MM2 = 4  # ± tolerance for area match check

# # ========== READ FILE ==========
# with open(INPUT_FILE, "r", encoding="utf-8", errors="ignore") as f:
#     text = f.read()

# # Split file into column blocks
# blocks = re.split(r"C\s*O\s*L\s*U\s*M\s*N\s+N\s*O\.\s*(\d+)", text)
# columns = {}

# for i in range(1, len(blocks), 2):
#     col_no = blocks[i].strip()
#     block = blocks[i + 1]

#     def find(pattern, flags=0):
#         m = re.search(pattern, block, flags)
#         return m.group(1).replace(",", "") if m else None

#     # Basic values
#     guiding_load = find(r"GUIDING\s*LOAD\s*CASE:\s*(\d+)")
#     guiding_joint = find(r"END\s*JOINT:\s*(\d+)")
#     req_steel = find(r"REQD\.\s*STEEL\s*AREA\s*[:=]?\s*([\d.]+)")
#     req_conc = find(r"REQD\.\s*CONCRETE\s*AREA\s*[:=]?\s*([\d.]+)")
#     prov_steel = find(r"MAIN\s*REINFORCEMENT\s*:\s*Provide.*?\((?:\s*[\d.]+%[, ]+\s*)?([\d.]+)", flags=re.DOTALL)
#     ratio = find(r"MAIN\s*REINFORCEMENT\s*:\s*Provide.*?\(\s*([\d.]+)%", flags=re.DOTALL)
#     no_bars = find(r"Provide\s+(\d+)\s*-\s*\d+\s*dia")
#     bar_dia = find(r"Provide\s+\d+\s*-\s*(\d+)\s*dia")

#     # Confinement
#     conf_dia = find(r"CONFINING\s*REINFORCEMENT\s*:\s*Provide\s*(\d+)")
#     conf_space = find(r"rectangular\s*ties\s*@\s*(\d+)")
#     conf_len = find(r"over\s+a\s+length\s*([\d.]+)\s*mm")

#     # Tie
#     tie_dia = find(r"TIE\s*REINFORCEMENT\s*.*?Provide\s*(\d+)", flags=re.DOTALL)
#     tie_space = find(r"TIE\s*REINFORCEMENT\s*.*?@\s*(\d+)", flags=re.DOTALL)

#     # Calculations
#     bar_area = total_area = None
#     if bar_dia and no_bars:
#         bar_area = math.pi * (float(bar_dia) ** 2) / 4
#         total_area = bar_area * float(no_bars)

#     # Area check logic (±4 tolerance)
#     area_check = ""
#     if prov_steel and total_area:
#         try:
#             prov_val = float(prov_steel)
#             if abs(prov_val - total_area) <= TOLERANCE_MM2:
#                 area_check = "YES"
#             else:
#                 area_check = "NO"
#         except ValueError:
#             pass

#     # Design Checks
#     min_ratio_check = max_ratio_check = min_bar_check = min_dia_check = ""

#     try:
#         ratio_val = float(ratio) if ratio else None
#         bars_val = int(no_bars) if no_bars else None
#         dia_val = float(bar_dia) if bar_dia else None

#         # 1️⃣ Minimum steel ratio check (>0.8)
#         if ratio_val is not None:
#             min_ratio_check = "YES" if ratio_val > 0.8 else "NO"

#         # 2️⃣ Maximum steel ratio check (<4)
#         if ratio_val is not None:
#             max_ratio_check = "YES" if ratio_val < 4 else "NO"

#         # 3️⃣ Minimum 4 bars check
#         if bars_val is not None:
#             min_bar_check = "YES" if bars_val >= 4 else "NO"

#         # 4️⃣ Minimum bar diameter check
#         if dia_val is not None:
#             min_dia_check = "YES" if dia_val >= 12 else "NO"

#     except Exception:
#         pass

#     # Add / overwrite unique by column number
#     columns[col_no] = {
#         "Column No": col_no,
#         "Guiding Load Case": guiding_load,
#         "Guiding Joint": guiding_joint,
#         "Req Steel Area (Sq.mm)": req_steel,
#         "Req Concrete Area (Sq.mm)": req_conc,
#         "Provided Steel Area (Sq.mm)": prov_steel,
#         "Steel Ratio (%)": ratio,
#         "No of Bars": no_bars,
#         "Bar Dia (mm)": bar_dia,
#         "Area 1 Bar (mm²)": round(bar_area, 2) if bar_area else "",
#         "Total Bar Area (mm²)": round(total_area, 2) if total_area else "",
#         "Area Check": area_check,
#         "Confinement Dia (mm)": conf_dia,
#         "Confinement Spacing (mm c/c)": conf_space,
#         "Confinement Length (mm)": conf_len,
#         "Tie Dia (mm)": tie_dia,
#         "Tie Spacing (mm c/c)": tie_space,
#         "Min Steel Ratio (>0.8)": min_ratio_check,
#         "Max Steel Ratio (<4)": max_ratio_check,
#         "Min 4 Bars (>=4)": min_bar_check,
#         "Min Bar Dia (>=12)": min_dia_check,
#     }

# # ========== WRITE TO EXCEL ==========
# wb = openpyxl.Workbook()
# ws = wb.active
# ws.title = "Column_Design_Data"

# if not columns:
#     print("❌ No columns found in file!")
# else:
#     headers = list(next(iter(columns.values())).keys())
#     ws.append(headers)
#     for c in columns.values():
#         ws.append([c.get(h, "") for h in headers])

#     wb.save(OUTPUT_XLSX)
#     print(f"✅ Extracted {len(columns)} unique columns → {OUTPUT_XLSX}")



# import re
# import math
# import openpyxl
# from openpyxl.styles import PatternFill, Font
# from openpyxl.utils import get_column_letter

# # ---------- CONFIG ----------
# INPUT_FILE = "MR. DINESH KHATRI.ANL"     # your .ANL file
# OUTPUT_XLSX = "column_design_data_checked.xlsx"
# TOLERANCE_MM2 = 4                        # ± tolerance for provided area vs computed
# # ----------------------------

# # Optional: non-blocking toast on Windows (if available)
# def notify_summary(total, n_pass, n_fail, path):
#     """
#     Non-blocking toast if win10toast is available; otherwise just prints.
#     No hard import to avoid VS Code 'could not be resolved' warnings.
#     """
#     try:
#         win10toast = __import__("win10toast")
#         toaster = win10toast.ToastNotifier()
#         msg = f"Columns: {total}  ✓ Pass: {n_pass}  ✖ Fail: {n_fail}\nSaved: {path}"
#         toaster.show_toast("Column Design Check", msg, duration=5, threaded=True)
#     except Exception:
#         print(f"INFO: Columns={total}, Pass={n_pass}, Fail={n_fail} → {path}")


# # ---------- READ .ANL ----------
# with open(INPUT_FILE, "r", encoding="utf-8", errors="ignore") as f:
#     text = f.read()

# # Split file into column blocks
# blocks = re.split(r"C\s*O\s*L\s*U\s*M\s*N\s+N\s*O\.\s*(\d+)", text)
# columns = {}

# for i in range(1, len(blocks), 2):
#     col_no = blocks[i].strip()
#     block = blocks[i + 1]

#     def find(pattern, flags=0, group=1):
#         m = re.search(pattern, block, flags)
#         if not m:
#             return None
#         val = m.group(group)
#         return val.replace(",", "") if isinstance(val, str) else val

#     # Basic values
#     guiding_load = find(r"GUIDING\s*LOAD\s*CASE:\s*(\d+)")
#     guiding_joint = find(r"END\s*JOINT:\s*(\d+)")
#     req_steel = find(r"REQD\.\s*STEEL\s*AREA\s*[:=]?\s*([\d.]+)")
#     req_conc = find(r"REQD\.\s*CONCRETE\s*AREA\s*[:=]?\s*([\d.]+)")

#     # Provided steel & ratio from typical STAAD strings
#     prov_steel = find(r"MAIN\s*REINFORCEMENT\s*:\s*Provide.*?\((?:\s*[\d.]+%[, ]+\s*)?([\d.]+)", flags=re.DOTALL)
#     ratio = find(r"MAIN\s*REINFORCEMENT\s*:\s*Provide.*?\(\s*([\d.]+)%", flags=re.DOTALL)

#     no_bars = find(r"Provide\s+(\d+)\s*-\s*\d+\s*dia")
#     bar_dia = find(r"Provide\s+\d+\s*-\s*(\d+)\s*dia")

#     # Confinement / ties (best-effort)
#     conf_dia = find(r"CONFINING\s*REINFORCEMENT\s*:\s*Provide\s*(\d+)")
#     conf_space = find(r"rectangular\s*ties\s*@\s*(\d+)")
#     conf_len = find(r"over\s+a\s+length\s*([\d.]+)\s*mm")
#     tie_dia = find(r"TIE\s*REINFORCEMENT\s*.*?Provide\s*(\d+)", flags=re.DOTALL)
#     tie_space = find(r"TIE\s*REINFORCEMENT\s*.*?@\s*(\d+)", flags=re.DOTALL)

#     # Calculations
#     bar_area = total_area = None
#     if bar_dia and no_bars:
#         try:
#             dia = float(bar_dia)
#             count = float(no_bars)
#             bar_area = math.pi * (dia ** 2) / 4.0
#             total_area = bar_area * count
#         except Exception:
#             pass

#     # Area check (± tolerance)
#     area_check = ""
#     if prov_steel and total_area is not None:
#         try:
#             prov_val = float(prov_steel)
#             area_check = "YES" if abs(prov_val - total_area) <= TOLERANCE_MM2 else "NO"
#         except ValueError:
#             pass

#     # Design checks
#     min_ratio_check = max_ratio_check = min_bar_check = min_dia_check = ""
#     try:
#         ratio_val = float(ratio) if ratio else None
#         bars_val = int(no_bars) if no_bars else None
#         dia_val = float(bar_dia) if bar_dia else None

#         # 1) Minimum steel ratio check (>0.8)
#         if ratio_val is not None:
#             min_ratio_check = "YES" if ratio_val > 0.8 else "NO"

#         # 2) Maximum steel ratio check (<4)
#         if ratio_val is not None:
#             max_ratio_check = "YES" if ratio_val < 4 else "NO"

#         # 3) Minimum 4 bars check
#         if bars_val is not None:
#             min_bar_check = "YES" if bars_val >= 4 else "NO"

#         # 4) Minimum bar dia check
#         if dia_val is not None:
#             min_dia_check = "YES" if dia_val >= 12 else "NO"
#     except Exception:
#         pass

#     columns[col_no] = {
#         "Column No": col_no,
#         "Guiding Load Case": guiding_load,
#         "Guiding Joint": guiding_joint,
#         "Req Steel Area (Sq.mm)": req_steel,
#         "Req Concrete Area (Sq.mm)": req_conc,
#         "Provided Steel Area (Sq.mm)": prov_steel,
#         "Steel Ratio (%)": ratio,
#         "No of Bars": no_bars,
#         "Bar Dia (mm)": bar_dia,
#         "Area 1 Bar (mm²)": round(bar_area, 2) if bar_area else "",
#         "Total Bar Area (mm²)": round(total_area, 2) if total_area else "",
#         "Area Check": area_check,
#         "Confinement Dia (mm)": conf_dia,
#         "Confinement Spacing (mm c/c)": conf_space,
#         "Confinement Length (mm)": conf_len,
#         "Tie Dia (mm)": tie_dia,
#         "Tie Spacing (mm c/c)": tie_space,
#         "Min Steel Ratio (>0.8)": min_ratio_check,
#         "Max Steel Ratio (<4)": max_ratio_check,
#         "Min 4 Bars (>=4)": min_bar_check,
#         "Min Bar Dia (>=12)": min_dia_check,
#     }

# # ---------- WRITE EXCEL ----------
# wb = openpyxl.Workbook()

# # Main sheet
# ws = wb.active
# ws.title = "Column_Design_Data"

# if not columns:
#     print("❌ No columns found!")
#     wb.save(OUTPUT_XLSX)
# else:
#     headers = list(next(iter(columns.values())).keys())
#     ws.append(headers)
#     for c in columns.values():
#         ws.append([c.get(h, "") for h in headers])

#     # Auto width
#     for col_idx, header in enumerate(headers, start=1):
#         ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(str(header)) + 4, 16), 40)

#     # Style header
#     header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
#     header_font = Font(color="FFFFFF", bold=True)
#     for cell in ws[1]:
#         cell.fill = header_fill
#         cell.font = header_font

#     # --- Color “NO” cells (fast, non-blocking) ---
#     red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
#     bold_red = Font(color="9C0006", bold=True)

#     # Columns we consider as checks
#     check_cols = [
#         "Area Check",
#         "Min Steel Ratio (>0.8)",
#         "Max Steel Ratio (<4)",
#         "Min 4 Bars (>=4)",
#         "Min Bar Dia (>=12)",
#     ]
#     # Map header name -> column index
#     header_to_index = {h: i for i, h in enumerate(headers)}

#     # Track failing rows (any NO in any check col)
#     fail_rows = []

#     for r in range(2, ws.max_row + 1):
#         row_failed = False
#         for col_name in check_cols:
#             idx = header_to_index.get(col_name)
#             if idx is None:
#                 continue
#             cell = ws.cell(row=r, column=idx + 1)
#             if str(cell.value).strip().upper() == "NO":
#                 cell.fill = red_fill
#                 cell.font = bold_red
#                 row_failed = True
#         if row_failed:
#             fail_rows.append(r)

#     # Add a summary row on the first sheet
#     total = len(columns)
#     n_fail = len(fail_rows)
#     n_pass = total - n_fail
#     ws.append([])
#     ws.append(["Summary", f"Total = {total}", f"Pass = {n_pass}", f"Fail = {n_fail}"])
#     for c in ws[ws.max_row]:
#         c.font = Font(bold=True)

#     # Failures sheet (only rows that had any NO)
#     ws_fail = wb.create_sheet("Failures_Only")
#     ws_fail.append(headers + ["Fail Flags (which checks failed)"])
#     for row_idx in fail_rows:
#         values = [ws.cell(row=row_idx, column=j + 1).value for j in range(len(headers))]
#         # Collect which checks failed for clarity
#         flags = []
#         for col_name in check_cols:
#             idx = header_to_index[col_name]
#             v = ws.cell(row=row_idx, column=idx + 1).value
#             if str(v).strip().upper() == "NO":
#                 flags.append(col_name)
#         ws_fail.append(values + [", ".join(flags)])

#     # Style Failures header
#     for cell in ws_fail[1]:
#         cell.fill = header_fill
#         cell.font = header_font

#     wb.save(OUTPUT_XLSX)

#     # One clean console line + non-blocking toast
#     print(f"✅ Extracted {total} columns → {OUTPUT_XLSX}  |  ✓ Pass: {n_pass}  ✖ Fail: {n_fail}")
#     notify_summary(total, n_pass, n_fail, OUTPUT_XLSX)


# import re
# import math
# from typing import Optional

# import openpyxl
# from openpyxl.styles import PatternFill, Font
# from openpyxl.utils import get_column_letter

# # ================================
# # CONFIG
# # ================================
# INPUT_FILE   = "MR. DINESH KHATRI.ANL"      # path to your .ANL file
# OUTPUT_XLSX  = "column_design_data_checked.xlsx"
# TOLERANCE_MM2 = 4                           # ± tolerance for area match
# # Summary notification: "popup" (centered Tk window), "toast" (win10toast), or "none"
# POPUP_MODE: str = "popup"
# # Seconds to auto-close popup; None = wait for user to click OK
# POPUP_SECONDS: Optional[int] = None

# # ================================
# # OPTIONAL NOTIFY HELPERS
# # ================================
# def _centered_popup(title: str, message: str, stay_seconds: Optional[int] = None) -> bool:
#     """
#     User-friendly centered popup window with heading, padding, and Close button.
#     """
#     try:
#         import tkinter as tk
#         import threading

#         def _show():
#             root = tk.Tk()
#             root.title(title)
#             root.resizable(False, False)
#             root.attributes("-topmost", True)

#             # Main frame
#             frm = tk.Frame(root, padx=25, pady=20, bg="#ffffff")
#             frm.pack(fill="both", expand=True)

#             # Title at top
#             lbl_title = tk.Label(
#                 frm,
#                 text="✅ Please Check Sheet",
#                 font=("Segoe UI", 14, "bold"),
#                 bg="#ffffff",
#                 fg="#2C3E50"
#             )
#             lbl_title.pack(pady=(0, 10))

#             # Main message
#             lbl_msg = tk.Label(
#                 frm,
#                 text=message + "\n\nWe have failure cases — fast and efficient review needed.",
#                 font=("Segoe UI", 11),
#                 bg="#ffffff",
#                 fg="#333333",
#                 justify="left",
#             )
#             lbl_msg.pack()

#             # Close button
#             btn = tk.Button(
#                 frm,
#                 text="Close",
#                 width=12,
#                 font=("Segoe UI", 10, "bold"),
#                 command=root.destroy,
#                 bg="#3498DB",
#                 fg="white",
#                 relief="raised"
#             )
#             btn.pack(pady=(15, 0))

#             # Center window
#             root.update_idletasks()
#             w = root.winfo_width()
#             h = root.winfo_height()
#             sw = root.winfo_screenwidth()
#             sh = root.winfo_screenheight()
#             x = (sw // 2) - (w // 2)
#             y = (sh // 2) - (h // 3)
#             root.geometry(f"{w}x{h}+{x}+{y}")

#             # Auto-close timer
#             if stay_seconds is not None:
#                 root.after(int(stay_seconds * 1000), root.destroy)

#             root.mainloop()

#         threading.Thread(target=_show, daemon=True).start()
#         return True

#     except Exception as e:
#         print("Popup failed:", e)
#         return False


# def notify_summary(total: int, n_pass: int, n_fail: int, path: str,
#                    mode: str = POPUP_MODE, popup_seconds: Optional[int] = POPUP_SECONDS) -> None:
#     """
#     Shows a summary using chosen mode.
#     Always prints a single clean line to console as well.
#     """
#     msg = f"Columns: {total}  ✓ Pass: {n_pass}  ✖ Fail: {n_fail}\nSaved: {path}"

#     shown = False
#     if mode.lower() == "toast":
#         try:
#             win10toast = __import__("win10toast")
#             toaster = win10toast.ToastNotifier()
#             # Non-blocking Windows toast
#             toaster.show_toast("Column Design Check", msg, duration=5, threaded=True)
#             shown = True
#         except Exception:
#             shown = False

#     elif mode.lower() == "popup":
#         shown = _centered_popup("Column Design Check", msg, stay_seconds=popup_seconds)

#     # Always print a one-liner fallback/info
#     print(f"INFO: Columns={total}, Pass={n_pass}, Fail={n_fail} → {path}"
#           + ("" if shown else "  (notification fallback: console only)"))

# # ================================
# # READ .ANL
# # ================================
# with open(INPUT_FILE, "r", encoding="utf-8", errors="ignore") as f:
#     text = f.read()

# # Split file into column blocks: the split keeps the column numbers in odd indices
# blocks = re.split(r"C\s*O\s*L\s*U\s*M\s*N\s+N\s*O\.\s*(\d+)", text)
# columns = {}

# for i in range(1, len(blocks), 2):
#     col_no = blocks[i].strip()
#     block  = blocks[i + 1]

#     def find(pattern, flags=0, group=1):
#         m = re.search(pattern, block, flags)
#         if not m:
#             return None
#         val = m.group(group)
#         return val.replace(",", "") if isinstance(val, str) else val

#     # Basic values
#     guiding_load = find(r"GUIDING\s*LOAD\s*CASE:\s*(\d+)")
#     guiding_joint = find(r"END\s*JOINT:\s*(\d+)")
#     req_steel = find(r"REQD\.\s*STEEL\s*AREA\s*[:=]?\s*([\d.]+)")
#     req_conc = find(r"REQD\.\s*CONCRETE\s*AREA\s*[:=]?\s*([\d.]+)")

#     # Provided steel & ratio (typical STAAD patterns)
#     prov_steel = find(r"MAIN\s*REINFORCEMENT\s*:\s*Provide.*?\((?:\s*[\d.]+%[, ]+\s*)?([\d.]+)",
#                       flags=re.DOTALL)
#     ratio = find(r"MAIN\s*REINFORCEMENT\s*:\s*Provide.*?\(\s*([\d.]+)%",
#                  flags=re.DOTALL)

#     # Bars
#     no_bars = find(r"Provide\s+(\d+)\s*-\s*\d+\s*dia")
#     bar_dia = find(r"Provide\s+\d+\s*-\s*(\d+)\s*dia")

#     # Confinement / ties
#     conf_dia = find(r"CONFINING\s*REINFORCEMENT\s*:\s*Provide\s*(\d+)")
#     conf_space = find(r"rectangular\s*ties\s*@\s*(\d+)")
#     conf_len = find(r"over\s+a\s+length\s*([\d.]+)\s*mm")
#     tie_dia = find(r"TIE\s*REINFORCEMENT\s*.*?Provide\s*(\d+)", flags=re.DOTALL)
#     tie_space = find(r"TIE\s*REINFORCEMENT\s*.*?@\s*(\d+)", flags=re.DOTALL)

#     # Calculated areas
#     bar_area = total_area = None
#     if bar_dia and no_bars:
#         try:
#             dia = float(bar_dia)
#             count = float(no_bars)
#             bar_area = math.pi * (dia ** 2) / 4.0
#             total_area = bar_area * count
#         except Exception:
#             pass

#     # Area check (± tolerance)
#     area_check = ""
#     if prov_steel and total_area is not None:
#         try:
#             prov_val = float(prov_steel)
#             area_check = "YES" if abs(prov_val - total_area) <= TOLERANCE_MM2 else "NO"
#         except ValueError:
#             pass

#     # Design checks
#     min_ratio_check = max_ratio_check = min_bar_check = min_dia_check = ""
#     try:
#         ratio_val = float(ratio) if ratio else None
#         bars_val = int(no_bars) if no_bars else None
#         dia_val = float(bar_dia) if bar_dia else None

#         # 1) Minimum steel ratio check (>0.8)
#         if ratio_val is not None:
#             min_ratio_check = "YES" if ratio_val > 0.8 else "NO"

#         # 2) Maximum steel ratio check (<4)
#         if ratio_val is not None:
#             max_ratio_check = "YES" if ratio_val < 4 else "NO"

#         # 3) Minimum 4 bars check
#         if bars_val is not None:
#             min_bar_check = "YES" if bars_val >= 4 else "NO"

#         # 4) Minimum bar dia check
#         if dia_val is not None:
#             min_dia_check = "YES" if dia_val >= 12 else "NO"
#     except Exception:
#         pass

#     columns[col_no] = {
#         "Column No": col_no,
#         "Guiding Load Case": guiding_load,
#         "Guiding Joint": guiding_joint,
#         "Req Steel Area (Sq.mm)": req_steel,
#         "Req Concrete Area (Sq.mm)": req_conc,
#         "Provided Steel Area (Sq.mm)": prov_steel,
#         "Steel Ratio (%)": ratio,
#         "No of Bars": no_bars,
#         "Bar Dia (mm)": bar_dia,
#         "Area 1 Bar (mm²)": round(bar_area, 2) if bar_area else "",
#         "Total Bar Area (mm²)": round(total_area, 2) if total_area else "",
#         "Area Check": area_check,
#         "Confinement Dia (mm)": conf_dia,
#         "Confinement Spacing (mm c/c)": conf_space,
#         "Confinement Length (mm)": conf_len,
#         "Tie Dia (mm)": tie_dia,
#         "Tie Spacing (mm c/c)": tie_space,
#         "Min Steel Ratio (>0.8)": min_ratio_check,
#         "Max Steel Ratio (<4)": max_ratio_check,
#         "Min 4 Bars (>=4)": min_bar_check,
#         "Min Bar Dia (>=12)": min_dia_check,
#     }

# # ================================
# # WRITE EXCEL
# # ================================
# wb = openpyxl.Workbook()

# # Main sheet
# ws = wb.active
# ws.title = "Column_Design_Data"

# if not columns:
#     print("❌ No columns found in file!")
#     wb.save(OUTPUT_XLSX)
# else:
#     headers = list(next(iter(columns.values())).keys())
#     ws.append(headers)
#     for c in columns.values():
#         ws.append([c.get(h, "") for h in headers])

#     # Auto width
#     for col_idx, header in enumerate(headers, start=1):
#         ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(str(header)) + 4, 16), 40)

#     # Style header
#     header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
#     header_font = Font(color="FFFFFF", bold=True)
#     for cell in ws[1]:
#         cell.fill = header_fill
#         cell.font = header_font

#     # === Color “NO” cells (fast, non-blocking) ===
#     red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
#     bold_red = Font(color="9C0006", bold=True)

#     # These columns are treated as checks
#     check_cols = [
#         "Area Check",
#         "Min Steel Ratio (>0.8)",
#         "Max Steel Ratio (<4)",
#         "Min 4 Bars (>=4)",
#         "Min Bar Dia (>=12)",
#     ]
#     # Map header name -> column index
#     header_to_index = {h: i for i, h in enumerate(headers)}

#     # Track failing rows (any NO in any check col)
#     fail_rows = []

#     for r in range(2, ws.max_row + 1):
#         row_failed = False
#         for col_name in check_cols:
#             idx = header_to_index.get(col_name)
#             if idx is None:
#                 continue
#             cell = ws.cell(row=r, column=idx + 1)
#             if str(cell.value).strip().upper() == "NO":
#                 cell.fill = red_fill
#                 cell.font = bold_red
#                 row_failed = True
#         if row_failed:
#             fail_rows.append(r)

#     # Summary row on main sheet
#     total = len(columns)
#     n_fail = len(fail_rows)
#     n_pass = total - n_fail
#     ws.append([])
#     ws.append(["Summary", f"Total = {total}", f"Pass = {n_pass}", f"Fail = {n_fail}"])
#     for c in ws[ws.max_row]:
#         c.font = Font(bold=True)

#     # Failures-only sheet
#     ws_fail = wb.create_sheet("Failures_Only")
#     ws_fail.append(headers + ["Fail Flags (which checks failed)"])
#     for row_idx in fail_rows:
#         values = [ws.cell(row=row_idx, column=j + 1).value for j in range(len(headers))]
#         flags = []
#         for col_name in check_cols:
#             idx = header_to_index[col_name]
#             v = ws.cell(row=row_idx, column=idx + 1).value
#             if str(v).strip().upper() == "NO":
#                 flags.append(col_name)
#         ws_fail.append(values + [", ".join(flags)])

#     # Style Failures header
#     for cell in ws_fail[1]:
#         cell.fill = header_fill
#         cell.font = header_font

#     # Save once
#     wb.save(OUTPUT_XLSX)

#     # One clean console line + popup/toast
#     print(f"✅ Extracted {total} columns → {OUTPUT_XLSX}  |  ✓ Pass: {n_pass}  ✖ Fail: {n_fail}")
#     notify_summary(total, n_pass, n_fail, OUTPUT_XLSX, mode=POPUP_MODE, popup_seconds=POPUP_SECONDS)


# import re
# import math
# from typing import Optional

# import openpyxl
# from openpyxl.styles import PatternFill, Font
# from openpyxl.utils import get_column_letter

# # ================================
# # CONFIG
# # ================================
# INPUT_FILE   = "MR. DINESH KHATRI.ANL"      # path to your .ANL file
# OUTPUT_XLSX  = "column_design_data_checked.xlsx"
# TOLERANCE_MM2 = 4                           # ± tolerance for area match

# POPUP_MODE: str = "popup"
# POPUP_SECONDS: Optional[int] = None         # keep popup open until close button

# # ================================
# # ALWAYS-WORKING POPUP
# # ================================
# def _centered_popup(title: str, message: str, stay_seconds: Optional[int] = None) -> bool:
#     """
#     GUARANTEED WORKING POPUP (Tkinter main-thread)
#     """
#     try:
#         import tkinter as tk

#         root = tk.Tk()
#         root.title(title)
#         root.attributes("-topmost", True)
#         root.resizable(False, False)
#         root.configure(bg="#ffffff")

#         # Frame
#         frm = tk.Frame(root, padx=25, pady=20, bg="#ffffff")
#         frm.pack(fill="both", expand=True)

#         # Title
#         lbl_title = tk.Label(
#             frm,
#             text="✅ Please Check Sheet",
#             font=("Segoe UI", 14, "bold"),
#             bg="#ffffff",
#             fg="#2C3E50"
#         )
#         lbl_title.pack(pady=(0, 10))

#         # Main message
#         lbl_msg = tk.Label(
#             frm,
#             text=message + "\n\nWe have failure cases — kindly review quickly.",
#             font=("Segoe UI", 11),
#             bg="#ffffff",
#             fg="#333333",
#             justify="left"
#         )
#         lbl_msg.pack()

#         # Close button
#         btn = tk.Button(
#             frm,
#             text="Close",
#             width=12,
#             font=("Segoe UI", 10, "bold"),
#             command=root.destroy,
#             bg="#3498DB",
#             fg="white"
#         )
#         btn.pack(pady=(15, 0))

#         # Centering
#         root.update_idletasks()
#         w = root.winfo_width()
#         h = root.winfo_height()
#         sw = root.winfo_screenwidth()
#         sh = root.winfo_screenheight()
#         x = (sw // 2) - (w // 2)
#         y = (sh // 2) - (h // 3)
#         root.geometry(f"{w}x{h}+{x}+{y}")

#         # Auto-close
#         if stay_seconds is not None:
#             root.after(int(stay_seconds * 1000), root.destroy)

#         root.mainloop()
#         return True

#     except Exception as e:
#         print("Popup error:", e)
#         return False


# # ================================
# # SUMMARY NOTIFICATION
# # ================================
# def notify_summary(total: int, n_pass: int, n_fail: int, path: str,
#                    mode: str = POPUP_MODE, popup_seconds: Optional[int] = POPUP_SECONDS) -> None:

#     msg = f"Columns: {total}  ✓ Pass: {n_pass}  ✖ Fail: {n_fail}\nSaved: {path}"

#     shown = False
#     if mode.lower() == "popup":
#         shown = _centered_popup("Column Design Check", msg, stay_seconds=popup_seconds)

#     print(f"INFO: Columns={total}, Pass={n_pass}, Fail={n_fail} → {path}"
#           + ("" if shown else "  (popup failed → console only)"))


# # ================================
# # READ .ANL
# # ================================
# with open(INPUT_FILE, "r", encoding="utf-8", errors="ignore") as f:
#     text = f.read()

# # Split into blocks
# blocks = re.split(r"C\s*O\s*L\s*U\s*M\s*N\s+N\s*O\.\s*(\d+)", text)
# columns = {}

# for i in range(1, len(blocks), 2):
#     col_no = blocks[i].strip()
#     block  = blocks[i + 1]

#     def find(pattern, flags=0, group=1):
#         m = re.search(pattern, block, flags)
#         if not m:
#             return None
#         val = m.group(group)
#         return val.replace(",", "") if isinstance(val, str) else val

#     guiding_load = find(r"GUIDING\s*LOAD\s*CASE:\s*(\d+)")
#     guiding_joint = find(r"END\s*JOINT:\s*(\d+)")
#     req_steel = find(r"REQD\.\s*STEEL\s*AREA\s*[:=]?\s*([\d.]+)")
#     req_conc = find(r"REQD\.\s*CONCRETE\s*AREA\s*[:=]?\s*([\d.]+)")

#     prov_steel = find(r"MAIN\s*REINFORCEMENT\s*:\s*Provide.*?\((?:\s*[\d.]+%[, ]+\s*)?([\d.]+)",
#                       flags=re.DOTALL)
#     ratio = find(r"MAIN\s*REINFORCEMENT\s*:\s*Provide.*?\(\s*([\d.]+)%",
#                  flags=re.DOTALL)

#     no_bars = find(r"Provide\s+(\d+)\s*-\s*\d+\s*dia")
#     bar_dia = find(r"Provide\s+\d+\s*-\s*(\d+)\s*dia")

#     conf_dia = find(r"CONFINING\s*REINFORCEMENT\s*:\s*Provide\s*(\d+)")
#     conf_space = find(r"rectangular\s*ties\s*@\s*(\d+)")
#     conf_len = find(r"over\s+a\s+length\s*([\d.]+)\s*mm")
#     tie_dia = find(r"TIE\s*REINFORCEMENT\s*.*?Provide\s*(\d+)", flags=re.DOTALL)
#     tie_space = find(r"TIE\s*REINFORCEMENT\s*.*?@\s*(\d+)", flags=re.DOTALL)

#     bar_area = total_area = None
#     if bar_dia and no_bars:
#         try:
#             dia = float(bar_dia)
#             count = float(no_bars)
#             bar_area = math.pi * (dia ** 2) / 4.0
#             total_area = bar_area * count
#         except:
#             pass

#     area_check = ""
#     if prov_steel and total_area is not None:
#         try:
#             prov_val = float(prov_steel)
#             area_check = "YES" if abs(prov_val - total_area) <= TOLERANCE_MM2 else "NO"
#         except:
#             pass

#     min_ratio_check = max_ratio_check = min_bar_check = min_dia_check = ""
#     try:
#         ratio_val = float(ratio) if ratio else None
#         bars_val = int(no_bars) if no_bars else None
#         dia_val = float(bar_dia) if bar_dia else None

#         if ratio_val is not None:
#             min_ratio_check = "YES" if ratio_val > 0.8 else "NO"
#             max_ratio_check = "YES" if ratio_val < 4 else "NO"

#         if bars_val is not None:
#             min_bar_check = "YES" if bars_val >= 4 else "NO"

#         if dia_val is not None:
#             min_dia_check = "YES" if dia_val >= 12 else "NO"
#     except:
#         pass

#     columns[col_no] = {
#         "Column No": col_no,
#         "Guiding Load Case": guiding_load,
#         "Guiding Joint": guiding_joint,
#         "Req Steel Area (Sq.mm)": req_steel,
#         "Req Concrete Area (Sq.mm)": req_conc,
#         "Provided Steel Area (Sq.mm)": prov_steel,
#         "Steel Ratio (%)": ratio,
#         "No of Bars": no_bars,
#         "Bar Dia (mm)": bar_dia,
#         "Area 1 Bar (mm²)": round(bar_area, 2) if bar_area else "",
#         "Total Bar Area (mm²)": round(total_area, 2) if total_area else "",
#         "Area Check": area_check,
#         "Confinement Dia (mm)": conf_dia,
#         "Confinement Spacing (mm c/c)": conf_space,
#         "Confinement Length (mm)": conf_len,
#         "Tie Dia (mm)": tie_dia,
#         "Tie Spacing (mm c/c)": tie_space,
#         "Min Steel Ratio (>0.8)": min_ratio_check,
#         "Max Steel Ratio (<4)": max_ratio_check,
#         "Min 4 Bars (>=4)": min_bar_check,
#         "Min Bar Dia (>=12)": min_dia_check,
#     }

# # ================================
# # WRITE EXCEL
# # ================================
# wb = openpyxl.Workbook()
# ws = wb.active
# ws.title = "Column_Design_Data"

# if not columns:
#     print("❌ No columns found!")
#     wb.save(OUTPUT_XLSX)
# else:
#     headers = list(next(iter(columns.values())).keys())
#     ws.append(headers)

#     for c in columns.values():
#         ws.append([c.get(h, "") for h in headers])

#     for col_idx, header in enumerate(headers, start=1):
#         ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(str(header)) + 4, 16), 40)

#     header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
#     header_font = Font(color="FFFFFF", bold=True)

#     for cell in ws[1]:
#         cell.fill = header_fill
#         cell.font = header_font

#     red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
#     bold_red = Font(color="9C0006", bold=True)

#     check_cols = [
#         "Area Check",
#         "Min Steel Ratio (>0.8)",
#         "Max Steel Ratio (<4)",
#         "Min 4 Bars (>=4)",
#         "Min Bar Dia (>=12)",
#     ]
#     header_to_index = {h: i for i, h in enumerate(headers)}

#     fail_rows = []

#     for r in range(2, ws.max_row + 1):
#         row_failed = False
#         for col_name in check_cols:
#             idx = header_to_index.get(col_name)
#             if idx is None:
#                 continue
#             cell = ws.cell(row=r, column=idx + 1)
#             if str(cell.value).strip().upper() == "NO":
#                 cell.fill = red_fill
#                 cell.font = bold_red
#                 row_failed = True
#         if row_failed:
#             fail_rows.append(r)

#     total = len(columns)
#     n_fail = len(fail_rows)
#     n_pass = total - n_fail

#     ws.append([])
#     ws.append(["Summary", f"Total = {total}", f"Pass = {n_pass}", f"Fail = {n_fail}"])
#     for c in ws[ws.max_row]:
#         c.font = Font(bold=True)

#     ws_fail = wb.create_sheet("Failures_Only")
#     ws_fail.append(headers + ["Fail Flags"])

#     for row_idx in fail_rows:
#         values = [ws.cell(row=row_idx, column=j + 1).value for j in range(len(headers))]
#         flags = []
#         for col_name in check_cols:
#             idx = header_to_index[col_name]
#             v = ws.cell(row=row_idx, column=idx + 1).value
#             if str(v).strip().upper() == "NO":
#                 flags.append(col_name)
#         ws_fail.append(values + [", ".join(flags)])

#     for cell in ws_fail[1]:
#         cell.fill = header_fill
#         cell.font = header_font

#     wb.save(OUTPUT_XLSX)

#     print(f"✅ Extracted {total} columns → {OUTPUT_XLSX}  |  ✓ Pass: {n_pass}  ✖ Fail: {n_fail}")

#     notify_summary(total, n_pass, n_fail, OUTPUT_XLSX, mode=POPUP_MODE, popup_seconds=POPUP_SECONDS)

