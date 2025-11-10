

# import re
# import math
# from typing import Optional

# import openpyxl
# from openpyxl.styles import PatternFill, Font
# from openpyxl.utils import get_column_letter

# # ================================
# # CONFIG
# # ================================
# INPUT_FILE   = "MR. DINESH KHATRI.ANL"
# OUTPUT_XLSX  = "column_design_data_checked.xlsx"
# TOLERANCE_MM2 = 4

# POPUP_MODE: str = "popup"
# POPUP_SECONDS: Optional[int] = None    # Keep popup open

# # ================================
# # ALWAYS-WORKING POPUP
# # ================================
# def _centered_popup(title: str, message: str, stay_seconds: Optional[int] = None) -> bool:
#     try:
#         import tkinter as tk

#         root = tk.Tk()
#         root.title(title)
#         root.attributes("-topmost", True)
#         root.resizable(False, False)
#         root.configure(bg="#ffffff")

#         frm = tk.Frame(root, padx=25, pady=20, bg="#ffffff")
#         frm.pack(fill="both", expand=True)

#         lbl_title = tk.Label(
#             frm, text="âœ… Please Check Sheet",
#             font=("Segoe UI", 14, "bold"),
#             bg="#ffffff", fg="#2C3E50"
#         )
#         lbl_title.pack(pady=(0, 10))

#         lbl_msg = tk.Label(
#             frm,
#             text=message + "\n\n kindly review .",
#             font=("Segoe UI", 11),
#             bg="#ffffff", fg="#333333",
#             justify="left"
#         )
#         lbl_msg.pack()

#         btn = tk.Button(
#             frm, text="Close", width=12,
#             font=("Segoe UI", 10, "bold"),
#             command=root.destroy,
#             bg="#3498DB", fg="white"
#         )
#         btn.pack(pady=(15, 0))

#         root.update_idletasks()
#         w = root.winfo_width()
#         h = root.winfo_height()
#         sw = root.winfo_screenwidth()
#         sh = root.winfo_screenheight()
#         x = (sw // 2) - (w // 2)
#         y = (sh // 3) - (h // 3)
#         root.geometry(f"{w}x{h}+{x}+{y}")

#         if stay_seconds is not None:
#             root.after(int(stay_seconds * 1000), root.destroy)

#         root.mainloop()
#         return True

#     except Exception as e:
#         print("Popup error:", e)
#         return False


# # ================================
# # SUMMARY NOTIFICATION
# # ================================
# def notify_summary(total: int, n_pass: int, n_fail: int, path: str,
#                    mode: str = POPUP_MODE, popup_seconds: Optional[int] = POPUP_SECONDS) -> None:

#     msg = f"Columns: {total}  âœ“ Pass: {n_pass}  âœ– Fail: {n_fail}\nSaved: {path}"

#     shown = False
#     if mode.lower() == "popup":
#         shown = _centered_popup("Column Design Check", msg, stay_seconds=popup_seconds)

#     print(f"INFO: Columns={total}, Pass={n_pass}, Fail={n_fail} â†’ {path}"
#           + ("" if shown else "  (popup failed â†’ console only)"))


# # ================================
# # READ .ANL
# # ================================
# with open(INPUT_FILE, "r", encoding="utf-8", errors="ignore") as f:
#     text = f.read()

# # Split into blocks
# blocks = re.split(r"C\s*O\s*L\s*U\s*M\s*N\s+N\s*O\.\s*(\d+)", text)
# columns = {}

# for i in range(1, len(blocks), 2):
#     col_no = blocks[i].strip()
#     block  = blocks[i + 1]

#     def find(pattern, flags=0, group=1):
#         m = re.search(pattern, block, flags)
#         if not m:
#             return None
#         val = m.group(group)
#         return val.replace(",", "") if isinstance(val, str) else val

#     guiding_load = find(r"GUIDING\s*LOAD\s*CASE:\s*(\d+)")
#     guiding_joint = find(r"END\s*JOINT:\s*(\d+)")
#     req_steel = find(r"REQD\.\s*STEEL\s*AREA\s*[:=]?\s*([\d.]+)")
#     req_conc = find(r"REQD\.\s*CONCRETE\s*AREA\s*[:=]?\s*([\d.]+)")

#     # Provided steel
#     prov_steel = find(
#         r"MAIN\s*REINFORCEMENT\s*:\s*Provide.*?\((?:\s*[\d.]+%[, ]+\s*)?([\d.]+)",
#         flags=re.DOTALL
#     )

#     # Bars
#     no_bars = find(r"Provide\s+(\d+)\s*-\s*\d+\s*dia")
#     bar_dia = find(r"Provide\s+\d+\s*-\s*(\d+)\s*dia")

#     # Confinement / ties
#     conf_dia = find(r"CONFINING\s*REINFORCEMENT\s*:\s*Provide\s*(\d+)")
#     conf_space = find(r"rectangular\s*ties\s*@\s*(\d+)")
#     conf_len = find(r"over\s+a\s+length\s*([\d.]+)\s*mm")
#     tie_dia = find(r"TIE\s*REINFORCEMENT\s*.*?Provide\s*(\d+)", flags=re.DOTALL)
#     tie_space = find(r"TIE\s*REINFORCEMENT\s*.*?@\s*(\d+)", flags=re.DOTALL)

#     # Area calculations
#     bar_area = total_area = None
#     if bar_dia and no_bars:
#         try:
#             dia = float(bar_dia)
#             count = float(no_bars)
#             bar_area = math.pi * (dia ** 2) / 4.0
#             total_area = bar_area * count
#         except:
#             pass

#     # Area check
#     area_check = ""
#     if prov_steel and total_area is not None:
#         try:
#             prov_val = float(prov_steel)
#             area_check = "YES" if abs(prov_val - total_area) <= TOLERANCE_MM2 else "NO"
#         except:
#             pass

#     # âœ… NEW â€” CALCULATED STEEL RATIO %
#     calc_ratio = ""
#     try:
#         if prov_steel and req_conc:
#             prov_val = float(prov_steel)
#             conc_val = float(req_conc)
#             if conc_val > 0:
#                 calc_ratio = round((prov_val / conc_val) * 100, 3)
#     except:
#         calc_ratio = ""

#     # âœ… CHECKS BASED ON CALCULATED RATIO
#     min_ratio_check = max_ratio_check = ""
#     try:
#         if calc_ratio != "":
#             min_ratio_check = "YES" if calc_ratio > 0.8 else "NO"
#             max_ratio_check = "YES" if calc_ratio < 4 else "NO"
#     except:
#         pass

#     # Bars & Dia checks
#     min_bar_check = min_dia_check = ""
#     try:
#         bars_val = int(no_bars) if no_bars else None
#         dia_val = float(bar_dia) if bar_dia else None

#         if bars_val is not None:
#             min_bar_check = "YES" if bars_val >= 4 else "NO"
#         if dia_val is not None:
#             min_dia_check = "YES" if dia_val >= 12 else "NO"
#     except:
#         pass

#     # âœ… Store Data
#     columns[col_no] = {
#         "Column No": col_no,
#         "Guiding Load Case": guiding_load,
#         "Guiding Joint": guiding_joint,
#         "Req Steel Area (Sq.mm)": req_steel,
#         "Req Concrete Area (Sq.mm)": req_conc,
#         "Provided Steel Area (Sq.mm)": prov_steel,
#         "Steel Ratio (%)": calc_ratio,       # âœ… UPDATED
#         "No of Bars": no_bars,
#         "Bar Dia (mm)": bar_dia,
#         "Area 1 Bar (mmÂ²)": round(bar_area, 2) if bar_area else "",
#         "Total Bar Area (mmÂ²)": round(total_area, 2) if total_area else "",
#         "Area Check": area_check,
#         "Confinement Dia (mm)": conf_dia,
#         "Confinement Spacing (mm c/c)": conf_space,
#         "Confinement Length (mm)": conf_len,
#         "Tie Dia (mm)": tie_dia,
#         "Tie Spacing (mm c/c)": tie_space,
#         "Min Steel Ratio (>0.8)": min_ratio_check,
#         "Max Steel Ratio (<4)": max_ratio_check,
#         "Min 4 Bars (>=4)": min_bar_check,
#         "Min Bar Dia (>=12)": min_dia_check,
#     }

# # ================================
# # WRITE EXCEL
# # ================================
# wb = openpyxl.Workbook()
# ws = wb.active
# ws.title = "Column_Design_Data"

# if not columns:
#     print("âŒ No columns found!")
#     wb.save(OUTPUT_XLSX)
# else:
#     headers = list(next(iter(columns.values())).keys())
#     ws.append(headers)

#     for c in columns.values():
#         ws.append([c.get(h, "") for h in headers])

#     # Auto column width
#     for col_idx, header in enumerate(headers, start=1):
#         ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(str(header)) + 4, 16), 40)

#     # Style header
#     header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
#     header_font = Font(color="FFFFFF", bold=True)
#     for cell in ws[1]:
#         cell.fill = header_fill
#         cell.font = header_font

#     # Fail highlight color
#     red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
#     bold_red = Font(color="9C0006", bold=True)

#     check_cols = [
#         "Area Check",
#         "Min Steel Ratio (>0.8)",
#         "Max Steel Ratio (<4)",
#         "Min 4 Bars (>=4)",
#         "Min Bar Dia (>=12)",
#     ]

#     header_to_index = {h: i for i, h in enumerate(headers)}

#     fail_rows = []

#     for r in range(2, ws.max_row + 1):
#         row_failed = False
#         for col_name in check_cols:
#             idx = header_to_index.get(col_name)
#             if idx is None:
#                 continue
#             cell = ws.cell(row=r, column=idx + 1)
#             if str(cell.value).strip().upper() == "NO":
#                 cell.fill = red_fill
#                 cell.font = bold_red
#                 row_failed = True
#         if row_failed:
#             fail_rows.append(r)

#     total = len(columns)
#     n_fail = len(fail_rows)
#     n_pass = total - n_fail

#     ws.append([])
#     ws.append(["Summary", f"Total = {total}", f"Pass = {n_pass}", f"Fail = {n_fail}"])
#     for c in ws[ws.max_row]:
#         c.font = Font(bold=True)

#     # Failures only sheet
#     ws_fail = wb.create_sheet("Failures_Only")
#     ws_fail.append(headers + ["Fail Flags"])

#     for row_idx in fail_rows:
#         values = [ws.cell(row=row_idx, column=j + 1).value for j in range(len(headers))]
#         flags = []

#         for col_name in check_cols:
#             idx = header_to_index[col_name]
#             v = ws.cell(row=row_idx, column=idx + 1).value
#             if str(v).strip().upper() == "NO":
#                 flags.append(col_name)

#         ws_fail.append(values + [", ".join(flags)])

#     for cell in ws_fail[1]:
#         cell.fill = header_fill
#         cell.font = header_font

#     wb.save(OUTPUT_XLSX)

#     print(f"âœ… Extracted {total} columns â†’ {OUTPUT_XLSX}  |  âœ“ Pass: {n_pass}  âœ– Fail: {n_fail}")

#     notify_summary(total, n_pass, n_fail, OUTPUT_XLSX, mode=POPUP_MODE, popup_seconds=POPUP_SECONDS)


# import re
# import math
# from typing import Optional
# import openpyxl
# from openpyxl.styles import PatternFill, Font
# from openpyxl.utils import get_column_letter
# import pythoncom
# from win32com.client import Dispatch, VARIANT

# # ================================
# # CONFIG
# # ================================
# INPUT_FILE   = "MR. DINESH KHATRI.ANL"
# OUTPUT_XLSX  = "column_design_data_checked.xlsx"
# TOLERANCE_MM2 = 4

# POPUP_MODE: str = "popup"
# POPUP_SECONDS: Optional[int] = None

# # ================================
# # POPUP FUNCTION
# # ================================
# def _centered_popup(title: str, message: str, stay_seconds: Optional[int] = None) -> bool:
#     try:
#         import tkinter as tk
#         root = tk.Tk()
#         root.title(title)
#         root.attributes("-topmost", True)
#         root.resizable(False, False)
#         root.configure(bg="#ffffff")

#         frm = tk.Frame(root, padx=25, pady=20, bg="#ffffff")
#         frm.pack(fill="both", expand=True)

#         lbl_title = tk.Label(frm, text="âœ… Please Check Sheet",
#                              font=("Segoe UI", 14, "bold"),
#                              bg="#ffffff", fg="#2C3E50")
#         lbl_title.pack(pady=(0, 10))

#         lbl_msg = tk.Label(frm,
#                            text=message + "\n\nkindly review.",
#                            font=("Segoe UI", 11),
#                            bg="#ffffff", fg="#333333", justify="left")
#         lbl_msg.pack()

#         btn = tk.Button(frm, text="Close", width=12,
#                         font=("Segoe UI", 10, "bold"),
#                         command=root.destroy,
#                         bg="#3498DB", fg="white")
#         btn.pack(pady=(15, 0))

#         root.update_idletasks()
#         w = root.winfo_width()
#         h = root.winfo_height()
#         sw = root.winfo_screenwidth()
#         sh = root.winfo_screenheight()
#         x = (sw // 2) - (w // 2)
#         y = (sh // 3) - (h // 3)
#         root.geometry(f"{w}x{h}+{x}+{y}")

#         if stay_seconds is not None:
#             root.after(int(stay_seconds * 1000), root.destroy)

#         root.mainloop()
#         return True
#     except Exception as e:
#         print("Popup error:", e)
#         return False


# def notify_summary(total: int, n_pass: int, n_fail: int, path: str,
#                    mode: str = POPUP_MODE, popup_seconds: Optional[int] = POPUP_SECONDS) -> None:
#     msg = f"Columns: {total}  âœ“ Pass: {n_pass}  âœ– Fail: {n_fail}\nSaved: {path}"
#     shown = False
#     if mode.lower() == "popup":
#         shown = _centered_popup("Column Design Check", msg, stay_seconds=popup_seconds)
#     print(f"INFO: Columns={total}, Pass={n_pass}, Fail={n_fail} â†’ {path}"
#           + ("" if shown else "  (popup failed â†’ console only)"))

# # ================================
# # READ .ANL FILE
# # ================================
# with open(INPUT_FILE, "r", encoding="utf-8", errors="ignore") as f:
#     text = f.read()

# blocks = re.split(r"C\s*O\s*L\s*U\s*M\s*N\s+N\s*O\.\s*(\d+)", text)
# columns = {}

# for i in range(1, len(blocks), 2):
#     col_no = blocks[i].strip()
#     block  = blocks[i + 1]

#     def find(pattern, flags=0, group=1):
#         m = re.search(pattern, block, flags)
#         if not m:
#             return None
#         val = m.group(group)
#         return val.replace(",", "") if isinstance(val, str) else val

#     guiding_load = find(r"GUIDING\s*LOAD\s*CASE:\s*(\d+)")
#     guiding_joint = find(r"END\s*JOINT:\s*(\d+)")
#     req_steel = find(r"REQD\.\s*STEEL\s*AREA\s*[:=]?\s*([\d.]+)")
#     req_conc = find(r"REQD\.\s*CONCRETE\s*AREA\s*[:=]?\s*([\d.]+)")

#     prov_steel = find(r"MAIN\s*REINFORCEMENT\s*:\s*Provide.*?\((?:\s*[\d.]+%[, ]+\s*)?([\d.]+)", flags=re.DOTALL)
#     no_bars = find(r"Provide\s+(\d+)\s*-\s*\d+\s*dia")
#     bar_dia = find(r"Provide\s+\d+\s*-\s*(\d+)\s*dia")

#     conf_dia = find(r"CONFINING\s*REINFORCEMENT\s*:\s*Provide\s*(\d+)")
#     conf_space = find(r"rectangular\s*ties\s*@\s*(\d+)")
#     conf_len = find(r"over\s+a\s+length\s*([\d.]+)\s*mm")
#     tie_dia = find(r"TIE\s*REINFORCEMENT\s*.*?Provide\s*(\d+)", flags=re.DOTALL)
#     tie_space = find(r"TIE\s*REINFORCEMENT\s*.*?@\s*(\d+)", flags=re.DOTALL)

#     bar_area = total_area = None
#     if bar_dia and no_bars:
#         try:
#             dia = float(bar_dia)
#             count = float(no_bars)
#             bar_area = math.pi * (dia ** 2) / 4.0
#             total_area = bar_area * count
#         except:
#             pass

#     area_check = ""
#     if prov_steel and total_area is not None:
#         try:
#             prov_val = float(prov_steel)
#             area_check = "YES" if abs(prov_val - total_area) <= TOLERANCE_MM2 else "NO"
#         except:
#             pass

#     calc_ratio = ""
#     try:
#         if prov_steel and req_conc:
#             prov_val = float(prov_steel)
#             conc_val = float(req_conc)
#             if conc_val > 0:
#                 calc_ratio = round((prov_val / conc_val) * 100, 3)
#     except:
#         calc_ratio = ""

#     min_ratio_check = max_ratio_check = ""
#     try:
#         if calc_ratio != "":
#             min_ratio_check = "YES" if calc_ratio > 0.8 else "NO"
#             max_ratio_check = "YES" if calc_ratio < 4 else "NO"
#     except:
#         pass

#     min_bar_check = min_dia_check = ""
#     try:
#         bars_val = int(no_bars) if no_bars else None
#         dia_val = float(bar_dia) if bar_dia else None
#         if bars_val is not None:
#             min_bar_check = "YES" if bars_val >= 4 else "NO"
#         if dia_val is not None:
#             min_dia_check = "YES" if dia_val >= 12 else "NO"
#     except:
#         pass

#     columns[col_no] = {
#         "Column No": col_no,
#         "Guiding Load Case": guiding_load,
#         "Guiding Joint": guiding_joint,
#         "Req Steel Area (Sq.mm)": req_steel,
#         "Req Concrete Area (Sq.mm)": req_conc,
#         "Provided Steel Area (Sq.mm)": prov_steel,
#         "Steel Ratio (%)": calc_ratio,
#         "No of Bars": no_bars,
#         "Bar Dia (mm)": bar_dia,
#         "Area 1 Bar (mmÂ²)": round(bar_area, 2) if bar_area else "",
#         "Total Bar Area (mmÂ²)": round(total_area, 2) if total_area else "",
#         "Area Check": area_check,
#         "Confinement Dia (mm)": conf_dia,
#         "Confinement Spacing (mm c/c)": conf_space,
#         "Confinement Length (mm)": conf_len,
#         "Tie Dia (mm)": tie_dia,
#         "Tie Spacing (mm c/c)": tie_space,
#         "Min Steel Ratio (>0.8)": min_ratio_check,
#         "Max Steel Ratio (<4)": max_ratio_check,
#         "Min 4 Bars (>=4)": min_bar_check,
#         "Min Bar Dia (>=12)": min_dia_check,
#     }

# # ============================================================
# # ADD COLUMN DEPTH & WIDTH FROM STAAD
# # ============================================================
# def dispid(obj, name):
#     try:
#         return obj._oleobj_.GetIDsOfNames(name)
#     except Exception:
#         return None

# def get_prismatic_section_mm(member_no, propUI, prop, UNIT_TO_MM=25.4):
#     sig = (
#         (pythoncom.VT_I4, 0),
#         (pythoncom.VT_VARIANT, 1), (pythoncom.VT_VARIANT, 1),
#         (pythoncom.VT_VARIANT, 1), (pythoncom.VT_VARIANT, 1), (pythoncom.VT_VARIANT, 1),
#         (pythoncom.VT_VARIANT, 1), (pythoncom.VT_VARIANT, 1), (pythoncom.VT_VARIANT, 1),
#     )
#     def _try_iface(iface):
#         if iface is None:
#             return False, None, None
#         did = dispid(iface, "GetBeamProperty")
#         if not did:
#             return False, None, None
#         W = VARIANT(pythoncom.VT_BYREF | pythoncom.VT_R8, 0.0)
#         D = VARIANT(pythoncom.VT_BYREF | pythoncom.VT_R8, 0.0)
#         Ax = Ay = Az = Ix = Iy = Iz = VARIANT(pythoncom.VT_BYREF | pythoncom.VT_R8, 0.0)
#         try:
#             ret = iface._oleobj_.InvokeTypes(
#                 did, 0, pythoncom.DISPATCH_METHOD, (pythoncom.VT_I4, 0),
#                 sig, int(member_no), W, D, Ax, Ay, Az, Ix, Iy, Iz
#             )
#             ok = int(ret[0] if isinstance(ret, tuple) else ret)
#             if ok == 1:
#                 return True, float(W.value)*UNIT_TO_MM, float(D.value)*UNIT_TO_MM
#         except Exception:
#             pass
#         return False, None, None
#     for iface in (propUI, prop):
#         ok, w, d = _try_iface(iface)
#         if ok:
#             return True, w, d
#     return False, None, None

# print("\nðŸ”— Connecting to STAAD to fetch section sizes...")
# pythoncom.CoInitialize()
# os_app = Dispatch("StaadPro.OpenSTAAD")
# propUI = getattr(os_app, "PropertyUI", None)
# prop   = getattr(os_app, "Property",   None)

# for col_no, data in columns.items():
#     member_id = int(col_no)
#     ok, w_mm_localZ, d_mm_localY = get_prismatic_section_mm(member_id, propUI, prop)
#     if not ok:
#         w_mm_localZ, d_mm_localY = 0.0, 0.0
#     data["Column Width (mm)"] = round(w_mm_localZ)
#     data["Column Depth (mm)"] = round(d_mm_localY)

# print("âœ… Added Column Depth (mm) and Column Width (mm).")

# # ============================================================
# # WRITE TO EXCEL
# # ============================================================
# wb = openpyxl.Workbook()
# ws = wb.active
# ws.title = "Column_Design_Data"

# if not columns:
#     print("âŒ No columns found!")
#     wb.save(OUTPUT_XLSX)
# else:
#     headers = list(next(iter(columns.values())).keys())
#     ws.append(headers)
#     for c in columns.values():
#         ws.append([c.get(h, "") for h in headers])

#     for col_idx, header in enumerate(headers, start=1):
#         ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(str(header)) + 4, 16), 40)

#     header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
#     header_font = Font(color="FFFFFF", bold=True)
#     for cell in ws[1]:
#         cell.fill = header_fill
#         cell.font = header_font

#     red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
#     bold_red = Font(color="9C0006", bold=True)

#     check_cols = [
#         "Area Check",
#         "Min Steel Ratio (>0.8)",
#         "Max Steel Ratio (<4)",
#         "Min 4 Bars (>=4)",
#         "Min Bar Dia (>=12)",
#     ]
#     header_to_index = {h: i for i, h in enumerate(headers)}
#     fail_rows = []
#     for r in range(2, ws.max_row + 1):
#         row_failed = False
#         for col_name in check_cols:
#             idx = header_to_index.get(col_name)
#             if idx is None:
#                 continue
#             cell = ws.cell(row=r, column=idx + 1)
#             if str(cell.value).strip().upper() == "NO":
#                 cell.fill = red_fill
#                 cell.font = bold_red
#                 row_failed = True
#         if row_failed:
#             fail_rows.append(r)

#     total = len(columns)
#     n_fail = len(fail_rows)
#     n_pass = total - n_fail

#     ws.append([])
#     ws.append(["Summary", f"Total = {total}", f"Pass = {n_pass}", f"Fail = {n_fail}"])
#     for c in ws[ws.max_row]:
#         c.font = Font(bold=True)

#     ws_fail = wb.create_sheet("Failures_Only")
#     ws_fail.append(headers + ["Fail Flags"])
#     for row_idx in fail_rows:
#         values = [ws.cell(row=row_idx, column=j + 1).value for j in range(len(headers))]
#         flags = []
#         for col_name in check_cols:
#             idx = header_to_index[col_name]
#             v = ws.cell(row=row_idx, column=idx + 1).value
#             if str(v).strip().upper() == "NO":
#                 flags.append(col_name)
#         ws_fail.append(values + [", ".join(flags)])
#     for cell in ws_fail[1]:
#         cell.fill = header_fill
#         cell.font = header_font

#     wb.save(OUTPUT_XLSX)
#     print(f"âœ… Extracted {total} columns â†’ {OUTPUT_XLSX}  |  âœ“ Pass: {n_pass}  âœ– Fail: {n_fail}")
#     notify_summary(total, n_pass, n_fail, OUTPUT_XLSX, mode=POPUP_MODE, popup_seconds=POPUP_SECONDS)







# import re
# import math
# from typing import Optional
# import openpyxl
# from openpyxl.styles import PatternFill, Font
# from openpyxl.utils import get_column_letter
# import pythoncom
# from win32com.client import Dispatch, VARIANT

# # ================================
# # CONFIG
# # ================================
# INPUT_FILE   = "MR. DINESH KHATRI.ANL"
# OUTPUT_XLSX  = "column_design_data_checked.xlsx"
# TOLERANCE_MM2 = 4

# POPUP_MODE: str = "popup"
# POPUP_SECONDS: Optional[int] = None

# # cover & stirrup for deductions (two faces each)
# COVER = 40            # mm
# STIRRUP_DIA = 8       # mm
# DEDUCT = 2*COVER + 2*STIRRUP_DIA  # 80 + 16 = 96

# MIN_CLEAR = 30.0      # mm
# MAX_SPACING = 300.0   # mm

# # ================================
# # POPUP FUNCTION
# # ================================
# def _centered_popup(title: str, message: str, stay_seconds: Optional[int] = None) -> bool:
#     try:
#         import tkinter as tk
#         root = tk.Tk()
#         root.title(title)
#         root.attributes("-topmost", True)
#         root.resizable(False, False)
#         root.configure(bg="#ffffff")

#         frm = tk.Frame(root, padx=25, pady=20, bg="#ffffff")
#         frm.pack(fill="both", expand=True)

#         lbl_title = tk.Label(frm, text="âœ… Please Check Sheet",
#                              font=("Segoe UI", 14, "bold"),
#                              bg="#ffffff", fg="#2C3E50")
#         lbl_title.pack(pady=(0, 10))

#         lbl_msg = tk.Label(frm,
#                            text=message + "\n\nkindly review.",
#                            font=("Segoe UI", 11),
#                            bg="#ffffff", fg="#333333", justify="left")
#         lbl_msg.pack()

#         btn = tk.Button(frm, text="Close", width=12,
#                         font=("Segoe UI", 10, "bold"),
#                         command=root.destroy,
#                         bg="#3498DB", fg="white")
#         btn.pack(pady=(15, 0))

#         root.update_idletasks()
#         w = root.winfo_width()
#         h = root.winfo_height()
#         sw = root.winfo_screenwidth()
#         sh = root.winfo_screenheight()
#         x = (sw // 2) - (w // 2)
#         y = (sh // 3) - (h // 3)
#         root.geometry(f"{w}x{h}+{x}+{y}")

#         if stay_seconds is not None:
#             root.after(int(stay_seconds * 1000), root.destroy)

#         root.mainloop()
#         return True
#     except Exception as e:
#         print("Popup error:", e)
#         return False


# def notify_summary(total: int, n_pass: int, n_fail: int, path: str,
#                    mode: str = POPUP_MODE, popup_seconds: Optional[int] = POPUP_SECONDS) -> None:
#     msg = f"Columns: {total}  âœ“ Pass: {n_pass}  âœ– Fail: {n_fail}\nSaved: {path}"
#     shown = False
#     if mode.lower() == "popup":
#         shown = _centered_popup("Column Design Check", msg, stay_seconds=popup_seconds)
#     print(f"INFO: Columns={total}, Pass={n_pass}, Fail={n_fail} â†’ {path}"
#           + ("" if shown else "  (popup failed â†’ console only)"))

# # ================================
# # READ .ANL FILE
# # ================================
# with open(INPUT_FILE, "r", encoding="utf-8", errors="ignore") as f:
#     text = f.read()

# blocks = re.split(r"C\s*O\s*L\s*U\s*M\s*N\s+N\s*O\.\s*(\d+)", text)
# columns = {}

# for i in range(1, len(blocks), 2):
#     col_no = blocks[i].strip()
#     block  = blocks[i + 1]

#     def find(pattern, flags=0, group=1):
#         m = re.search(pattern, block, flags)
#         if not m:
#             return None
#         val = m.group(group)
#         return val.replace(",", "") if isinstance(val, str) else val

#     guiding_load = find(r"GUIDING\s*LOAD\s*CASE:\s*(\d+)")
#     guiding_joint = find(r"END\s*JOINT:\s*(\d+)")
#     req_steel = find(r"REQD\.\s*STEEL\s*AREA\s*[:=]?\s*([\d.]+)")
#     req_conc = find(r"REQD\.\s*CONCRETE\s*AREA\s*[:=]?\s*([\d.]+)")
#     prov_steel = find(r"MAIN\s*REINFORCEMENT\s*:\s*Provide.*?\((?:\s*[\d.]+%[, ]+\s*)?([\d.]+)", flags=re.DOTALL)
#     no_bars = find(r"Provide\s+(\d+)\s*-\s*\d+\s*dia")
#     bar_dia = find(r"Provide\s+\d+\s*-\s*(\d+)\s*dia")

#     conf_dia = find(r"CONFINING\s*REINFORCEMENT\s*:\s*Provide\s*(\d+)")
#     conf_space = find(r"rectangular\s*ties\s*@\s*(\d+)")
#     conf_len = find(r"over\s+a\s+length\s*([\d.]+)\s*mm")
#     tie_dia = find(r"TIE\s*REINFORCEMENT\s*.*?Provide\s*(\d+)", flags=re.DOTALL)
#     tie_space = find(r"TIE\s*REINFORCEMENT\s*.*?@\s*(\d+)", flags=re.DOTALL)

#     bar_area = total_area = None
#     if bar_dia and no_bars:
#         try:
#             dia = float(bar_dia)
#             count = float(no_bars)
#             bar_area = math.pi * (dia ** 2) / 4.0
#             total_area = bar_area * count
#         except:
#             pass

#     area_check = ""
#     if prov_steel and total_area is not None:
#         try:
#             prov_val = float(prov_steel)
#             area_check = "YES" if abs(prov_val - total_area) <= TOLERANCE_MM2 else "NO"
#         except:
#             pass

#     calc_ratio = ""
#     try:
#         if prov_steel and req_conc:
#             prov_val = float(prov_steel)
#             conc_val = float(req_conc)
#             if conc_val > 0:
#                 calc_ratio = round((prov_val / conc_val) * 100, 3)
#     except:
#         calc_ratio = ""

#     min_ratio_check = max_ratio_check = ""
#     try:
#         if calc_ratio != "":
#             min_ratio_check = "YES" if calc_ratio > 0.8 else "NO"
#             max_ratio_check = "YES" if calc_ratio < 4 else "NO"
#     except:
#         pass

#     min_bar_check = min_dia_check = ""
#     try:
#         bars_val = int(no_bars) if no_bars else None
#         dia_val = float(bar_dia) if bar_dia else None
#         if bars_val is not None:
#             min_bar_check = "YES" if bars_val >= 4 else "NO"
#         if dia_val is not None:
#             min_dia_check = "YES" if dia_val >= 12 else "NO"
#     except:
#         pass

#     columns[col_no] = {
#         "Column No": col_no,
#         "Guiding Load Case": guiding_load,
#         "Guiding Joint": guiding_joint,
#         "Req Steel Area (Sq.mm)": req_steel,
#         "Req Concrete Area (Sq.mm)": req_conc,
#         "Provided Steel Area (Sq.mm)": prov_steel,
#         "Steel Ratio (%)": calc_ratio,
#         "No of Bars": no_bars,
#         "Bar Dia (mm)": bar_dia,
#         "Area 1 Bar (mmÂ²)": round(bar_area, 2) if bar_area else "",
#         "Total Bar Area (mmÂ²)": round(total_area, 2) if total_area else "",
#         "Area Check": area_check,
#         "Confinement Dia (mm)": conf_dia,
#         "Confinement Spacing (mm c/c)": conf_space,
#         "Confinement Length (mm)": conf_len,
#         "Tie Dia (mm)": tie_dia,
#         "Tie Spacing (mm c/c)": tie_space,
#         "Min Steel Ratio (>0.8)": min_ratio_check,
#         "Max Steel Ratio (<4)": max_ratio_check,
#         "Min 4 Bars (>=4)": min_bar_check,
#         "Min Bar Dia (>=12)": min_dia_check,
#     }

# # ============================================================
# # ADD COLUMN DEPTH & WIDTH FROM STAAD
# # ============================================================
# def dispid(obj, name):
#     try:
#         return obj._oleobj_.GetIDsOfNames(name)
#     except Exception:
#         return None

# def get_prismatic_section_mm(member_no, propUI, prop, UNIT_TO_MM=25.4):
#     sig = (
#         (pythoncom.VT_I4, 0),
#         (pythoncom.VT_VARIANT, 1), (pythoncom.VT_VARIANT, 1),
#         (pythoncom.VT_VARIANT, 1), (pythoncom.VT_VARIANT, 1), (pythoncom.VT_VARIANT, 1),
#         (pythoncom.VT_VARIANT, 1), (pythoncom.VT_VARIANT, 1), (pythoncom.VT_VARIANT, 1),
#     )
#     def _try_iface(iface):
#         if iface is None:
#             return False, None, None
#         did = dispid(iface, "GetBeamProperty")
#         if not did:
#             return False, None, None
#         W = VARIANT(pythoncom.VT_BYREF | pythoncom.VT_R8, 0.0)
#         D = VARIANT(pythoncom.VT_BYREF | pythoncom.VT_R8, 0.0)
#         Ax = Ay = Az = Ix = Iy = Iz = VARIANT(pythoncom.VT_BYREF | pythoncom.VT_R8, 0.0)
#         try:
#             ret = iface._oleobj_.InvokeTypes(
#                 did, 0, pythoncom.DISPATCH_METHOD, (pythoncom.VT_I4, 0),
#                 sig, int(member_no), W, D, Ax, Ay, Az, Ix, Iy, Iz
#             )
#             ok = int(ret[0] if isinstance(ret, tuple) else ret)
#             if ok == 1:
#                 return True, float(W.value)*UNIT_TO_MM, float(D.value)*UNIT_TO_MM
#         except Exception:
#             pass
#         return False, None, None
#     for iface in (propUI, prop):
#         ok, w, d = _try_iface(iface)
#         if ok:
#             return True, w, d
#     return False, None, None

# print("\nðŸ”— Connecting to STAAD to fetch section sizes...")
# pythoncom.CoInitialize()
# os_app = Dispatch("StaadPro.OpenSTAAD")
# propUI = getattr(os_app, "PropertyUI", None)
# prop   = getattr(os_app, "Property",   None)

# for col_no, data in columns.items():
#     member_id = int(col_no)
#     ok, w_mm_localZ, d_mm_localY = get_prismatic_section_mm(member_id, propUI, prop)
#     if not ok:
#         w_mm_localZ, d_mm_localY = 0.0, 0.0
#     data["Column Width (mm)"] = round(w_mm_localZ)
#     data["Column Depth (mm)"] = round(d_mm_localY)

# print("âœ… Added Column Depth (mm) and Column Width (mm).")

# # ============================================================
# # BAR SPACING LOGIC (faces: Width & Depth)
# # ============================================================
# # def _spacing_for_face(face_len_mm: float, total_bars: int, bar_dia_mm: float):
# #     """
# #     Returns dict with:
# #       s0, s, ok0, ok, recalc_used
# #     Logic:
# #       - Effective clear length = face_len - (2 cover + 2 stirrup) = face_len - 96
# #       - Bars per edge = floor((total_bars - 4 corners)/4) + 2 (incl. corners)
# #       - initial spacing s0 = (eff - bars_each_edge*bar_dia) / (bars_each_edge - 1)
# #       - if s0 not in (30, 300], recalc once:
# #           > 300  => add one intermediate per edge
# #           <= 30  => remove one intermediate per edge (not below 2 bars per edge)
# #     """
# #     result = {"s0": "", "s": "", "ok0": "", "ok": "", "recalc_used": "NO"}

# #     # guard
# #     if total_bars is None or bar_dia_mm is None:
# #         result.update(s0="", s="", ok0="NO", ok="NO")
# #         return result
# #     try:
# #         total_bars = int(total_bars)
# #         bar_dia_mm = float(bar_dia_mm)
# #     except:
# #         result.update(s0="", s="", ok0="NO", ok="NO")
# #         return result

# #     if total_bars < 4 or face_len_mm <= 0 or bar_dia_mm <= 0:
# #         result.update(s0="", s="", ok0="NO", ok="NO")
# #         return result

# #     eff = face_len_mm - DEDUCT
# #     if eff <= 0:
# #         result.update(s0="", s="", ok0="NO", ok="NO")
# #         return result

# #     remain = max(total_bars - 4, 0)  # after corners
# #     per_edge_intermediate = remain // 4
# #     bars_each_edge = per_edge_intermediate + 2  # 2 corners + intermediates
# #     if bars_each_edge < 2:
# #         bars_each_edge = 2

# #     def compute_spacing(bars_edge: int) -> float:
# #         spacing_count = max(bars_edge - 1, 1)
# #         used_bars_len = bars_edge * bar_dia_mm
# #         clear_len = eff - used_bars_len
# #         return clear_len / spacing_count

# #     def in_range(v: float) -> bool:
# #         return (v > MIN_CLEAR) and (v <= MAX_SPACING)

# #     # initial
# #     s0 = compute_spacing(bars_each_edge)
# #     ok0 = "YES" if in_range(s0) else "NO"

# #     # default final = initial
# #     s_final = s0
# #     ok_final = ok0
# #     recalc_used = "NO"

# #     if ok0 == "NO":
# #         recalc_used = "YES"
# #         # decide add/remove by failure direction
# #         if s0 > MAX_SPACING:
# #             # too wide -> add one intermediate bar per edge
# #             bars_each_edge_try = bars_each_edge + 1
# #         else:
# #             # too tight -> remove one intermediate per edge (keep corners)
# #             bars_each_edge_try = max(bars_each_edge - 1, 2)

# #         s_try = compute_spacing(bars_each_edge_try)
# #         if in_range(s_try):
# #             s_final = s_try
# #             ok_final = "YES"
# #         else:
# #             s_final = s_try
# #             ok_final = "NO"

# #     result["s0"] = round(s0, 2)
# #     result["s"] = round(s_final, 2)
# #     result["ok0"] = ok0
# #     result["ok"] = ok_final
# #     result["recalc_used"] = recalc_used
# #     return result

# def _spacing_for_face(face_len_mm: float, total_bars: int, bar_dia_mm: float):
#     """
#     Multi-iteration version:
#     Keeps recalculating until spacing fits 30â€“300 mm or bars_each_edge <= 2.
#     If spacing never fits, returns last values as NO.
#     """
#     result = {"s0": "", "s": "", "ok0": "", "ok": "", "recalc_used": "NO"}

#     if not face_len_mm or not total_bars or not bar_dia_mm:
#         result.update(s0="", s="", ok0="NO", ok="NO")
#         return result
#     try:
#         total_bars = int(total_bars)
#         bar_dia_mm = float(bar_dia_mm)
#     except:
#         result.update(s0="", s="", ok0="NO", ok="NO")
#         return result

#     if total_bars < 4:
#         result.update(s0="", s="", ok0="NO", ok="NO")
#         return result

#     eff = face_len_mm - DEDUCT
#     if eff <= 0:
#         result.update(s0="", s="", ok0="NO", ok="NO")
#         return result

#     remain = max(total_bars - 4, 0)
#     per_edge_intermediate = remain // 4
#     bars_each_edge = per_edge_intermediate + 2  # incl. corners

#     def compute_spacing(bars_edge: int) -> float:
#         spacing_count = max(bars_edge - 1, 1)
#         used_bars_len = bars_edge * bar_dia_mm
#         clear_len = eff - used_bars_len
#         return clear_len / spacing_count

#     def in_range(v: float) -> bool:
#         return (v > MIN_CLEAR) and (v <= MAX_SPACING)

#     # initial
#     s0 = compute_spacing(bars_each_edge)
#     ok0 = "YES" if in_range(s0) else "NO"

#     s_final = s0
#     ok_final = ok0
#     recalc_used = "NO"

#     # Iterate until spacing OK or we run out of bars
#     iter_count = 0
#     while not in_range(s_final) and bars_each_edge > 2:
#         recalc_used = "YES"
#         iter_count += 1

#         if s_final > MAX_SPACING:
#             bars_each_edge += 1   # spacing too large, add bar
#         elif s_final <= MIN_CLEAR:
#             bars_each_edge -= 1   # spacing too small, remove bar

#         s_final = compute_spacing(bars_each_edge)
#         ok_final = "YES" if in_range(s_final) else "NO"

#         if iter_count > 50:  # safety stop
#             break

#     result["s0"] = round(s0, 2)
#     result["s"] = round(s_final, 2)
#     result["ok0"] = ok0
#     result["ok"] = ok_final
#     result["recalc_used"] = recalc_used
#     return result


# # compute spacing + effective sizes for each column
# for col_no, data in columns.items():
#     try:
#         cw = float(data.get("Column Width (mm)", 0) or 0)
#         cd = float(data.get("Column Depth (mm)", 0) or 0)
#         nb = data.get("No of Bars")
#         bd = data.get("Bar Dia (mm)")

#         # Effective clear lengths (after cover+stirrups)
#         data["Eff Width (mm)"] = round(cw - DEDUCT, 2) if cw else ""
#         data["Eff Depth (mm)"] = round(cd - DEDUCT, 2) if cd else ""

#         faceW = _spacing_for_face(cw, int(nb) if nb else None, float(bd) if bd else None)
#         faceD = _spacing_for_face(cd, int(nb) if nb else None, float(bd) if bd else None)

#         # short column names for spacing columns
#         data["sW0 (mm)"] = faceW["s0"]
#         data["sD0 (mm)"] = faceD["s0"]
#         data["sW (mm)"]  = faceW["s"]
#         data["sD (mm)"]  = faceD["s"]

#         # overall flags
#         data["Spacing OK"] = "YES" if (faceW["ok"] == "YES" and faceD["ok"] == "YES") else "NO"
#         data["Recalc Used"] = "YES" if (faceW["recalc_used"] == "YES" or faceD["recalc_used"] == "YES") else "NO"

#     except Exception:
#         data["Spacing OK"] = "NO"
#         data["Recalc Used"] = "YES"  # conservative

# print("âœ… Added effective sizes, spacing, checks, and recalc flags.")

# # ============================================================
# # WRITE TO EXCEL
# # ============================================================
# wb = openpyxl.Workbook()
# ws = wb.active
# ws.title = "Column_Design_Data"

# if not columns:
#     print("âŒ No columns found!")
#     wb.save(OUTPUT_XLSX)
# else:
#     headers = list(next(iter(columns.values())).keys())
#     ws.append(headers)
#     for c in columns.values():
#         ws.append([c.get(h, "") for h in headers])

#     # column widths
#     for col_idx, header in enumerate(headers, start=1):
#         ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(str(header)) + 4, 16), 40)

#     # header style
#     header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
#     header_font = Font(color="FFFFFF", bold=True)
#     for cell in ws[1]:
#         cell.fill = header_fill
#         cell.font = header_font

#     # fail highlighting
#     red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
#     bold_red = Font(color="9C0006", bold=True)

#     check_cols = [
#         "Area Check",
#         "Min Steel Ratio (>0.8)",
#         "Max Steel Ratio (<4)",
#         "Min 4 Bars (>=4)",
#         "Min Bar Dia (>=12)",
#         "Spacing OK",
#     ]
#     header_to_index = {h: i for i, h in enumerate(headers)}
#     fail_rows = []
#     for r in range(2, ws.max_row + 1):
#         row_failed = False
#         for col_name in check_cols:
#             idx = header_to_index.get(col_name)
#             if idx is None:
#                 continue
#             cell = ws.cell(row=r, column=idx + 1)
#             if str(cell.value).strip().upper() == "NO":
#                 cell.fill = red_fill
#                 cell.font = bold_red
#                 row_failed = True
#         if row_failed:
#             fail_rows.append(r)

#     total = len(columns)
#     n_fail = len(fail_rows)
#     n_pass = total - n_fail

#     ws.append([])
#     ws.append(["Summary", f"Total = {total}", f"Pass = {n_pass}", f"Fail = {n_fail}"])
#     for c in ws[ws.max_row]:
#         c.font = Font(bold=True)

#     # Failures-only sheet
#     ws_fail = wb.create_sheet("Failures_Only")
#     ws_fail.append(headers + ["Fail Flags"])
#     for row_idx in fail_rows:
#         values = [ws.cell(row=row_idx, column=j + 1).value for j in range(len(headers))]
#         flags = []
#         for col_name in check_cols:
#             idx = header_to_index.get(col_name)
#             if idx is None:
#                 continue
#             v = ws.cell(row=row_idx, column=idx + 1).value
#             if str(v).strip().upper() == "NO":
#                 flags.append(col_name)
#         ws_fail.append(values + [", ".join(flags)])

#     for cell in ws_fail[1]:
#         cell.fill = header_fill
#         cell.font = header_font

#     wb.save(OUTPUT_XLSX)
#     print(f"âœ… Extracted {total} columns â†’ {OUTPUT_XLSX}  |  âœ“ Pass: {n_pass}  âœ– Fail: {n_fail}")
#     notify_summary(total, n_pass, n_fail, OUTPUT_XLSX, mode=POPUP_MODE, popup_seconds=POPUP_SECONDS)


# import re
# import math
# from typing import Optional
# import openpyxl
# from openpyxl.styles import PatternFill, Font
# from openpyxl.utils import get_column_letter
# import pythoncom
# from win32com.client import Dispatch, VARIANT

# # ================================
# # CONFIG
# # ================================
# INPUT_FILE   = "MR. DINESH KHATRI.ANL"
# OUTPUT_XLSX  = "column_design_data_checked.xlsx"
# TOLERANCE_MM2 = 4

# POPUP_MODE: str = "popup"
# POPUP_SECONDS: Optional[int] = None

# # cover & stirrup for deductions (two faces each)
# COVER = 40            # mm
# STIRRUP_DIA = 8       # mm
# DEDUCT = 2*COVER + 2*STIRRUP_DIA  # 80 + 16 = 96

# MIN_CLEAR = 30.0      # mm
# MAX_SPACING = 300.0   # mm

# # ================================
# # POPUP FUNCTION
# # ================================
# def _centered_popup(title: str, message: str, stay_seconds: Optional[int] = None) -> bool:
#     try:
#         import tkinter as tk
#         root = tk.Tk()
#         root.title(title)
#         root.attributes("-topmost", True)
#         root.resizable(False, False)
#         root.configure(bg="#ffffff")

#         frm = tk.Frame(root, padx=25, pady=20, bg="#ffffff")
#         frm.pack(fill="both", expand=True)

#         lbl_title = tk.Label(frm, text="âœ… Please Check Sheet",
#                              font=("Segoe UI", 14, "bold"),
#                              bg="#ffffff", fg="#2C3E50")
#         lbl_title.pack(pady=(0, 10))

#         lbl_msg = tk.Label(frm,
#                            text=message + "\n\nkindly review.",
#                            font=("Segoe UI", 11),
#                            bg="#ffffff", fg="#333333", justify="left")
#         lbl_msg.pack()

#         btn = tk.Button(frm, text="Close", width=12,
#                         font=("Segoe UI", 10, "bold"),
#                         command=root.destroy,
#                         bg="#3498DB", fg="white")
#         btn.pack(pady=(15, 0))

#         root.update_idletasks()
#         w = root.winfo_width()
#         h = root.winfo_height()
#         sw = root.winfo_screenwidth()
#         sh = root.winfo_screenheight()
#         x = (sw // 2) - (w // 2)
#         y = (sh // 3) - (h // 3)
#         root.geometry(f"{w}x{h}+{x}+{y}")

#         if stay_seconds is not None:
#             root.after(int(stay_seconds * 1000), root.destroy)

#         root.mainloop()
#         return True
#     except Exception as e:
#         print("Popup error:", e)
#         return False


# def notify_summary(total: int, n_pass: int, n_fail: int, path: str,
#                    mode: str = POPUP_MODE, popup_seconds: Optional[int] = POPUP_SECONDS) -> None:
#     msg = f"Columns: {total}  âœ“ Pass: {n_pass}  âœ– Fail: {n_fail}\nSaved: {path}"
#     shown = False
#     if mode.lower() == "popup":
#         shown = _centered_popup("Column Design Check", msg, stay_seconds=popup_seconds)
#     print(f"INFO: Columns={total}, Pass={n_pass}, Fail={n_fail} â†’ {path}"
#           + ("" if shown else "  (popup failed â†’ console only)"))

# # ================================
# # READ .ANL FILE
# # ================================
# with open(INPUT_FILE, "r", encoding="utf-8", errors="ignore") as f:
#     text = f.read()

# blocks = re.split(r"C\s*O\s*L\s*U\s*M\s*N\s+N\s*O\.\s*(\d+)", text)
# columns = {}

# for i in range(1, len(blocks), 2):
#     col_no = blocks[i].strip()
#     block  = blocks[i + 1]

#     def find(pattern, flags=0, group=1):
#         m = re.search(pattern, block, flags)
#         if not m:
#             return None
#         val = m.group(group)
#         return val.replace(",", "") if isinstance(val, str) else val

#     guiding_load = find(r"GUIDING\s*LOAD\s*CASE:\s*(\d+)")
#     guiding_joint = find(r"END\s*JOINT:\s*(\d+)")
#     req_steel = find(r"REQD\.\s*STEEL\s*AREA\s*[:=]?\s*([\d.]+)")
#     req_conc = find(r"REQD\.\s*CONCRETE\s*AREA\s*[:=]?\s*([\d.]+)")
#     prov_steel = find(r"MAIN\s*REINFORCEMENT\s*:\s*Provide.*?\((?:\s*[\d.]+%[, ]+\s*)?([\d.]+)", flags=re.DOTALL)
#     no_bars = find(r"Provide\s+(\d+)\s*-\s*\d+\s*dia")
#     bar_dia = find(r"Provide\s+\d+\s*-\s*(\d+)\s*dia")

#     conf_dia = find(r"CONFINING\s*REINFORCEMENT\s*:\s*Provide\s*(\d+)")
#     conf_space = find(r"rectangular\s*ties\s*@\s*(\d+)")
#     conf_len = find(r"over\s+a\s+length\s*([\d.]+)\s*mm")
#     tie_dia = find(r"TIE\s*REINFORCEMENT\s*.*?Provide\s*(\d+)", flags=re.DOTALL)
#     tie_space = find(r"TIE\s*REINFORCEMENT\s*.*?@\s*(\d+)", flags=re.DOTALL)

#     bar_area = total_area = None
#     if bar_dia and no_bars:
#         try:
#             dia = float(bar_dia)
#             count = float(no_bars)
#             bar_area = math.pi * (dia ** 2) / 4.0
#             total_area = bar_area * count
#         except:
#             pass

#     area_check = ""
#     if prov_steel and total_area is not None:
#         try:
#             prov_val = float(prov_steel)
#             area_check = "YES" if abs(prov_val - total_area) <= TOLERANCE_MM2 else "NO"
#         except:
#             pass

#     calc_ratio = ""
#     try:
#         if prov_steel and req_conc:
#             prov_val = float(prov_steel)
#             conc_val = float(req_conc)
#             if conc_val > 0:
#                 calc_ratio = round((prov_val / conc_val) * 100, 3)
#     except:
#         calc_ratio = ""

#     min_ratio_check = max_ratio_check = ""
#     try:
#         if calc_ratio != "":
#             min_ratio_check = "YES" if calc_ratio > 0.8 else "NO"
#             max_ratio_check = "YES" if calc_ratio < 4 else "NO"
#     except:
#         pass

#     min_bar_check = min_dia_check = ""
#     try:
#         bars_val = int(no_bars) if no_bars else None
#         dia_val = float(bar_dia) if bar_dia else None
#         if bars_val is not None:
#             min_bar_check = "YES" if bars_val >= 4 else "NO"
#         if dia_val is not None:
#             min_dia_check = "YES" if dia_val >= 12 else "NO"
#     except:
#         pass

#     columns[col_no] = {
#         "Column No": col_no,
#         "Guiding Load Case": guiding_load,
#         "Guiding Joint": guiding_joint,
#         "Req Steel Area (Sq.mm)": req_steel,
#         "Req Concrete Area (Sq.mm)": req_conc,
#         "Provided Steel Area (Sq.mm)": prov_steel,
#         "Steel Ratio (%)": calc_ratio,
#         "No of Bars": no_bars,
#         "Bar Dia (mm)": bar_dia,
#         "Area 1 Bar (mmÂ²)": round(bar_area, 2) if bar_area else "",
#         "Total Bar Area (mmÂ²)": round(total_area, 2) if total_area else "",
#         "Area Check": area_check,
#         "Confinement Dia (mm)": conf_dia,
#         "Confinement Spacing (mm c/c)": conf_space,
#         "Confinement Length (mm)": conf_len,
#         "Tie Dia (mm)": tie_dia,
#         "Tie Spacing (mm c/c)": tie_space,
#         "Min Steel Ratio (>0.8)": min_ratio_check,
#         "Max Steel Ratio (<4)": max_ratio_check,
#         "Min 4 Bars (>=4)": min_bar_check,
#         "Min Bar Dia (>=12)": min_dia_check,
#     }

# # ============================================================
# # ADD COLUMN DEPTH & WIDTH FROM STAAD
# # ============================================================
# def dispid(obj, name):
#     try:
#         return obj._oleobj_.GetIDsOfNames(name)
#     except Exception:
#         return None

# def get_prismatic_section_mm(member_no, propUI, prop, UNIT_TO_MM=25.4):
#     sig = (
#         (pythoncom.VT_I4, 0),
#         (pythoncom.VT_VARIANT, 1), (pythoncom.VT_VARIANT, 1),
#         (pythoncom.VT_VARIANT, 1), (pythoncom.VT_VARIANT, 1), (pythoncom.VT_VARIANT, 1),
#         (pythoncom.VT_VARIANT, 1), (pythoncom.VT_VARIANT, 1), (pythoncom.VT_VARIANT, 1),
#     )
#     def _try_iface(iface):
#         if iface is None:
#             return False, None, None
#         did = dispid(iface, "GetBeamProperty")
#         if not did:
#             return False, None, None
#         W = VARIANT(pythoncom.VT_BYREF | pythoncom.VT_R8, 0.0)
#         D = VARIANT(pythoncom.VT_BYREF | pythoncom.VT_R8, 0.0)
#         Ax = Ay = Az = Ix = Iy = Iz = VARIANT(pythoncom.VT_BYREF | pythoncom.VT_R8, 0.0)
#         try:
#             ret = iface._oleobj_.InvokeTypes(
#                 did, 0, pythoncom.DISPATCH_METHOD, (pythoncom.VT_I4, 0),
#                 sig, int(member_no), W, D, Ax, Ay, Az, Ix, Iy, Iz
#             )
#             ok = int(ret[0] if isinstance(ret, tuple) else ret)
#             if ok == 1:
#                 return True, float(W.value)*UNIT_TO_MM, float(D.value)*UNIT_TO_MM
#         except Exception:
#             pass
#         return False, None, None
#     for iface in (propUI, prop):
#         ok, w, d = _try_iface(iface)
#         if ok:
#             return True, w, d
#     return False, None, None

# print("\nðŸ”— Connecting to STAAD to fetch section sizes...")
# pythoncom.CoInitialize()
# os_app = Dispatch("StaadPro.OpenSTAAD")
# propUI = getattr(os_app, "PropertyUI", None)
# prop   = getattr(os_app, "Property",   None)

# for col_no, data in columns.items():
#     member_id = int(col_no)
#     ok, w_mm_localZ, d_mm_localY = get_prismatic_section_mm(member_id, propUI, prop)
#     if not ok:
#         w_mm_localZ, d_mm_localY = 0.0, 0.0
#     data["Column Width (mm)"] = round(w_mm_localZ)
#     data["Column Depth (mm)"] = round(d_mm_localY)

# print("âœ… Added Column Depth (mm) and Column Width (mm).")

# # ============================================================
# # BAR SPACING LOGIC (Widthâ€“Depth coupled, multi-iteration)
# # ============================================================
# def _spacing_balance(width_len_mm: float, depth_len_mm: float, total_bars: int, bar_dia_mm: float):
#     """
#     Distribute bars equally to 4 edges after 4 corners.
#     If either face (Width or Depth) fails 30 < s <= 300, iterate:
#       - Move ONE intermediate bar per edge from failing face to the other face.
#       - Stop if both faces OK, or when a face hits corners-only (bars_each_edge <= 2).
#     Returns dict with initial and final spacings & flags.
#     """
#     out = {
#         "sW0": "", "sD0": "", "sW": "", "sD": "",
#         "okW": "NO", "okD": "NO", "recalc_used": "NO"
#     }

#     # guards
#     try:
#         total_bars = int(total_bars) if total_bars is not None else None
#         bar_dia_mm = float(bar_dia_mm) if bar_dia_mm is not None else None
#     except:
#         total_bars = None
#         bar_dia_mm = None

#     if (not width_len_mm) or (not depth_len_mm) or (not total_bars) or (not bar_dia_mm) or total_bars < 4:
#         return out

#     effW = width_len_mm - DEDUCT
#     effD = depth_len_mm - DEDUCT
#     if effW <= 0 or effD <= 0:
#         return out

#     remain = max(total_bars - 4, 0)
#     per_edge_intermediate = remain // 4
#     # Start equal on both faces (each edge)
#     barsW = per_edge_intermediate + 2
#     barsD = per_edge_intermediate + 2
#     barsW = max(barsW, 2)
#     barsD = max(barsD, 2)

#     def spacing(eff_len, bars_edge):
#         spacing_count = max(bars_edge - 1, 1)
#         used_bars_len = bars_edge * bar_dia_mm
#         clear_len = eff_len - used_bars_len
#         return clear_len / spacing_count

#     def in_range(v): return (v > MIN_CLEAR) and (v <= MAX_SPACING)

#     # initial spacings
#     sW = spacing(effW, barsW)
#     sD = spacing(effD, barsD)
#     out["sW0"] = round(sW, 2)
#     out["sD0"] = round(sD, 2)
#     okW = in_range(sW)
#     okD = in_range(sD)

#     # iterate if any fail
#     recalc = False
#     iters = 0
#     while (not (okW and okD)) and (barsW > 2 or barsD > 2):
#         recalc = True
#         iters += 1

#         # decide direction based on which face fails and how
#         if not okW and sW <= MIN_CLEAR and barsW > 2:
#             # width too tight â†’ move one bar from Width to Depth
#             barsW -= 1
#             barsD += 1
#         elif not okD and sD <= MIN_CLEAR and barsD > 2:
#             # depth too tight â†’ move one bar from Depth to Width
#             barsD -= 1
#             barsW += 1
#         elif not okW and sW > MAX_SPACING and barsD > 2:
#             # width too wide â†’ move one bar from Depth to Width
#             barsW += 1
#             barsD -= 1
#         elif not okD and sD > MAX_SPACING and barsW > 2:
#             # depth too wide â†’ move one bar from Width to Depth
#             barsD += 1
#             barsW -= 1
#         else:
#             # cannot move further without breaking corners-only rule
#             break

#         sW = spacing(effW, barsW)
#         sD = spacing(effD, barsD)
#         okW = in_range(sW)
#         okD = in_range(sD)

#         if iters > 200:   # hard safety cap
#             break

#     out["sW"] = round(sW, 2)
#     out["sD"] = round(sD, 2)
#     out["okW"] = "YES" if okW else "NO"
#     out["okD"] = "YES" if okD else "NO"
#     out["recalc_used"] = "YES" if recalc else "NO"
#     return out

# # compute spacing + effective sizes for each column
# for col_no, data in columns.items():
#     try:
#         cw = float(data.get("Column Width (mm)", 0) or 0)
#         cd = float(data.get("Column Depth (mm)", 0) or 0)
#         nb = data.get("No of Bars")
#         bd = data.get("Bar Dia (mm)")

#         # Effective clear lengths (after cover+stirrups)
#         data["Eff Width (mm)"] = round(cw - DEDUCT, 2) if cw else ""
#         data["Eff Depth (mm)"] = round(cd - DEDUCT, 2) if cd else ""

#         res = _spacing_balance(cw, cd, int(nb) if nb else None, float(bd) if bd else None)

#         # short column names for spacing columns
#         data["sW0 (mm)"] = res["sW0"]
#         data["sD0 (mm)"] = res["sD0"]
#         data["sW (mm)"]  = res["sW"]
#         data["sD (mm)"]  = res["sD"]

#         # overall flags
#         data["Spacing OK"] = "YES" if (res["okW"] == "YES" and res["okD"] == "YES") else "NO"
#         data["Recalc Used"] = res["recalc_used"]

#     except Exception:
#         data["Spacing OK"] = "NO"
#         data["Recalc Used"] = "YES"  # conservative

# print("âœ… Added effective sizes, coupled spacing balance, checks, and recalc flags.")

# # ============================================================
# # WRITE TO EXCEL
# # ============================================================
# wb = openpyxl.Workbook()
# ws = wb.active
# ws.title = "Column_Design_Data"

# if not columns:
#     print("âŒ No columns found!")
#     wb.save(OUTPUT_XLSX)
# else:
#     headers = list(next(iter(columns.values())).keys())
#     ws.append(headers)
#     for c in columns.values():
#         ws.append([c.get(h, "") for h in headers])

#     # column widths
#     for col_idx, header in enumerate(headers, start=1):
#         ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(str(header)) + 4, 16), 40)

#     # header style
#     header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
#     header_font = Font(color="FFFFFF", bold=True)
#     for cell in ws[1]:
#         cell.fill = header_fill
#         cell.font = header_font

#     # fail highlighting
#     red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
#     bold_red = Font(color="9C0006", bold=True)

#     check_cols = [
#         "Area Check",
#         "Min Steel Ratio (>0.8)",
#         "Max Steel Ratio (<4)",
#         "Min 4 Bars (>=4)",
#         "Min Bar Dia (>=12)",
#         "Spacing OK",
#     ]
#     header_to_index = {h: i for i, h in enumerate(headers)}
#     fail_rows = []
#     for r in range(2, ws.max_row + 1):
#         row_failed = False
#         for col_name in check_cols:
#             idx = header_to_index.get(col_name)
#             if idx is None:
#                 continue
#             cell = ws.cell(row=r, column=idx + 1)
#             if str(cell.value).strip().upper() == "NO":
#                 cell.fill = red_fill
#                 cell.font = bold_red
#                 row_failed = True
#         if row_failed:
#             fail_rows.append(r)

#     total = len(columns)
#     n_fail = len(fail_rows)
#     n_pass = total - n_fail

#     ws.append([])
#     ws.append(["Summary", f"Total = {total}", f"Pass = {n_pass}", f"Fail = {n_fail}"])
#     for c in ws[ws.max_row]:
#         c.font = Font(bold=True)

#     # Failures-only sheet
#     ws_fail = wb.create_sheet("Failures_Only")
#     ws_fail.append(headers + ["Fail Flags"])
#     for row_idx in fail_rows:
#         values = [ws.cell(row=row_idx, column=j + 1).value for j in range(len(headers))]
#         flags = []
#         for col_name in check_cols:
#             idx = header_to_index.get(col_name)
#             if idx is None:
#                 continue
#             v = ws.cell(row=row_idx, column=idx + 1).value
#             if str(v).strip().upper() == "NO":
#                 flags.append(col_name)
#         ws_fail.append(values + [", ".join(flags)])

#     for cell in ws_fail[1]:
#         cell.fill = header_fill
#         cell.font = header_font

#     wb.save(OUTPUT_XLSX)
#     print(f"âœ… Extracted {total} columns â†’ {OUTPUT_XLSX}  |  âœ“ Pass: {n_pass}  âœ– Fail: {n_fail}")
#     notify_summary(total, n_pass, n_fail, OUTPUT_XLSX, mode=POPUP_MODE, popup_seconds=POPUP_SECONDS)





import re
import math
from typing import Optional
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import pythoncom
from win32com.client import Dispatch, VARIANT
import pandas as pd  # <-- ADDED: for reading the height sheet

# ================================
# CONFIG
# ================================
INPUT_FILE   = "MR. DINESH KHATRI.ANL"
OUTPUT_XLSX  = "column_design_data_checked.xlsx"
TOLERANCE_MM2 = 4

POPUP_MODE: str = "popup"
POPUP_SECONDS: Optional[int] = None

# cover & stirrup for deductions (two faces each)
COVER = 40            # mm
STIRRUP_DIA = 8       # mm
DEDUCT = 2*COVER + 2*STIRRUP_DIA  # 80 + 16 = 96

MIN_CLEAR = 30.0      # mm
MAX_SPACING = 300.0   # mm

def tau_bd_base_from_fck(fck: float) -> float:
    """IS 456 Table 21 base Ï„bd for plain bars in tension."""
    if fck <= 25:
        return 1.4
    if fck <= 30:
        return 1.5
    if fck <= 35:
        return 1.7
    if fck <= 40:
        return 1.9
    return 2.0

# ================================
# POPUP FUNCTION
# ================================
def _centered_popup(title: str, message: str, stay_seconds: Optional[int] = None) -> bool:
    try:
        import tkinter as tk
        root = tk.Tk()
        root.title(title)
        root.attributes("-topmost", True)
        root.resizable(False, False)
        root.configure(bg="#ffffff")

        frm = tk.Frame(root, padx=25, pady=20, bg="#ffffff")
        frm.pack(fill="both", expand=True)

        lbl_title = tk.Label(frm, text="âœ… Please Check Sheet",
                             font=("Segoe UI", 14, "bold"),
                             bg="#ffffff", fg="#2C3E50")
        lbl_title.pack(pady=(0, 10))

        lbl_msg = tk.Label(frm,
                           text=message + "\n\nkindly review.",
                           font=("Segoe UI", 11),
                           bg="#ffffff", fg="#333333", justify="left")
        lbl_msg.pack()

        btn = tk.Button(frm, text="Close", width=12,
                        font=("Segoe UI", 10, "bold"),
                        command=root.destroy,
                        bg="#3498DB", fg="white")
        btn.pack(pady=(15, 0))

        root.update_idletasks()
        w = root.winfo_width()
        h = root.winfo_height()
        sw = root.winfo_screenwidth()
        sh = root.winfo_screenheight()
        x = (sw // 2) - (w // 2)
        y = (sh // 3) - (h // 3)
        root.geometry(f"{w}x{h}+{x}+{y}")

        if stay_seconds is not None:
            root.after(int(stay_seconds * 1000), root.destroy)

        root.mainloop()
        return True
    except Exception as e:
        print("Popup error:", e)
        return False


def notify_summary(total: int, n_pass: int, n_fail: int, path: str,
                   mode: str = POPUP_MODE, popup_seconds: Optional[int] = POPUP_SECONDS) -> None:
    msg = f"Columns: {total}  âœ“ Pass: {n_pass}  âœ– Fail: {n_fail}\nSaved: {path}"
    shown = False
    if mode.lower() == "popup":
        shown = _centered_popup("Column Design Check", msg, stay_seconds=popup_seconds)
    print(f"INFO: Columns={total}, Pass={n_pass}, Fail={n_fail} â†’ {path}"
          + ("" if shown else "  (popup failed â†’ console only)"))

# ================================
# READ .ANL FILE
# ================================
with open(INPUT_FILE, "r", encoding="utf-8", errors="ignore") as f:
    text = f.read()

blocks = re.split(r"C\s*O\s*L\s*U\s*M\s*N\s+N\s*O\.\s*(\d+)", text)
columns = {}

for i in range(1, len(blocks), 2):
    col_no = blocks[i].strip()
    block  = blocks[i + 1]

    def find(pattern, flags=0, group=1):
        m = re.search(pattern, block, flags)
        if not m:
            return None
        val = m.group(group)
        return val.replace(",", "") if isinstance(val, str) else val

    guiding_load = find(r"GUIDING\s*LOAD\s*CASE:\s*(\d+)")
    guiding_joint = find(r"END\s*JOINT:\s*(\d+)")
    req_steel = find(r"REQD\.\s*STEEL\s*AREA\s*[:=]?\s*([\d.]+)")
    req_conc = find(r"REQD\.\s*CONCRETE\s*AREA\s*[:=]?\s*([\d.]+)")
    prov_steel = find(r"MAIN\s*REINFORCEMENT\s*:\s*Provide.*?\((?:\s*[\d.]+%[, ]+\s*)?([\d.]+)", flags=re.DOTALL)
    no_bars = find(r"Provide\s+(\d+)\s*-\s*\d+\s*dia")
    bar_dia = find(r"Provide\s+\d+\s*-\s*(\d+)\s*dia")

    conf_dia = find(r"CONFINING\s*REINFORCEMENT\s*:\s*Provide\s*(\d+)")
    conf_space = find(r"rectangular\s*ties\s*@\s*(\d+)")
    conf_len = find(r"over\s+a\s+length\s*([\d.]+)\s*mm")
    tie_dia = find(r"TIE\s*REINFORCEMENT\s*.*?Provide\s*(\d+)", flags=re.DOTALL)
    tie_space = find(r"TIE\s*REINFORCEMENT\s*.*?@\s*(\d+)", flags=re.DOTALL)

    bar_area = total_area = None
    if bar_dia and no_bars:
        try:
            dia = float(bar_dia)
            count = float(no_bars)
            bar_area = math.pi * (dia ** 2) / 4.0
            total_area = bar_area * count
        except:
            pass

    area_check = ""
    if prov_steel and total_area is not None:
        try:
            prov_val = float(prov_steel)
            area_check = "YES" if abs(prov_val - total_area) <= TOLERANCE_MM2 else "NO"
        except:
            pass

    calc_ratio = ""
    try:
        if prov_steel and req_conc:
            prov_val = float(prov_steel)
            conc_val = float(req_conc)
            if conc_val > 0:
                calc_ratio = round((prov_val / conc_val) * 100, 3)
    except:
        calc_ratio = ""

    min_ratio_check = max_ratio_check = ""
    try:
        if calc_ratio != "":
            min_ratio_check = "YES" if calc_ratio > 0.8 else "NO"
            max_ratio_check = "YES" if calc_ratio < 4 else "NO"
    except:
        pass

    min_bar_check = min_dia_check = ""
    try:
        bars_val = int(no_bars) if no_bars else None
        dia_val = float(bar_dia) if bar_dia else None
        if bars_val is not None:
            min_bar_check = "YES" if bars_val >= 4 else "NO"
        if dia_val is not None:
            min_dia_check = "YES" if dia_val >= 12 else "NO"
    except:
        pass

    columns[col_no] = {
        "Column No": col_no,
        "Guiding Load Case": guiding_load,
        "Guiding Joint": guiding_joint,
        "Req Steel Area (Sq.mm)": req_steel,
        "Req Concrete Area (Sq.mm)": req_conc,
        "Provided Steel Area (Sq.mm)": prov_steel,
        "Steel Ratio (%)": calc_ratio,
        "No of Bars": no_bars,
        "Bar Dia (mm)": bar_dia,
        "Area 1 Bar (mmÂ²)": round(bar_area, 2) if bar_area else "",
        "Total Bar Area (mmÂ²)": round(total_area, 2) if total_area else "",
        "Area Check": area_check,
        "Confinement Dia (mm)": conf_dia,
        "Confinement Spacing (mm c/c)": conf_space,
        "Confinement Length (mm)": conf_len,
        "Tie Dia (mm)": tie_dia,
        "Tie Spacing (mm c/c)": tie_space,
        "Min Steel Ratio (>0.8)": min_ratio_check,
        "Max Steel Ratio (<4)": max_ratio_check,
        "Min 4 Bars (>=4)": min_bar_check,
        "Min Bar Dia (>=12)": min_dia_check,
    }

# ============================================================
# ADD COLUMN DEPTH & WIDTH FROM STAAD
# ============================================================
def dispid(obj, name):
    try:
        return obj._oleobj_.GetIDsOfNames(name)
    except Exception:
        return None

def get_prismatic_section_mm(member_no, propUI, prop, UNIT_TO_MM=25.4):
    sig = (
        (pythoncom.VT_I4, 0),
        (pythoncom.VT_VARIANT, 1), (pythoncom.VT_VARIANT, 1),
        (pythoncom.VT_VARIANT, 1), (pythoncom.VT_VARIANT, 1), (pythoncom.VT_VARIANT, 1),
        (pythoncom.VT_VARIANT, 1), (pythoncom.VT_VARIANT, 1), (pythoncom.VT_VARIANT, 1),
    )
    def _try_iface(iface):
        if iface is None:
            return False, None, None
        did = dispid(iface, "GetBeamProperty")
        if not did:
            return False, None, None
        W = VARIANT(pythoncom.VT_BYREF | pythoncom.VT_R8, 0.0)
        D = VARIANT(pythoncom.VT_BYREF | pythoncom.VT_R8, 0.0)
        Ax = Ay = Az = Ix = Iy = Iz = VARIANT(pythoncom.VT_BYREF | pythoncom.VT_R8, 0.0)
        try:
            ret = iface._oleobj_.InvokeTypes(
                did, 0, pythoncom.DISPATCH_METHOD, (pythoncom.VT_I4, 0),
                sig, int(member_no), W, D, Ax, Ay, Az, Ix, Iy, Iz
            )
            ok = int(ret[0] if isinstance(ret, tuple) else ret)
            if ok == 1:
                return True, float(W.value)*UNIT_TO_MM, float(D.value)*UNIT_TO_MM
        except Exception:
            pass
        return False, None, None
    for iface in (propUI, prop):
        ok, w, d = _try_iface(iface)
        if ok:
            return True, w, d
    return False, None, None

print("\nðŸ”— Connecting to STAAD to fetch section sizes...")
pythoncom.CoInitialize()
os_app = Dispatch("StaadPro.OpenSTAAD")
propUI = getattr(os_app, "PropertyUI", None)
prop   = getattr(os_app, "Property",   None)

for col_no, data in columns.items():
    member_id = int(col_no)
    ok, w_mm_localZ, d_mm_localY = get_prismatic_section_mm(member_id, propUI, prop)
    if not ok:
        w_mm_localZ, d_mm_localY = 0.0, 0.0
    data["Column Width (mm)"] = round(w_mm_localZ)
    data["Column Depth (mm)"] = round(d_mm_localY)

print("âœ… Added Column Depth (mm) and Column Width (mm).")



# ============================================================
# MERGE FOOTING MEMBER INFO (Member ID == Column No)
# ============================================================
def merge_footing_members(columns, footing_excel):
    import os
    if not os.path.exists(footing_excel):
        print(f"âš ï¸ Footing member sheet not found: {footing_excel}")
        return

    try:
        df_f = pd.read_excel(footing_excel)
        df_f.columns = [c.strip() for c in df_f.columns]

        if "Member ID" not in df_f.columns or "Footing Member (YES/NO)" not in df_f.columns:
            print("âš ï¸ Footing Excel must contain 'Member ID' and 'Footing Member (YES/NO)'.")
            return

        footing_map = dict(zip(df_f["Member ID"].astype(str), df_f["Footing Member (YES/NO)"]))

        for col_no, data in columns.items():
            data["Footing Member (YES/NO)"] = footing_map.get(str(col_no), "NO")

        print("âœ… Added Footing Member (YES/NO).")

    except Exception as e:
        print(f"âš ï¸ Failed to merge footing members: {e}")


# ============================================================
# MERGE COLUMN HEIGHT FROM EXCEL (Column ID == Column No)
# ============================================================
# try:
#     print("ðŸ“‚ Reading columns_with_beam_clear_height.xlsx ...")
#     df_h = pd.read_excel("columns_with_beam_clear_height.xlsx")
#     df_h.columns = [c.strip() for c in df_h.columns]
#     if "Column ID" not in df_h.columns or "Column_Height (mm)" not in df_h.columns:
#         raise ValueError("Input sheet must have 'Column ID' and 'Column_Height (mm)'.")
#     height_map = dict(zip(df_h["Column ID"].astype(str), df_h["Column_Height (mm)"]))
#     # attach to rows by Column No
#     for col_no, data in columns.items():
#         h = height_map.get(str(col_no))
#         data["Column_Height (mm)"] = round(float(h), 2) if pd.notna(h) else ""
#     print("âœ… Added Column_Height (mm) from Excel.")
    
# except Exception as e:
#     print(f"âš ï¸ Skipped Column_Height merge: {e}")
# ============================================================
# MERGE COLUMN HEIGHT FROM EXCEL (Column ID == Column No)
# ============================================================
try:
    print("ðŸ“‚ Reading columns_with_beam_clear_height.xlsx ...")
    df_h = pd.read_excel("columns_with_beam_clear_height.xlsx")
    df_h.columns = [c.strip() for c in df_h.columns]

    if "Column ID" not in df_h.columns or "Column_Height (mm)" not in df_h.columns:
        raise ValueError("Input sheet must have 'Column ID' and 'Column_Height (mm)'.")

    df_h["Column ID"] = df_h["Column ID"].astype(str)
    height_map = dict(zip(df_h["Column ID"], df_h["Column_Height (mm)"]))

    for col_no, data in columns.items():
        h = height_map.get(str(col_no))
        data["Column_Height (mm)"] = round(float(h), 2) if pd.notna(h) else ""

    print("âœ… Added Column_Height (mm) from Excel.")

    # Store Excel IDs for later missing check
    excel_ids_for_check = set(df_h["Column ID"])

except Exception as e:
    print(f"âš ï¸ Skipped Column_Height merge: {e}")
    excel_ids_for_check = set()


# ============================================================
# MERGE COLUMN CLEAR HEIGHT FROM EXCEL (Column ID == Column No)
# ============================================================
try:
    print("ðŸ“‚ Reading columns_with_beam_clear_height.xlsx for Clear Height...")
    df_h = pd.read_excel("columns_with_beam_clear_height.xlsx")
    df_h.columns = [c.strip() for c in df_h.columns]
    if "Column ID" not in df_h.columns or "Column_Clear_Height (mm)" not in df_h.columns:
        raise ValueError("Input sheet must have 'Column ID' and 'Column_Clear_Height (mm)'.")
    clear_map = dict(zip(df_h["Column ID"].astype(str), df_h["Column_Clear_Height (mm)"]))
    for col_no, data in columns.items():
        h = clear_map.get(str(col_no))
        data["Column_Clear_Height (mm)"] = round(float(h), 2) if pd.notna(h) else ""
    print("âœ… Added Column_Clear_Height (mm) from Excel.")
    merge_footing_members(columns, "footing_members.xlsx")
except Exception as e:
    print(f"âš ï¸ Skipped Column_Clear_Height merge: {e}")

# ============================================================
# BAR SPACING LOGIC (Widthâ€“Depth coupled, multi-iteration)
# ============================================================
def _spacing_balance(width_len_mm: float, depth_len_mm: float, total_bars: int, bar_dia_mm: float):
    """
    Distribute bars equally to 4 edges after 4 corners.
    If either face (Width or Depth) fails 30 < s <= 300, iterate:
      - Move ONE intermediate bar per edge from failing face to the other face.
      - Stop if both faces OK, or when a face hits corners-only (bars_each_edge <= 2).
    Returns dict with initial and final spacings & flags.
    """
    out = {
        "sW0": "", "sD0": "", "sW": "", "sD": "",
        "okW": "NO", "okD": "NO", "recalc_used": "NO"
    }

    # guards
    try:
        total_bars = int(total_bars) if total_bars is not None else None
        bar_dia_mm = float(bar_dia_mm) if bar_dia_mm is not None else None
    except:
        total_bars = None
        bar_dia_mm = None

    if (not width_len_mm) or (not depth_len_mm) or (not total_bars) or (not bar_dia_mm) or total_bars < 4:
        return out

    effW = width_len_mm - DEDUCT
    effD = depth_len_mm - DEDUCT
    if effW <= 0 or effD <= 0:
        return out

    remain = max(total_bars - 4, 0)
    per_edge_intermediate = remain // 4
    # Start equal on both faces (each edge)
    barsW = per_edge_intermediate + 2
    barsD = per_edge_intermediate + 2
    barsW = max(barsW, 2)
    barsD = max(barsD, 2)

    def spacing(eff_len, bars_edge):
        spacing_count = max(bars_edge - 1, 1)
        used_bars_len = bars_edge * bar_dia_mm
        clear_len = eff_len - used_bars_len
        return clear_len / spacing_count

    def in_range(v): return (v > MIN_CLEAR) and (v <= MAX_SPACING)

    # initial spacings
    sW = spacing(effW, barsW)
    sD = spacing(effD, barsD)
    out["sW0"] = round(sW, 2)
    out["sD0"] = round(sD, 2)
    okW = in_range(sW)
    okD = in_range(sD)

    # iterate if any fail
    recalc = False
    iters = 0
    while (not (okW and okD)) and (barsW > 2 or barsD > 2):
        recalc = True
        iters += 1

        # decide direction based on which face fails and how
        if not okW and sW <= MIN_CLEAR and barsW > 2:
            # width too tight â†’ move one bar from Width to Depth
            barsW -= 1
            barsD += 1
        elif not okD and sD <= MIN_CLEAR and barsD > 2:
            # depth too tight â†’ move one bar from Depth to Width
            barsD -= 1
            barsW += 1
        elif not okW and sW > MAX_SPACING and barsD > 2:
            # width too wide â†’ move one bar from Depth to Width
            barsW += 1
            barsD -= 1
        elif not okD and sD > MAX_SPACING and barsW > 2:
            # depth too wide â†’ move one bar from Width to Depth
            barsD += 1
            barsW -= 1
        else:
            # cannot move further without breaking corners-only rule
            break

        sW = spacing(effW, barsW)
        sD = spacing(effD, barsD)
        okW = in_range(sW)
        okD = in_range(sD)

        if iters > 200:   # hard safety cap
            break

    out["sW"] = round(sW, 2)
    out["sD"] = round(sD, 2)
    out["okW"] = "YES" if okW else "NO"
    out["okD"] = "YES" if okD else "NO"
    out["recalc_used"] = "YES" if recalc else "NO"
    out["barsW_final"], out["barsD_final"] = barsW, barsD
    return out

# compute spacing + effective sizes for each column
# for col_no, data in columns.items():
#     try:
#         cw = float(data.get("Column Width (mm)", 0) or 0)
#         cd = float(data.get("Column Depth (mm)", 0) or 0)
#         nb = data.get("No of Bars")
#         bd = data.get("Bar Dia (mm)")

#         # Effective clear lengths (after cover+stirrups)
#         data["Eff Width (mm)"] = round(cw - DEDUCT, 2) if cw else ""
#         data["Eff Depth (mm)"] = round(cd - DEDUCT, 2) if cd else ""

#         res = _spacing_balance(cw, cd, int(nb) if nb else None, float(bd) if bd else None)

#         # short column names for spacing columns
#         data["sW0 (mm)"] = res["sW0"]
#         data["sD0 (mm)"] = res["sD0"]
#         data["sW (mm)"]  = res["sW"]
#         data["sD (mm)"]  = res["sD"]
#         data["barsW_final"] = res.get("barsW_final", "")
#         data["barsD_final"] = res.get("barsD_final", "")

#         # overall flags
#         data["Spacing OK"] = "YES" if (res["okW"] == "YES" and res["okD"] == "YES") else "NO"
#         data["Recalc Used"] = res["recalc_used"]
#         # =========================
#         # EXTRA LENGTH CALCULATIONS
#         # =========================
#         try:
#             col_height = float(data.get("Column_Height (mm)", 0) or 0)
#             no_bars = int(data.get("No of Bars") or 0)
#             bar_dia = float(data.get("Bar Dia (mm)") or 0)

#             # 2. Total longitudinal length
#             total_long_len = no_bars * col_height if no_bars and col_height else 0

#             # 3. Ld calculation
#             fy = 550
#             tau_bd = 1.5
#             factored_tau = tau_bd * 1.6
#             comp_factor = 1.25
#             Ld = (bar_dia * 0.87 * fy) / (4 * factored_tau * comp_factor) if bar_dia else 0

#             # 4. Splicing length per bar
#             splicing_len_each = max(Ld, 24 * bar_dia)

#             # 5. Total splicing length
#             total_splice_len = no_bars * splicing_len_each if no_bars else 0

#             # 6. Total length
#             total_len = total_long_len + total_splice_len

#             data["Ld (mm)"] = round(Ld, 2)
#             data["Splicing Length (mm)"] = round(splicing_len_each, 2)
#             data["Total Longitudinal Length (mm)"] = round(total_long_len, 2)
#             data["Total Splicing Length (mm)"] = round(total_splice_len, 2)
#             data["Total Length (mm)"] = round(total_len, 2)

#         except Exception:
#             data["Ld (mm)"] = data["Splicing Length (mm)"] = ""
#             data["Total Longitudinal Length (mm)"] = data["Total Splicing Length (mm)"] = data["Total Length (mm)"] = ""

#         # =========================
#         # CONFINEMENT ZONE CALCULATIONS
#         # =========================
#         try:
#             b = float(data.get("Column Width (mm)", 0) or 0)
#             D = float(data.get("Column Depth (mm)", 0) or 0)
#             bar_dia = float(data.get("Bar Dia (mm)", 0) or 0)
#             lclear = float(data.get("Column_Clear_Height (mm)", 0) or 0)

#             # (a) Confinement length (mm)
#             conf_len = max(max(b, D), lclear / 6.0, 450.0)

#             # (b) Confinement spacing (mm)
#             min_term = min(0.25 * min(b, D), 6 * bar_dia, 100.0)
#             conf_spacing = min_term

#             # (c) Number of stirrups
#             # n_stirrups = conf_len / conf_spacing if conf_spacing else 0
#             n_stirrups = math.ceil(conf_len / conf_spacing) + 1 if conf_spacing else 0

#             # (d) Length of one stirrup (centre-to-centre)
#             eff_b = b - 88
#             eff_D = D - 88
#             stirrup_len_each = 2 * (eff_b + eff_D)

#             # (e) Total length of stirrups (both faces)
#             total_stirrup_len = stirrup_len_each * n_stirrups * 2

#             data["Confinement Length (mm)"] = round(conf_len, 2)
#             data["Confinement Spacing (mm)"] = round(conf_spacing, 2)
#             data["No of Stirrups"] = round(n_stirrups, 2)
#             data["Length Each Stirrup (mm)"] = round(stirrup_len_each, 2)
#             data["Total Length Stirrups (mm)"] = round(total_stirrup_len, 2)

#         except Exception as e:
#             data["Confinement Length (mm)"] = ""
#             data["Confinement Spacing (mm)"] = ""
#             data["No of Stirrups"] = ""
#             data["Length Each Stirrup (mm)"] = ""
#             data["Total Length Stirrups (mm)"] = ""


#     except Exception:
#         data["Spacing OK"] = "NO"
#         data["Recalc Used"] = "YES"  # conservative

# compute spacing + effective sizes for each column
for col_no, data in columns.items():
    try:
        cw = float(data.get("Column Width (mm)", 0) or 0)
        cd = float(data.get("Column Depth (mm)", 0) or 0)
        nb = data.get("No of Bars")
        bd = data.get("Bar Dia (mm)")

        # Effective clear lengths (after cover+stirrups)
        data["Eff Width (mm)"] = round(cw - DEDUCT, 2) if cw else ""
        data["Eff Depth (mm)"] = round(cd - DEDUCT, 2) if cd else ""

        res = _spacing_balance(cw, cd, int(nb) if nb else None, float(bd) if bd else None)

        # short column names for spacing columns
        data["sW0 (mm)"] = res["sW0"]
        data["sD0 (mm)"] = res["sD0"]
        data["sW (mm)"]  = res["sW"]
        data["sD (mm)"]  = res["sD"]
        data["barsW_final"] = res.get("barsW_final", "")
        data["barsD_final"] = res.get("barsD_final", "")

        # overall flags
        data["Spacing OK"] = "YES" if (res["okW"] == "YES" and res["okD"] == "YES") else "NO"
        data["Recalc Used"] = res["recalc_used"]

        # =========================
        # EXTRA LENGTH CALCULATIONS
        # =========================
        try:
            col_height = float(data.get("Column_Height (mm)", 0) or 0)
            no_bars = int(data.get("No of Bars") or 0)
            bar_dia = float(data.get("Bar Dia (mm)") or 0)

            # 2. Total longitudinal length
            total_long_len = no_bars * col_height if no_bars and col_height else 0

            # 3. Ld calculation
            # fy = 550
            # tau_bd = 1.5
            # factored_tau = tau_bd * 1.6
            # comp_factor = 1.25
            # Ld = (bar_dia * 0.87 * fy) / (4 * factored_tau * comp_factor) if bar_dia else 0
            # ============================================================
            # DEVELOPMENT LENGTH (Ld) CALCULATION BASED ON fck
            # ============================================================
            fy = 550                     # Steel grade (MPa)
            fck = 30                     # Concrete grade (MPa) â€” you can make this configurable
            tau_bd_base = tau_bd_base_from_fck(fck)   # Get Ï„bd from IS 456 Table 21
            tau_bd = tau_bd_base         # Start from base Ï„bd
            factored_tau = tau_bd * 1.6  # For deformed bars (as per IS 456 Note)
            comp_factor = 1.25           # For compression members (if applicable)

            Ld = (bar_dia * 0.87 * fy) / (4 * factored_tau * comp_factor) if bar_dia else 0


            # 4. Splicing length per bar
            splicing_len_each = max(Ld, 24 * bar_dia)

            # 5. Total splicing length
            total_splice_len = no_bars * splicing_len_each if no_bars else 0

            

            Anchorage = round(Ld, 2)
            total_anchorage = Anchorage * no_bars

            # 6. Total length
            total_len = total_long_len + total_splice_len + total_anchorage

            data["Ld (mm)"] = round(Ld, 2)
            if data.get("Footing Member (YES/NO)") == "YES":
                data["Bottom Anchorage (mm)"] = round(Ld, 2)
            else:
                data["Bottom Anchorage (mm)"] = ""

            data["Splicing Length (mm)"] = round(splicing_len_each, 2)
            data["Splicing Length (mm)"] = round(splicing_len_each, 2)
            data["Total Longitudinal Length (mm)"] = round(total_long_len, 2)
            
            data["Total Splicing Length (mm)"] = round(total_splice_len, 2)
            data["total anchorage"]= total_anchorage
            data["Total Length (mm)"] = round(total_len, 2)
            data["Total Longitudinal kg "] = (0.006165 * bar_dia * bar_dia* round(total_len, 2))/1000
            ld_kg = (0.006165 * bar_dia * bar_dia* round(total_len, 2))/1000
            

        except Exception:
            data["Ld (mm)"] = data["Splicing Length (mm)"] = ""
            data["Total Longitudinal Length (mm)"] = data["Total Splicing Length (mm)"] = data["Total Length (mm)"] = ""

        # =========================
        # CONFINEMENT ZONE CALCULATIONS
        # =========================
        # try:
        #     import math
        #     b = float(data.get("Column Width (mm)", 0) or 0)
        #     D = float(data.get("Column Depth (mm)", 0) or 0)
        #     bar_dia = float(data.get("Bar Dia (mm)", 0) or 0)
        #     lclear = float(data.get("Column_Clear_Height (mm)", 0) or 0)

        #     conf_len = max(max(b, D), lclear / 6.0, 450.0)
        #     conf_spacing = min(0.25 * min(b, D), 6 * bar_dia, 100.0)
        #     n_stirrups = math.ceil(conf_len / conf_spacing) + 1 if conf_spacing else 0

        #     eff_b = b - 88
        #     eff_D = D - 88
        #     stirrup_len_each = 2 * (eff_b + eff_D)
        #     total_stirrup_len = stirrup_len_each * n_stirrups * 2

        #     data["Confinement Length"] = round(conf_len, 2)
        #     data["Confinement Spacing (mm)"] = round(conf_spacing, 2)
        #     data["No of End Stirrups"] = int(n_stirrups)
        #     data["Length Each Stirrup (mm)"] = round(stirrup_len_each, 2)
        #     data["Total Length End Stirrups (mm)"] = round(total_stirrup_len, 2)

        # except Exception as e:
        #     data["Confinement Length (mm)"] = ""
        #     data["Confinement Spacing (mm)"] = ""
        #     data["No of Stirrups"] = ""
        #     data["Length Each Stirrup (mm)"] = ""
        #     data["Total Length Stirrups (mm)"] = ""
        # =========================
        # CONFINEMENT ZONE + MID-STIRRUP CALCULATIONS
        # =========================
        try:
            import math

            # ---- Base parameters ----
            b = float(data.get("Column Width (mm)", 0) or 0)
            D = float(data.get("Column Depth (mm)", 0) or 0)
            bar_dia = float(data.get("Bar Dia (mm)", 0) or 0)
            lclear = float(data.get("Column_Height (mm)", 0) or 0)

            # ---- (A) Confinement Zone (End Regions) ----
            conf_len = max(max(b, D), lclear / 6.0, 450.0)
            conf_spacing = min(0.25 * min(b, D), 6 * bar_dia, 100.0)
            n_stirrups = math.ceil(conf_len / conf_spacing) + 1 if conf_spacing else 0

            # Effective dimensions for stirrup perimeter
            eff_b = b - 88   # 2*(cover + Â½*stirrup dia)
            eff_D = D - 88
            stirrup_len_each = 2 * (eff_b + eff_D)

            # Total stirrup length for both confinement zones (top + bottom)
            total_stirrup_len = stirrup_len_each * n_stirrups * 2
            hook_length1 = 10*10*2 * n_stirrups*2

            # Save confinement results
            data["Confinement Length"] = round(conf_len, 2) 
            data["Confinement Spacing (mm)"] = round(conf_spacing, 2)
            data["No of End Stirrups"] = int(n_stirrups)
            data["Length Each Stirrup (mm)"] = round(stirrup_len_each, 2)
            data["confinement hook_length (mm)"] = hook_length1
            data["Total Length End Stirrups (mm)"] = round(total_stirrup_len, 2) + hook_length1
            data["Total confinement kg "] = (0.006165 * 10 *10* (round(total_stirrup_len, 2) + hook_length1))/1000
            conf_kg = (0.006165 * 10 *10* (round(total_stirrup_len, 2) + hook_length1))/1000

            # ---- (B) Mid-Region Stirrup Calculations ----

            # 1ï¸âƒ£ Confinement Mid Length
            conf_mid_len = max(lclear - 2 * conf_len, 0.0)
            data["Confinement Mid Length (mm)"] = round(conf_mid_len, 2)
            hook_length2 = 10*8*2

            # 2ï¸âƒ£ Mid Spacing â‰¤ min{ min(b,D), 16Ï•, 300 }
            if bar_dia and b and D:
                mid_spacing = min(0.5 *min(b, D), 16 * bar_dia, 150.0)
            else:
                mid_spacing = 0.0
            data["Mid Spacing (mm)"] = round(mid_spacing, 2)

            # 3ï¸âƒ£ Number of Mid Stirrups
            n_mid_stirrups = math.ceil(conf_mid_len / mid_spacing) + 1 if mid_spacing else 0
            hook_length2 = hook_length2 *n_mid_stirrups
            data["No of Mid Stirrups"] = int(n_mid_stirrups)

            data["Mid stirrup hook_length (mm)"] = hook_length2

            # 4ï¸âƒ£ Total Length of Mid Stirrups
            total_mid_stirrup_len = n_mid_stirrups * stirrup_len_each
            data["Total Length Mid Stirrups (mm)"] = round(total_mid_stirrup_len, 2)+hook_length2
            data["Total mid stirrup kg "] = (0.006165 * 8 *8* (round(total_mid_stirrup_len, 2)+hook_length2))/1000
            mid_kg = (0.006165 * 8 *8* (round(total_mid_stirrup_len, 2)+hook_length2))/1000
            # ---- (C) Grand Total (End + Mid) ----
            total_all_stirrups = total_stirrup_len + total_mid_stirrup_len
            data["Total Length All Stirrups (mm)"] = round(total_all_stirrups, 2)
            data["Total stirrup kg "] = conf_kg + mid_kg + ld_kg

            

        except Exception as e:
            print(f"âš ï¸ Error computing confinement/mid-stirrups for column {col_no}: {e}")
            data["Confinement Length (mm)"] = data["Confinement Spacing (mm)"] = ""
            data["No of End Stirrups"] = data["Length Each Stirrup (mm)"] = ""
            data["Total Length End Stirrups (mm)"] = ""
            data["Confinement Mid Length (mm)"] = data["Mid Spacing (mm)"] = ""
            data["No of Mid Stirrups"] = data["Total Length Mid Stirrups (mm)"] = ""
            data["Total Length All Stirrups (mm)"] = ""

        

    except Exception:
        data["Spacing OK"] = "NO"
        data["Recalc Used"] = "YES"  # conservative

print("âœ… Added effective sizes, spacing, and confinement data.")
        

print("âœ… Added effective sizes, coupled spacing balance, checks, and recalc flags.")

# # ============================================================
# # WRITE TO EXCEL (UNCHANGED incl. Failures_Only and popup)
# # ============================================================
# wb = openpyxl.Workbook()
# ws = wb.active
# ws.title = "Column_Design_Data"

# if not columns:
#     print("âŒ No columns found!")
#     wb.save(OUTPUT_XLSX)
# else:
#     # headers = list(next(iter(columns.values())).keys())
#     # all_keys = list(next(iter(columns.values())).keys())
#     # headers = [h for h in all_keys if h != "Recalc Used"]
#     all_keys = list(next(iter(columns.values())).keys())

# # Force Footing Member column to appear after Column No
#     if "Footing Member (YES/NO)" in all_keys:
#         all_keys.remove("Footing Member (YES/NO)")
#         insert_pos = all_keys.index("Column No") + 1
#         all_keys.insert(insert_pos, "Footing Member (YES/NO)")

#     headers = [h for h in all_keys if h != "Recalc Used"]

#     ws.append(headers)  
#     for c in columns.values():
#         ws.append([c.get(h, "") for h in headers])

#     # column widths
#     for col_idx, header in enumerate(headers, start=1):
#         ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(str(header)) + 4, 16), 40)

#     # header style
#     header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
#     header_font = Font(color="FFFFFF", bold=True)
#     for cell in ws[1]:
#         cell.fill = header_fill
#         cell.font = header_font

#     # fail highlighting
#     red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
#     bold_red = Font(color="9C0006", bold=True)

#     check_cols = [
#         "Area Check",
#         "Min Steel Ratio (>0.8)",
#         "Max Steel Ratio (<4)",
#         "Min 4 Bars (>=4)",
#         "Min Bar Dia (>=12)",
#         "Spacing OK",
#     ]
#     header_to_index = {h: i for i, h in enumerate(headers)}
#     fail_rows = []
#     for r in range(2, ws.max_row + 1):
#         row_failed = False
#         for col_name in check_cols:
#             idx = header_to_index.get(col_name)
#             if idx is None:
#                 continue
#             cell = ws.cell(row=r, column=idx + 1)
#             if str(cell.value).strip().upper() == "NO":
#                 cell.fill = red_fill
#                 cell.font = bold_red
#                 row_failed = True
#         if row_failed:
#             fail_rows.append(r)

#     total = len(columns)
#     n_fail = len(fail_rows)
#     n_pass = total - n_fail

#     ws.append([])
#     ws.append(["Summary", f"Total = {total}", f"Pass = {n_pass}", f"Fail = {n_fail}"])
#     for c in ws[ws.max_row]:
#         c.font = Font(bold=True)

#     # Failures-only sheet
#     ws_fail = wb.create_sheet("Failures_Only")
#     ws_fail.append(headers + ["Fail Flags"])
#     for row_idx in fail_rows:
#         values = [ws.cell(row=row_idx, column=j + 1).value for j in range(len(headers))]
#         flags = []
#         for col_name in check_cols:
#             idx = header_to_index.get(col_name)
#             if idx is None:
#                 continue
#             v = ws.cell(row=row_idx, column=idx + 1).value
#             if str(v).strip().upper() == "NO":
#                 flags.append(col_name)
#         ws_fail.append(values + [", ".join(flags)])

#     for cell in ws_fail[1]:
#         cell.fill = header_fill
#         cell.font = header_font

#     wb.save(OUTPUT_XLSX)
#     print(f"âœ… Extracted {total} columns â†’ {OUTPUT_XLSX}  |  âœ“ Pass: {n_pass}  âœ– Fail: {n_fail}")
#     notify_summary(total, n_pass, n_fail, OUTPUT_XLSX, mode=POPUP_MODE, popup_seconds=POPUP_SECONDS)

# ============================================================
# WRITE TO EXCEL (UNCHANGED incl. Failures_Only and popup)
# ============================================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Column_Design_Data"

if not columns:
    print("âŒ No columns found!")
    wb.save(OUTPUT_XLSX)

else:
    # Extract all keys from first row
    all_keys = list(next(iter(columns.values())).keys())

    # âœ… FORCE Footing Member column right after Column No
    if "Footing Member (YES/NO)" in all_keys:
        all_keys.remove("Footing Member (YES/NO)")
        insert_pos = all_keys.index("Column No") + 1
        all_keys.insert(insert_pos, "Footing Member (YES/NO)")

    # Remove "Recalc Used" from output
    headers = [h for h in all_keys if h != "Recalc Used"]

    # Write header row
    ws.append(headers)

    # Write data rows
    for c in columns.values():
        ws.append([c.get(h, "") for h in headers])

    # Adjust column widths
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(str(header)) + 4, 16), 40)

    # Header styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Fail highlighting
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    bold_red = Font(color="9C0006", bold=True)

    check_cols = [
        "Area Check",
        "Min Steel Ratio (>0.8)",
        "Max Steel Ratio (<4)",
        "Min 4 Bars (>=4)",
        "Min Bar Dia (>=12)",
        "Spacing OK",
    ]

    header_to_index = {h: i for i, h in enumerate(headers)}
    fail_rows = []

    # Identify failing rows
    for r in range(2, ws.max_row + 1):
        row_failed = False
        for col_name in check_cols:
            idx = header_to_index.get(col_name)
            if idx is None:
                continue

            cell = ws.cell(row=r, column=idx + 1)
            if str(cell.value).strip().upper() == "NO":
                cell.fill = red_fill
                cell.font = bold_red
                row_failed = True

        if row_failed:
            fail_rows.append(r)

    total = len(columns)
    n_fail = len(fail_rows)
    n_pass = total - n_fail

    # Summary
    ws.append([])
    ws.append(["Summary", f"Total = {total}", f"Pass = {n_pass}", f"Fail = {n_fail}"])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)

    # ============================================================
    # FAILURES ONLY SHEET
    # ============================================================
    ws_fail = wb.create_sheet("Failures_Only")
    ws_fail.append(headers + ["Fail Flags"])

    for row_idx in fail_rows:
        values = [ws.cell(row=row_idx, column=j + 1).value for j in range(len(headers))]
        flags = []

        for col_name in check_cols:
            idx = header_to_index.get(col_name)
            if idx is None:
                continue

            v = ws.cell(row=row_idx, column=idx + 1).value
            if str(v).strip().upper() == "NO":
                flags.append(col_name)

        ws_fail.append(values + [", ".join(flags)])

    # Style header in failure sheet
    for cell in ws_fail[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Save workbook
    wb.save(OUTPUT_XLSX)
    # ============================================================
# CHECK MISSING COLUMN IDs (Excel vs Main Output)
# ============================================================
    try:
        output_excel = OUTPUT_XLSX
        df_output = pd.read_excel(output_excel, sheet_name=0)

        if "Column No" not in df_output.columns:
            raise ValueError("'Column No' not found in output file.")

        output_ids = set(df_output["Column No"].astype(str))
        missing_in_output = sorted(excel_ids_for_check - output_ids, key=lambda x: int(x))

        if missing_in_output:
            # Extract details of missing columns from input Excel
            df_missing = df_h[df_h["Column ID"].astype(str).isin(missing_in_output)].copy()
            df_missing["Remark"] = "Missing in main output"
            with pd.ExcelWriter(output_excel, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                df_missing.to_excel(writer, sheet_name="Missing_Columns", index=False)
            print(f"âš ï¸ Found {len(missing_in_output)} missing Column IDs â†’ written to 'Missing_Columns' sheet.")
        else:
            print("âœ… No missing Column IDs found â€” all matched successfully.")

    except Exception as e:
        print(f"âš ï¸ Missing Columns check skipped: {e}")


    print(f"âœ… Extracted {total} columns â†’ {OUTPUT_XLSX}  |  âœ“ Pass: {n_pass}  âœ– Fail: {n_fail}")
    notify_summary(total, n_pass, n_fail, OUTPUT_XLSX, mode=POPUP_MODE, popup_seconds=POPUP_SECONDS)
