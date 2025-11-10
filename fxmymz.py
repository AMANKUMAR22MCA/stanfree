# import os
# import time
# import pythoncom
# from win32com.client import Dispatch, VARIANT
# import openpyxl
# import pandas as pd

# # ============================================================
# # SETTINGS
# # ============================================================
# BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# OUTPUT_XLSX = os.path.join(BASE_DIR, "llk.xlsx")
# INPUT_CSV = os.path.join(BASE_DIR, "column_actions_per_combo_allmembers.csv")

# UNIT_TO_MM = 25.4
# TOL = 1e-3

# # ============================================================
# # HELPERS
# # ============================================================
# def approx_eq(a, b, tol=TOL):
#     return abs(a - b) <= tol

# def dispid(obj, name):
#     try:
#         return obj._oleobj_.GetIDsOfNames(name)
#     except Exception:
#         return None

# def safe_save(wb_obj, path):
#     try:
#         wb_obj.save(path)
#     except PermissionError:
#         alt = path.replace(".xlsx", f"_backup_{int(time.time())}.xlsx")
#         print(f"⚠ File locked, saved as backup: {alt}")
#         wb_obj.save(alt)

# # ============================================================
# # LOAD CSV DATA
# # ============================================================
# print("\n📄 Reading CSV data for top column forces...")
# df_csv = pd.read_csv(INPUT_CSV)
# df_csv.columns = [c.strip() for c in df_csv.columns]

# # Ensure numeric types
# df_csv["member"] = df_csv["member"].astype(int)
# df_csv["Fx (kN)"] = pd.to_numeric(df_csv["Fx (kN)"], errors="coerce")
# df_csv["My (kN-m)"] = pd.to_numeric(df_csv["My (kN-m)"], errors="coerce")
# df_csv["Mz (kN-m)"] = pd.to_numeric(df_csv["Mz (kN-m)"], errors="coerce")

# # Filter bottom joint rows
# df_bottom = df_csv[df_csv["station_label"].str.upper() == "BOTTOM_JOINT_CENTRE"]

# # For each member, find row with max Fx
# df_top_forces = (
#     df_bottom.loc[df_bottom.groupby("member")["Fx (kN)"].idxmax(), ["member", "Fx (kN)", "My (kN-m)", "Mz (kN-m)"]]
#     .rename(columns={
#         "member": "Top Column ID",
#         "Fx (kN)": "Top Col Max Fx (kN)",
#         "My (kN-m)": "Top Col My (kN-m)",
#         "Mz (kN-m)": "Top Col Mz (kN-m)"
#     })
# )
# print(f"✅ Loaded {len(df_top_forces)} top column force entries.")

# # ============================================================
# # OPENSTAAD ATTACH
# # ============================================================
# print("\n🔗 Connecting to STAAD via OpenSTAAD...")
# pythoncom.CoInitialize()
# os_app = Dispatch("StaadPro.OpenSTAAD")
# geom = os_app.Geometry
# prop = getattr(os_app, "Property", None)
# print("✅ Connected to STAAD model.")

# # ============================================================
# # GEOMETRY HELPERS
# # ============================================================
# def get_member_incidence(member_no):
#     try:
#         dpid = geom._oleobj_.GetIDsOfNames("GetMemberIncidence")
#         n1 = VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)
#         n2 = VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)
#         geom._oleobj_.InvokeTypes(
#             dpid, 0, pythoncom.DISPATCH_METHOD, (pythoncom.VT_EMPTY, 0),
#             ((pythoncom.VT_I4, 0),
#              (pythoncom.VT_VARIANT, 1),
#              (pythoncom.VT_VARIANT, 1)),
#             int(member_no), n1, n2
#         )
#         return int(n1.value), int(n2.value)
#     except Exception:
#         return None, None

# def get_node_coordinates_mm(node_id):
#     try:
#         dpid = geom._oleobj_.GetIDsOfNames("GetNodeCoordinates")
#         x = VARIANT(pythoncom.VT_BYREF | pythoncom.VT_R8, 0.0)
#         y = VARIANT(pythoncom.VT_BYREF | pythoncom.VT_R8, 0.0)
#         z = VARIANT(pythoncom.VT_BYREF | pythoncom.VT_R8, 0.0)
#         geom._oleobj_.InvokeTypes(
#             dpid, 0, pythoncom.DISPATCH_METHOD, (pythoncom.VT_EMPTY, 0),
#             ((pythoncom.VT_I4, 0),
#              (pythoncom.VT_VARIANT, 1),
#              (pythoncom.VT_VARIANT, 1),
#              (pythoncom.VT_VARIANT, 1)),
#             int(node_id), x, y, z
#         )
#         return float(x.value) * UNIT_TO_MM, float(y.value) * UNIT_TO_MM, float(z.value) * UNIT_TO_MM
#     except Exception:
#         return None, None, None

# # ============================================================
# # STEP 1: FETCH ALL MEMBERS
# # ============================================================
# total_members = geom.GetMemberCount
# print(f"📦 Total members found: {total_members}")

# member_data = {}

# for m in range(1, int(total_members) + 1):
#     n1, n2 = get_member_incidence(m)
#     if not n1 or not n2:
#         continue
#     x1, y1, z1 = get_node_coordinates_mm(n1)
#     x2, y2, z2 = get_node_coordinates_mm(n2)
#     if None in (x1, y1, z1, x2, y2, z2):
#         continue
#     member_data[m] = {
#         "n1": n1, "n2": n2,
#         "x1": x1, "y1": y1, "z1": z1,
#         "x2": x2, "y2": y2, "z2": z2
#     }

# # ============================================================
# # STEP 2: FILTER ONLY COLUMNS (vertical members)
# # ============================================================
# columns = {}
# for mid, info in member_data.items():
#     if approx_eq(info["x1"], info["x2"]) and approx_eq(info["z1"], info["z2"]) and not approx_eq(info["y1"], info["y2"]):
#         columns[mid] = info

# print(f"🧱 Total column members detected: {len(columns)}")

# # ============================================================
# # STEP 3: BUILD NODE-WISE MAPPING (only nodes part of columns)
# # ============================================================
# results = []
# node_map = {}

# for col_id, info in columns.items():
#     n1, n2 = info["n1"], info["n2"]
#     y1, y2 = info["y1"], info["y2"]

#     # Determine top & bottom node based on Y
#     if y1 > y2:
#         top_node, bottom_node = n1, n2
#     else:
#         top_node, bottom_node = n2, n1

#     # For bottom node: record bottom column id
#     if bottom_node not in node_map:
#         node_map[bottom_node] = {"node_id": bottom_node, "top_column_id": 0, "bottom_column_id": col_id}
#     else:
#         node_map[bottom_node]["bottom_column_id"] = col_id

#     # For top node: record top column id
#     if top_node not in node_map:
#         node_map[top_node] = {"node_id": top_node, "top_column_id": col_id, "bottom_column_id": 0}
#     else:
#         node_map[top_node]["top_column_id"] = col_id

# # Convert node_map to results list
# for nid, info in node_map.items():
#     top_col = info["top_column_id"]
#     bottom_col = info["bottom_column_id"]
    

#     # Lookup top column data from CSV
#     fx = my = mz = None
#     if top_col in df_top_forces["Top Column ID"].values:
#         rec = df_top_forces[df_top_forces["Top Column ID"] == top_col].iloc[0]
#         fx = rec["Top Col Max Fx (kN)"]
#         my = rec["Top Col My (kN-m)"]
#         mz = rec["Top Col Mz (kN-m)"]

#     results.append([
#         info["node_id"],
#         top_col,
#         bottom_col,
#         fx,
#         my,
#         mz
#     ])

# # ============================================================
# # STEP 4: WRITE TO EXCEL
# # ============================================================
# wb = openpyxl.Workbook()
# ws = wb.active
# ws.title = "Column_Node_Map"

# headers = [
#     "Node ID","Top Column ID" , "Bottom Column ID",
#     "Top Col Max Fx (kN)", "Top Col My (kN-m)", "Top Col Mz (kN-m)"
# ]
# ws.append(headers)

# for row in results:
#     ws.append(row)

# safe_save(wb, OUTPUT_XLSX)
# print(f"✅ Column-node mapping with top-column forces saved to: {OUTPUT_XLSX}")


# import os
# import time
# import pythoncom
# from win32com.client import Dispatch, VARIANT
# import openpyxl
# import pandas as pd

# # ============================================================
# # SETTINGS
# # ============================================================
# BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# OUTPUT_XLSX = os.path.join(BASE_DIR, "llk_fixed.xlsx")
# INPUT_CSV = os.path.join(BASE_DIR, "column_actions_per_combo_allmembers.csv")

# UNIT_TO_MM = 25.4
# TOL = 1e-3

# # ============================================================
# # HELPERS
# # ============================================================
# def approx_eq(a, b, tol=TOL):
#     return abs(a - b) <= tol

# def safe_save(wb_obj, path):
#     try:
#         wb_obj.save(path)
#     except PermissionError:
#         alt = path.replace(".xlsx", f"_backup_{int(time.time())}.xlsx")
#         print(f"⚠ File locked, saved as backup: {alt}")
#         wb_obj.save(alt)

# # ============================================================
# # LOAD CSV DATA
# # ============================================================
# print("\n📄 Reading CSV data for top column forces...")
# df_csv = pd.read_csv(INPUT_CSV)
# df_csv.columns = [c.strip() for c in df_csv.columns]

# # Ensure numeric types
# df_csv["member"] = df_csv["member"].astype(int)
# df_csv["Fx (kN)"] = pd.to_numeric(df_csv["Fx (kN)"], errors="coerce")
# df_csv["My (kN-m)"] = pd.to_numeric(df_csv["My (kN-m)"], errors="coerce")
# df_csv["Mz (kN-m)"] = pd.to_numeric(df_csv["Mz (kN-m)"], errors="coerce")

# # Filter only bottom joint rows
# df_bottom = df_csv[df_csv["station_label"].str.upper() == "BOTTOM_JOINT_CENTRE"]

# # For each member, find row with max Fx
# df_top_forces = (
#     df_bottom.loc[df_bottom.groupby("member")["Fx (kN)"].idxmax(), ["member", "Fx (kN)", "My (kN-m)", "Mz (kN-m)"]]
#     .rename(columns={
#         "member": "Top Column ID",
#         "Fx (kN)": "Top Col Max Fx (kN)",
#         "My (kN-m)": "Top Col My (kN-m)",
#         "Mz (kN-m)": "Top Col Mz (kN-m)"
#     })
# )
# print(f"✅ Loaded {len(df_top_forces)} top column force entries.")

# # ============================================================
# # OPENSTAAD ATTACH
# # ============================================================
# print("\n🔗 Connecting to STAAD via OpenSTAAD...")
# pythoncom.CoInitialize()
# os_app = Dispatch("StaadPro.OpenSTAAD")
# geom = os_app.Geometry
# print("✅ Connected to STAAD model.")

# # ============================================================
# # GEOMETRY HELPERS
# # ============================================================
# def get_member_incidence(member_no):
#     try:
#         dpid = geom._oleobj_.GetIDsOfNames("GetMemberIncidence")
#         n1 = VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)
#         n2 = VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)
#         geom._oleobj_.InvokeTypes(
#             dpid, 0, pythoncom.DISPATCH_METHOD, (pythoncom.VT_EMPTY, 0),
#             ((pythoncom.VT_I4, 0),
#              (pythoncom.VT_VARIANT, 1),
#              (pythoncom.VT_VARIANT, 1)),
#             int(member_no), n1, n2
#         )
#         return int(n1.value), int(n2.value)
#     except Exception:
#         return None, None

# def get_node_coordinates_mm(node_id):
#     try:
#         dpid = geom._oleobj_.GetIDsOfNames("GetNodeCoordinates")
#         x = VARIANT(pythoncom.VT_BYREF | pythoncom.VT_R8, 0.0)
#         y = VARIANT(pythoncom.VT_BYREF | pythoncom.VT_R8, 0.0)
#         z = VARIANT(pythoncom.VT_BYREF | pythoncom.VT_R8, 0.0)
#         geom._oleobj_.InvokeTypes(
#             dpid, 0, pythoncom.DISPATCH_METHOD, (pythoncom.VT_EMPTY, 0),
#             ((pythoncom.VT_I4, 0),
#              (pythoncom.VT_VARIANT, 1),
#              (pythoncom.VT_VARIANT, 1),
#              (pythoncom.VT_VARIANT, 1)),
#             int(node_id), x, y, z
#         )
#         return float(x.value) * UNIT_TO_MM, float(y.value) * UNIT_TO_MM, float(z.value) * UNIT_TO_MM
#     except Exception:
#         return None, None, None

# # ============================================================
# # STEP 1: FETCH ALL MEMBERS
# # ============================================================
# total_members = geom.GetMemberCount
# print(f"📦 Total members found: {total_members}")

# member_data = {}

# for m in range(1, int(total_members) + 1):
#     n1, n2 = get_member_incidence(m)
#     if not n1 or not n2:
#         continue
#     x1, y1, z1 = get_node_coordinates_mm(n1)
#     x2, y2, z2 = get_node_coordinates_mm(n2)
#     if None in (x1, y1, z1, x2, y2, z2):
#         continue
#     member_data[m] = {
#         "n1": n1, "n2": n2,
#         "x1": x1, "y1": y1, "z1": z1,
#         "x2": x2, "y2": y2, "z2": z2
#     }

# # ============================================================
# # STEP 2: FILTER ONLY COLUMNS (vertical members)
# # ============================================================
# columns = {}
# for mid, info in member_data.items():
#     if approx_eq(info["x1"], info["x2"]) and approx_eq(info["z1"], info["z2"]) and not approx_eq(info["y1"], info["y2"]):
#         columns[mid] = info

# print(f"🧱 Total column members detected: {len(columns)}")

# # ============================================================
# # STEP 3: BUILD NODE-WISE MAPPING (reversed logic)
# # ============================================================
# results = []
# node_map = {}

# for col_id, info in columns.items():
#     n1, n2 = info["n1"], info["n2"]
#     y1, y2 = info["y1"], info["y2"]

#     # Reverse logic: higher Y = bottom column, lower Y = top column
#     if y1 > y2:
#         bottom_node, top_node = n1, n2
#     else:
#         bottom_node, top_node = n2, n1

#     # Record top/bottom columns correctly
#     if bottom_node not in node_map:
#         node_map[bottom_node] = {"node_id": bottom_node, "top_column_id": 0, "bottom_column_id": col_id}
#     else:
#         node_map[bottom_node]["bottom_column_id"] = col_id

#     if top_node not in node_map:
#         node_map[top_node] = {"node_id": top_node, "top_column_id": col_id, "bottom_column_id": 0}
#     else:
#         node_map[top_node]["top_column_id"] = col_id

# # ============================================================
# # STEP 4: LOOKUP TOP COLUMN FORCES (use correct top member)
# # ============================================================
# for nid, info in node_map.items():
#     top_col = info["top_column_id"]
#     fx = my = mz = None

#     if top_col in df_top_forces["Top Column ID"].values:
#         rec = df_top_forces[df_top_forces["Top Column ID"] == top_col].iloc[0]
#         fx = rec["Top Col Max Fx (kN)"]
#         my = rec["Top Col My (kN-m)"]
#         mz = rec["Top Col Mz (kN-m)"]

#     info["fx"] = fx
#     info["my"] = my
#     info["mz"] = mz

# # ============================================================
# # STEP 5: WRITE TO EXCEL
# # ============================================================
# wb = openpyxl.Workbook()
# ws = wb.active
# ws.title = "Column_Node_Map"

# headers = [
#     "Node ID", "Top Column ID", "Bottom Column ID",
#     "Top Col Max Fx (kN)", "Top Col My (kN-m)", "Top Col Mz (kN-m)"
# ]
# ws.append(headers)

# for nid, info in node_map.items():
#     ws.append([
#         info["node_id"],
#         info["top_column_id"],
#         info["bottom_column_id"],
#         info["fx"],
#         info["my"],
#         info["mz"]
#     ])

# safe_save(wb, OUTPUT_XLSX)
# print(f"✅ Fixed output saved: {OUTPUT_XLSX}")




import os
import time
import pythoncom
from win32com.client import Dispatch, VARIANT
import openpyxl
import pandas as pd

# ============================================================
# SETTINGS
# ============================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_XLSX = os.path.join(BASE_DIR, "llk_fixed_with_y_zplane.xlsx")
INPUT_CSV = os.path.join(BASE_DIR, "column_actions_per_combo_allmembers.csv")

UNIT_TO_MM = 25.4
TOL = 1e-3

# ============================================================
# HELPERS
# ============================================================
def approx_eq(a, b, tol=TOL):
    return abs(a - b) <= tol

def safe_save(wb_obj, path):
    try:
        wb_obj.save(path)
    except PermissionError:
        alt = path.replace(".xlsx", f"_backup_{int(time.time())}.xlsx")
        print(f"⚠ File locked, saved as backup: {alt}")
        wb_obj.save(alt)

# ============================================================
# LOAD CSV DATA
# ============================================================
print("\n📄 Reading CSV data for top/bottom column forces...")
df_csv = pd.read_csv(INPUT_CSV)
df_csv.columns = [c.strip() for c in df_csv.columns]

# Ensure numeric types
df_csv["member"] = df_csv["member"].astype(int)
df_csv["Fx (kN)"] = pd.to_numeric(df_csv["Fx (kN)"], errors="coerce")
df_csv["My (kN-m)"] = pd.to_numeric(df_csv["My (kN-m)"], errors="coerce")
df_csv["Mz (kN-m)"] = pd.to_numeric(df_csv["Mz (kN-m)"], errors="coerce")

# -------------------------------
# TOP COLUMN FORCES  (BOTTOM_JOINT_CENTRE)
# -------------------------------
df_top_label = df_csv[df_csv["station_label"].str.upper() == "BOTTOM_JOINT_CENTRE"]
df_top_forces = (
    df_top_label.loc[df_top_label.groupby("member")["Fx (kN)"].idxmax(),
                     ["member", "Fx (kN)", "My (kN-m)", "Mz (kN-m)"]]
    .rename(columns={
        "member": "Top Column ID",
        "Fx (kN)": "Top Col Max Fx (kN)",
        "My (kN-m)": "Top Col My (kN-m)",
        "Mz (kN-m)": "Top Col Mz (kN-m)"
    })
)
print(f"✅ Loaded {len(df_top_forces)} top column force entries (BOTTOM_JOINT_CENTRE).")

# -------------------------------
# BOTTOM COLUMN FORCES (TOP_FACE_AT_JOINT)
# -------------------------------
df_bottom_label = df_csv[df_csv["station_label"].str.upper() == "TOP_FACE_AT_JOINT"]
df_bottom_forces = (
    df_bottom_label.loc[df_bottom_label.groupby("member")["Fx (kN)"].idxmax(),
                        ["member", "Fx (kN)", "My (kN-m)", "Mz (kN-m)"]]
    .rename(columns={
        "member": "Bottom Column ID",
        "Fx (kN)": "Bottom Col Max Fx (kN)",
        "My (kN-m)": "Bottom Col My (kN-m)",
        "Mz (kN-m)": "Bottom Col Mz (kN-m)"
    })
)
print(f"✅ Loaded {len(df_bottom_forces)} bottom column force entries (TOP_FACE_AT_JOINT).")

# ============================================================
# OPENSTAAD ATTACH
# ============================================================
print("\n🔗 Connecting to STAAD via OpenSTAAD...")
pythoncom.CoInitialize()
os_app = Dispatch("StaadPro.OpenSTAAD")
geom = os_app.Geometry
print("✅ Connected to STAAD model.")

# ============================================================
# GEOMETRY HELPERS
# ============================================================
def get_member_incidence(member_no):
    try:
        dpid = geom._oleobj_.GetIDsOfNames("GetMemberIncidence")
        n1 = VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)
        n2 = VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)
        geom._oleobj_.InvokeTypes(
            dpid, 0, pythoncom.DISPATCH_METHOD, (pythoncom.VT_EMPTY, 0),
            ((pythoncom.VT_I4, 0),
             (pythoncom.VT_VARIANT, 1),
             (pythoncom.VT_VARIANT, 1)),
            int(member_no), n1, n2
        )
        return int(n1.value), int(n2.value)
    except Exception:
        return None, None

def get_node_coordinates_mm(node_id):
    try:
        dpid = geom._oleobj_.GetIDsOfNames("GetNodeCoordinates")
        x = VARIANT(pythoncom.VT_BYREF | pythoncom.VT_R8, 0.0)
        y = VARIANT(pythoncom.VT_BYREF | pythoncom.VT_R8, 0.0)
        z = VARIANT(pythoncom.VT_BYREF | pythoncom.VT_R8, 0.0)
        geom._oleobj_.InvokeTypes(
            dpid, 0, pythoncom.DISPATCH_METHOD, (pythoncom.VT_EMPTY, 0),
            ((pythoncom.VT_I4, 0),
             (pythoncom.VT_VARIANT, 1),
             (pythoncom.VT_VARIANT, 1),
             (pythoncom.VT_VARIANT, 1)),
            int(node_id), x, y, z
        )
        return float(x.value) * UNIT_TO_MM, float(y.value) * UNIT_TO_MM, float(z.value) * UNIT_TO_MM
    except Exception:
        return None, None, None

# ============================================================
# STEP 1: FETCH ALL MEMBERS
# ============================================================
total_members = geom.GetMemberCount
print(f"📦 Total members found: {total_members}")

member_data = {}
for m in range(1, int(total_members) + 1):
    n1, n2 = get_member_incidence(m)
    if not n1 or not n2:
        continue
    x1, y1, z1 = get_node_coordinates_mm(n1)
    x2, y2, z2 = get_node_coordinates_mm(n2)
    if None in (x1, y1, z1, x2, y2, z2):
        continue
    member_data[m] = {
        "n1": n1, "n2": n2,
        "x1": x1, "y1": y1, "z1": z1,
        "x2": x2, "y2": y2, "z2": z2
    }

# ============================================================
# STEP 2: FILTER ONLY COLUMNS (vertical members)
# ============================================================
columns = {}
for mid, info in member_data.items():
    if approx_eq(info["x1"], info["x2"]) and approx_eq(info["z1"], info["z2"]) and not approx_eq(info["y1"], info["y2"]):
        columns[mid] = info
print(f"🧱 Total column members detected: {len(columns)}")

# ============================================================
# STEP 3: BUILD NODE-WISE MAPPING
# ============================================================
node_map = {}
for col_id, info in columns.items():
    n1, n2 = info["n1"], info["n2"]
    y1, y2 = info["y1"], info["y2"]

    # Reverse logic: higher Y = bottom, lower Y = top
    if y1 > y2:
        bottom_node, top_node = n1, n2
    else:
        bottom_node, top_node = n2, n1

    # bottom_node
    if bottom_node not in node_map:
        node_map[bottom_node] = {"node_id": bottom_node, "top_column_id": 0, "bottom_column_id": col_id}
    else:
        node_map[bottom_node]["bottom_column_id"] = col_id

    # top_node
    if top_node not in node_map:
        node_map[top_node] = {"node_id": top_node, "top_column_id": col_id, "bottom_column_id": 0}
    else:
        node_map[top_node]["top_column_id"] = col_id

# ============================================================
# STEP 4: LOOKUP FORCES FOR BOTH COLUMNS
# ============================================================
for nid, info in node_map.items():
    top_col = info["top_column_id"]
    bottom_col = info["bottom_column_id"]

    fx_t = my_t = mz_t = None
    fx_b = my_b = mz_b = None

    # --- Top column from df_top_forces ---
    if top_col in df_top_forces["Top Column ID"].values:
        rec_t = df_top_forces[df_top_forces["Top Column ID"] == top_col].iloc[0]
        fx_t, my_t, mz_t = rec_t["Top Col Max Fx (kN)"], rec_t["Top Col My (kN-m)"], rec_t["Top Col Mz (kN-m)"]

    # --- Bottom column from df_bottom_forces ---
    if bottom_col in df_bottom_forces["Bottom Column ID"].values:
        rec_b = df_bottom_forces[df_bottom_forces["Bottom Column ID"] == bottom_col].iloc[0]
        fx_b, my_b, mz_b = rec_b["Bottom Col Max Fx (kN)"], rec_b["Bottom Col My (kN-m)"], rec_b["Bottom Col Mz (kN-m)"]

    info.update({
        "Top Fx": fx_t, "Top My": my_t, "Top Mz": mz_t,
        "Bottom Fx": fx_b, "Bottom My": my_b, "Bottom Mz": mz_b
    })

# # ============================================================
# # STEP 4B: LOAD COLUMN DESIGN & GEOMETRY DATA AND COMPUTE VALUES
# # ============================================================
# import math

# # ---- 1️⃣ Read column design data (for Bar Dia & No of Bars)
# DESIGN_XLSX = os.path.join(BASE_DIR, "column_design_data_checked.xlsx")
# print("\n📘 Reading column design data...")
# df_design = pd.read_excel(DESIGN_XLSX)
# df_design.columns = [c.strip() for c in df_design.columns]

# df_design["Column No"] = pd.to_numeric(df_design["Column No"], errors="coerce")
# df_design["Bar Dia (mm)"] = pd.to_numeric(df_design["Bar Dia (mm)"], errors="coerce")
# df_design["No of Bars"] = pd.to_numeric(df_design["No of Bars"], errors="coerce")

# design_lookup = {
#     int(row["Column No"]): {
#         "Bar Dia (mm)": row["Bar Dia (mm)"],
#         "No of Bars": row["No of Bars"]
#     }
#     for _, row in df_design.iterrows() if not pd.isna(row["Column No"])
# }

# # ---- 2️⃣ Read geometry data (Depth, Width, Beta)
# GEOM_XLSX = os.path.join(BASE_DIR, "columns_with_beam_clear_height.xlsx")
# print("\n📗 Reading column geometry data...")
# df_geom = pd.read_excel(GEOM_XLSX)
# df_geom.columns = [c.strip() for c in df_geom.columns]

# df_geom["Column ID"] = pd.to_numeric(df_geom["Column ID"], errors="coerce")
# df_geom["Column Depth (mm)"] = pd.to_numeric(df_geom["Column Depth (mm)"], errors="coerce")
# df_geom["Column Width (mm)"] = pd.to_numeric(df_geom["Column Width (mm)"], errors="coerce")
# df_geom["Column Beta (deg)"] = pd.to_numeric(df_geom["Column Beta (deg)"], errors="coerce")

# geom_lookup = {
#     int(row["Column ID"]): {
#         "Column Depth (mm)": row["Column Depth (mm)"],
#         "Column Width (mm)": row["Column Width (mm)"],
#         "Column Beta (deg)": row["Column Beta (deg)"]
#     }
#     for _, row in df_geom.iterrows() if not pd.isna(row["Column ID"])
# }

# # ---- 3️⃣ Perform computations for each node
# for nid, info in node_map.items():
#     bottom_col = info.get("bottom_column_id", 0)

#     # --- Get data from both sources
#     design_data = design_lookup.get(bottom_col, {})
#     geom_data = geom_lookup.get(bottom_col, {})

#     bar_dia = design_data.get("Bar Dia (mm)")
#     no_of_bars = design_data.get("No of Bars")
#     col_depth = geom_data.get("Column Depth (mm)")
#     col_width = geom_data.get("Column Width (mm)")
#     col_beta = geom_data.get("Column Beta (deg)")

#     # --- Compute Bottom Y
#     if bar_dia is not None and not pd.isna(bar_dia):
#         bottom_y = 40 + 10 + (bar_dia / 2.0)
#     else:
#         bottom_y = None

#     # --- Compute geometric properties if available
#     bot_col_A = bot_col_W = bot_col_sigma_plus = None
#     if col_depth and col_width and not pd.isna(col_depth) and not pd.isna(col_width):
#         bot_col_b = col_width
#         bot_col_D = col_depth
#         bot_col_A = bot_col_b * bot_col_D
#         bot_col_W = bot_col_b * ((bot_col_D * bot_col_D) / 6.0)

#         fx_b = info.get("Bottom Fx")
#         my_b = info.get("Bottom My")
#         mz_b = info.get("Bottom Mz")

#         if fx_b and bot_col_A and bot_col_W:
#             if col_beta == 0:
#                 bot_col_sigma_plus = (fx_b / bot_col_A) + (mz_b / bot_col_W if mz_b else 0)
#             elif col_beta == 90:
#                 bot_col_sigma_plus = (fx_b / bot_col_A) + (my_b / bot_col_W if my_b else 0)
#             else:
#                 bot_col_sigma_plus = None

#     # --- Compute steel area
#     col_bot_as = None
#     if no_of_bars and bar_dia:
#         col_bot_as = no_of_bars * (math.pi * (bar_dia ** 2) / 4.0)

#     # --- Store computed values
#     info.update({
#         "Bar Dia (mm)": bar_dia,
#         "No of Bars": no_of_bars,
#         "Bottom Y (mm)": bottom_y,
#         "Column Depth (mm)": col_depth,
#         "Column Width (mm)": col_width,
#         "Column Beta (deg)": col_beta,
#         "bot_col_A": bot_col_A,
#         "bot_col_W": bot_col_W,
#         "bot_col_sigma_plus": bot_col_sigma_plus,
#         "col_bot_as": col_bot_as
#     })

# print("✅ Computed Bottom Y, bot_col_sigma_plus, and col_bot_as for all matching columns.")

# # ============================================================
# # STEP 5: WRITE TO EXCEL
# # ============================================================
# wb = openpyxl.Workbook()
# ws = wb.active
# ws.title = "Column_Node_Map"

# headers = [
#     "Node ID", "Top Column ID", "Bottom Column ID",
#     "Top Col Max Fx (kN)", "Top Col My (kN-m)", "Top Col Mz (kN-m)",
#     "Bottom Col Max Fx (kN)", "Bottom Col My (kN-m)", "Bottom Col Mz (kN-m)",
#     "Bar Dia (mm)", "No of Bars", "Bottom Y (mm)",
#     "Column Depth (mm)", "Column Width (mm)", "Column Beta (deg)",
#     "bot_col_A", "bot_col_W", "bot_col_sigma_plus", "col_bot_as"
# ]
# ws.append(headers)

# for nid, info in node_map.items():
#     ws.append([
#         info["node_id"],
#         info["top_column_id"],
#         info["bottom_column_id"],
#         info["Top Fx"],
#         info["Top My"],
#         info["Top Mz"],
#         info["Bottom Fx"],
#         info["Bottom My"],
#         info["Bottom Mz"],
#         info.get("Bar Dia (mm)"),
#         info.get("No of Bars"),
#         info.get("Bottom Y (mm)"),
#         info.get("Column Depth (mm)"),
#         info.get("Column Width (mm)"),
#         info.get("Column Beta (deg)"),
#         info.get("bot_col_A"),
#         info.get("bot_col_W"),
#         info.get("bot_col_sigma_plus"),
#         info.get("col_bot_as")
#     ])

# safe_save(wb, OUTPUT_XLSX)
# print(f"✅ Output with computed column geometry and stresses saved: {OUTPUT_XLSX}")


# # ============================================================
# # STEP 4B: LOAD COLUMN DESIGN & GEOMETRY DATA AND COMPUTE VALUES
# # (updated: includes solving for neutral axis x and computing cc, fs1, fs2, ac, a1, a2, mc1)
# # ============================================================
# import math

# # Material / constants
# fck = 30.0        # N/mm^2
# es = 2.0e5        # N/mm^2
# fy = 550.0        # N/mm^2

# # ---- 1️⃣ Read column design data (for Bar Dia & No of Bars)
# DESIGN_XLSX = os.path.join(BASE_DIR, "column_design_data_checked.xlsx")
# print("\n📘 Reading column design data...")
# df_design = pd.read_excel(DESIGN_XLSX)
# df_design.columns = [c.strip() for c in df_design.columns]

# df_design["Column No"] = pd.to_numeric(df_design["Column No"], errors="coerce")
# df_design["Bar Dia (mm)"] = pd.to_numeric(df_design["Bar Dia (mm)"], errors="coerce")
# df_design["No of Bars"] = pd.to_numeric(df_design["No of Bars"], errors="coerce")

# design_lookup = {
#     int(row["Column No"]): {
#         "Bar Dia (mm)": row["Bar Dia (mm)"],
#         "No of Bars": row["No of Bars"]
#     }
#     for _, row in df_design.iterrows() if not pd.isna(row["Column No"])
# }

# # ---- 2️⃣ Read geometry data (Depth, Width, Beta)
# GEOM_XLSX = os.path.join(BASE_DIR, "columns_with_beam_clear_height.xlsx")
# print("\n📗 Reading column geometry data...")
# df_geom = pd.read_excel(GEOM_XLSX)
# df_geom.columns = [c.strip() for c in df_geom.columns]

# df_geom["Column ID"] = pd.to_numeric(df_geom["Column ID"], errors="coerce")
# df_geom["Column Depth (mm)"] = pd.to_numeric(df_geom["Column Depth (mm)"], errors="coerce")
# df_geom["Column Width (mm)"] = pd.to_numeric(df_geom["Column Width (mm)"], errors="coerce")
# df_geom["Column Beta (deg)"] = pd.to_numeric(df_geom["Column Beta (deg)"], errors="coerce")

# geom_lookup = {
#     int(row["Column ID"]): {
#         "Column Depth (mm)": row["Column Depth (mm)"],
#         "Column Width (mm)": row["Column Width (mm)"],
#         "Column Beta (deg)": row["Column Beta (deg)"]
#     }
#     for _, row in df_geom.iterrows() if not pd.isna(row["Column ID"])
# }

# # helper: compute cc, fs1, fs2 in kN for a given x (mm)
# def compute_cc_fs(F_bkN, b_mm, D_mm, bottom_y_mm, As_mm2, x_mm):
#     # cc in kN
#     cc_N = 0.36 * fck * b_mm * x_mm         # N
#     cc_kN = cc_N / 1000.0

#     # s1 stress (N/mm2)
#     # guard against x_mm == 0
#     if x_mm <= 1e-12:
#         s1 = 0.0
#         s2 = 0.0
#     else:
#         s1_raw = es * 0.0035 * (1.0 - (bottom_y_mm / x_mm))
#         s1 = min(s1_raw, 0.87 * fy)

#         s2_raw = es * 0.0035 * (1.0 - ((D_mm - bottom_y_mm) / x_mm))
#         s2 = min(s2_raw, 0.87 * fy)

#     # fs in kN: As (mm2) * stress (N/mm2) = N -> /1000 -> kN
#     fs1_kN = (As_mm2 * s1) / 1000.0
#     fs2_kN = (As_mm2 * s2) / 1000.0

#     return cc_kN, fs1_kN, fs2_kN

# # solver using bisection for root of residual(x) = cc + fs1 + fs2 - F_bkN
# def solve_x_for_equilibrium(F_bkN, b_mm, D_mm, bottom_y_mm, As_mm2):
#     # Residual function
#     def residual(x):
#         cc_kN, fs1_kN, fs2_kN = compute_cc_fs(F_bkN, b_mm, D_mm, bottom_y_mm, As_mm2, x)
#         return (cc_kN + fs1_kN + fs2_kN) - F_bkN

#     # quick checks
#     if As_mm2 is None or b_mm is None or D_mm is None or bottom_y_mm is None:
#         return None

#     # bracket search
#     low = 1e-6
#     high = max(D_mm * 2.0, 1.0)  # try up to twice depth initially

#     rlow = residual(low)
#     rhigh = residual(high)

#     # expand high until sign change or until a reasonable limit
#     max_expand = 10
#     expand_count = 0
#     while rlow * rhigh > 0 and expand_count < max_expand:
#         high *= 2.0
#         rhigh = residual(high)
#         expand_count += 1

#     if rlow * rhigh > 0:
#         # cannot find sign change -> return None (no root found in bracket)
#         return None

#     # bisection
#     tol = 1e-6
#     max_iter = 100
#     for _ in range(max_iter):
#         mid = 0.5 * (low + high)
#         rmid = residual(mid)
#         if abs(rmid) <= tol:
#             return mid
#         if rlow * rmid <= 0:
#             high = mid
#             rhigh = rmid
#         else:
#             low = mid
#             rlow = rmid
#     # return mid as approximation
#     return mid

# # ---- 3️⃣ Perform computations for each node (now with x, cc, fs1, fs2, ac, a1, a2, mc1)
# for nid, info in node_map.items():
#     bottom_col = info.get("bottom_column_id", 0)

#     # --- Get data from both sources
#     design_data = design_lookup.get(bottom_col, {})
#     geom_data = geom_lookup.get(bottom_col, {})

#     bar_dia = design_data.get("Bar Dia (mm)")
#     no_of_bars = design_data.get("No of Bars")
#     col_depth = geom_data.get("Column Depth (mm)")
#     col_width = geom_data.get("Column Width (mm)")
#     col_beta = geom_data.get("Column Beta (deg)")

#     # --- Compute Bottom Y
#     if bar_dia is not None and not pd.isna(bar_dia):
#         bottom_y = 40 + 10 + (bar_dia / 2.0)
#     else:
#         bottom_y = None

#     # --- Compute geometric properties if available
#     bot_col_A = bot_col_W = bot_col_sigma_plus = None
#     if col_depth and col_width and not pd.isna(col_depth) and not pd.isna(col_width):
#         bot_col_b = col_width
#         bot_col_D = col_depth
#         bot_col_A = bot_col_b * bot_col_D
#         bot_col_W = bot_col_b * ((bot_col_D * bot_col_D) / 6.0)

#         fx_b = info.get("Bottom Fx")
#         my_b = info.get("Bottom My")
#         mz_b = info.get("Bottom Mz")

#         if fx_b is not None and bot_col_A and bot_col_W:
#             # Note: fx_b is in kN (can be negative)
#             if col_beta == 0:
#                 bot_col_sigma_plus = (fx_b / bot_col_A) + ((mz_b or 0.0) / bot_col_W)
#             elif col_beta == 90:
#                 bot_col_sigma_plus = (fx_b / bot_col_A) + ((my_b or 0.0) / bot_col_W)
#             else:
#                 bot_col_sigma_plus = None

#     # --- Compute steel area (mm2)
#     col_bot_as = None
#     if no_of_bars and bar_dia:
#         col_bot_as = no_of_bars * (math.pi * (bar_dia ** 2) / 4.0)

#     # --- Solve for x using equilibrium if Bottom Fx and geometry present
#     x_mm = None
#     cc_kN = fs1_kN = fs2_kN = None
#     ac = a1 = a2 = mc1 = None

#     fx_bkN = info.get("Bottom Fx")  # already in kN
#     if fx_bkN is not None and col_width and col_depth and bottom_y is not None and col_bot_as is not None:
#         # call solver
#         x_mm = solve_x_for_equilibrium(fx_bkN, col_width, col_depth, bottom_y, col_bot_as)

#         if x_mm is not None:
#             cc_kN, fs1_kN, fs2_kN = compute_cc_fs(fx_bkN, col_width, col_depth, bottom_y, col_bot_as, x_mm)
#             # compute ac, a1, a2, mc1
#             ac = 0.42 * x_mm - (col_depth / 2.0)
#             a1 = bottom_y - (col_depth / 2.0)
#             a2 = (col_depth - bottom_y) - (col_depth / 2.0)  # = col_depth/2 - bottom_y
#             mc1 = (cc_kN * ac) + (fs1_kN * a1) + (fs2_kN * a2)  # units: kN * mm

#     # --- Store computed values
#     info.update({
#         "Bar Dia (mm)": bar_dia,
#         "No of Bars": no_of_bars,
#         "Bottom Y (mm)": bottom_y,
#         "Column Depth (mm)": col_depth,
#         "Column Width (mm)": col_width,
#         "Column Beta (deg)": col_beta,
#         "bot_col_A": bot_col_A,
#         "bot_col_W": bot_col_W,
#         "bot_col_sigma_plus": bot_col_sigma_plus,
#         "col_bot_as": col_bot_as,
#         # new fields
#         "x_mm": x_mm,
#         "cc_kN": cc_kN,
#         "fs1_kN": fs1_kN,
#         "fs2_kN": fs2_kN,
#         "ac_mm": ac,
#         "a1_mm": a1,
#         "a2_mm": a2,
#         "mc1_kNmm": mc1
#     })

# print("✅ Computed Bottom Y, bot_col_sigma_plus, col_bot_as and reinforced equilibrium results for all matching columns.")
# # ============================================================
# # STEP 5: WRITE TO EXCEL (updated headers include new fields)
# # ============================================================
# wb = openpyxl.Workbook()
# ws = wb.active
# ws.title = "Column_Node_Map"

# headers = [
#     "Node ID", "Top Column ID", "Bottom Column ID",
#     "Top Col Max Fx (kN)", "Top Col My (kN-m)", "Top Col Mz (kN-m)",
#     "Bottom Col Max Fx (kN)", "Bottom Col My (kN-m)", "Bottom Col Mz (kN-m)",
#     "Bar Dia (mm)", "No of Bars", "Bottom Y (mm)",
#     "Column Depth (mm)", "Column Width (mm)", "Column Beta (deg)",
#     "bot_col_A", "bot_col_W", "bot_col_sigma_plus", "col_bot_as",
#     # new fields
#     "x_mm", "cc_kN", "fs1_kN", "fs2_kN", "ac_mm", "a1_mm", "a2_mm", "mc1_kNmm"
# ]
# ws.append(headers)

# for nid, info in node_map.items():
#     ws.append([
#         info["node_id"],
#         info["top_column_id"],
#         info["bottom_column_id"],
#         info["Top Fx"],
#         info["Top My"],
#         info["Top Mz"],
#         info["Bottom Fx"],
#         info["Bottom My"],
#         info["Bottom Mz"],
#         info.get("Bar Dia (mm)"),
#         info.get("No of Bars"),
#         info.get("Bottom Y (mm)"),
#         info.get("Column Depth (mm)"),
#         info.get("Column Width (mm)"),
#         info.get("Column Beta (deg)"),
#         info.get("bot_col_A"),
#         info.get("bot_col_W"),
#         info.get("bot_col_sigma_plus"),
#         info.get("col_bot_as"),
#         info.get("x_mm"),
#         info.get("cc_kN"),
#         info.get("fs1_kN"),
#         info.get("fs2_kN"),
#         info.get("ac_mm"),
#         info.get("a1_mm"),
#         info.get("a2_mm"),
#         info.get("mc1_kNmm")
#     ])

# safe_save(wb, OUTPUT_XLSX)
# print(f"✅ Output with computed column geometry, stresses and equilibrium results saved: {OUTPUT_XLSX}")


# ============================================================
# STEP 4B: LOAD COLUMN DESIGN & GEOMETRY DATA AND COMPUTE VALUES
# (updated: includes solving for neutral axis x and computing cc, fs1, fs2, ac, a1, a2, mc1)
# ============================================================
# import math

# # Material / constants
# fck = 30.0        # N/mm^2
# es = 2.0e5        # N/mm^2
# fy = 550.0        # N/mm^2

# # ---- 1️⃣ Read column design data (for Bar Dia & No of Bars)
# DESIGN_XLSX = os.path.join(BASE_DIR, "column_design_data_checked.xlsx")
# print("\n📘 Reading column design data...")
# df_design = pd.read_excel(DESIGN_XLSX)
# df_design.columns = [c.strip() for c in df_design.columns]

# df_design["Column No"] = pd.to_numeric(df_design["Column No"], errors="coerce")
# df_design["Bar Dia (mm)"] = pd.to_numeric(df_design["Bar Dia (mm)"], errors="coerce")
# df_design["No of Bars"] = pd.to_numeric(df_design["No of Bars"], errors="coerce")

# design_lookup = {
#     int(row["Column No"]): {
#         "Bar Dia (mm)": row["Bar Dia (mm)"],
#         "No of Bars": row["No of Bars"]
#     }
#     for _, row in df_design.iterrows() if not pd.isna(row["Column No"])
# }

# # ---- 2️⃣ Read geometry data (Depth, Width, Beta)
# GEOM_XLSX = os.path.join(BASE_DIR, "columns_with_beam_clear_height.xlsx")
# print("\n📗 Reading column geometry data...")
# df_geom = pd.read_excel(GEOM_XLSX)
# df_geom.columns = [c.strip() for c in df_geom.columns]

# df_geom["Column ID"] = pd.to_numeric(df_geom["Column ID"], errors="coerce")
# df_geom["Column Depth (mm)"] = pd.to_numeric(df_geom["Column Depth (mm)"], errors="coerce")
# df_geom["Column Width (mm)"] = pd.to_numeric(df_geom["Column Width (mm)"], errors="coerce")
# df_geom["Column Beta (deg)"] = pd.to_numeric(df_geom["Column Beta (deg)"], errors="coerce")

# geom_lookup = {
#     int(row["Column ID"]): {
#         "Column Depth (mm)": row["Column Depth (mm)"],
#         "Column Width (mm)": row["Column Width (mm)"],
#         "Column Beta (deg)": row["Column Beta (deg)"]
#     }
#     for _, row in df_geom.iterrows() if not pd.isna(row["Column ID"])
# }

# # helper: compute cc, fs1, fs2 in kN for a given x (mm)
# def compute_cc_fs(F_bkN, b_mm, D_mm, bottom_y_mm, As_mm2, x_mm):
#     cc_N = 0.36 * fck * b_mm * x_mm
#     cc_kN = cc_N / 1000.0

#     if x_mm <= 1e-12:
#         s1 = s2 = 0.0
#     else:
#         s1_raw = es * 0.0035 * (1.0 - (bottom_y_mm / x_mm))
#         s1 = min(s1_raw, 0.87 * fy)

#         s2_raw = es * 0.0035 * (1.0 - ((D_mm - bottom_y_mm) / x_mm))
#         s2 = min(s2_raw, 0.87 * fy)

#     fs1_kN = (As_mm2 * s1) / 1000.0
#     fs2_kN = (As_mm2 * s2) / 1000.0

#     return cc_kN, fs1_kN, fs2_kN

# # solver using bisection for root of residual(x) = cc + fs1 + fs2 - F_bkN
# def solve_x_for_equilibrium(F_bkN, b_mm, D_mm, bottom_y_mm, As_mm2):
#     def residual(x):
#         cc_kN, fs1_kN, fs2_kN = compute_cc_fs(F_bkN, b_mm, D_mm, bottom_y_mm, As_mm2, x)
#         return (cc_kN + fs1_kN + fs2_kN) - F_bkN

#     if As_mm2 is None or b_mm is None or D_mm is None or bottom_y_mm is None:
#         return None

#     low = 1e-6
#     high = max(D_mm * 2.0, 1.0)

#     rlow = residual(low)
#     rhigh = residual(high)

#     max_expand = 10
#     expand_count = 0
#     while rlow * rhigh > 0 and expand_count < max_expand:
#         high *= 2.0
#         rhigh = residual(high)
#         expand_count += 1

#     if rlow * rhigh > 0:
#         return None

#     tol = 1e-6
#     max_iter = 100
#     for _ in range(max_iter):
#         mid = 0.5 * (low + high)
#         rmid = residual(mid)
#         if abs(rmid) <= tol:
#             return mid
#         if rlow * rmid <= 0:
#             high = mid
#             rhigh = rmid
#         else:
#             low = mid
#             rlow = rmid
#     return mid


# # ---- 3️⃣ Perform computations for each node (both top & bottom)
# for nid, info in node_map.items():
#     # -----------------------------
#     # 🔹 Bottom column
#     # -----------------------------
#     bottom_col = info.get("bottom_column_id", 0)
#     design_data_bot = design_lookup.get(bottom_col, {})
#     geom_data_bot = geom_lookup.get(bottom_col, {})

#     bot_bar_dia = design_data_bot.get("Bar Dia (mm)")
#     bot_no_of_bars = design_data_bot.get("No of Bars")
#     bot_depth = geom_data_bot.get("Column Depth (mm)")
#     bot_width = geom_data_bot.get("Column Width (mm)")
#     bot_beta = geom_data_bot.get("Column Beta (deg)")

#     bot_bottom_y = 40 + 10 + (bot_bar_dia / 2.0) if bot_bar_dia else None
#     bot_A = bot_W = bot_sigma_plus = None

#     if bot_depth and bot_width:
#         bot_A = bot_width * bot_depth
#         bot_W = bot_width * (bot_depth ** 2) / 6.0
#         fx_b = info.get("Bottom Fx")
#         my_b = info.get("Bottom My")
#         mz_b = info.get("Bottom Mz")
#         if fx_b is not None:
#             if bot_beta == 0:
#                 bot_sigma_plus = (fx_b / bot_A) + ((mz_b or 0.0) / bot_W)
#             elif bot_beta == 90:
#                 bot_sigma_plus = (fx_b / bot_A) + ((my_b or 0.0) / bot_W)

#     bot_As = None
#     if bot_no_of_bars and bot_bar_dia:
#         bot_As = bot_no_of_bars * (math.pi * (bot_bar_dia ** 2) / 4.0)

#     bot_x = bot_cc = bot_fs1 = bot_fs2 = bot_ac = bot_a1 = bot_a2 = bot_mc1 = None
#     fx_bkN = info.get("Bottom Fx")
#     if fx_bkN is not None and bot_width and bot_depth and bot_bottom_y and bot_As:
#         bot_x = solve_x_for_equilibrium(fx_bkN, bot_width, bot_depth, bot_bottom_y, bot_As)
#         if bot_x:
#             bot_cc, bot_fs1, bot_fs2 = compute_cc_fs(fx_bkN, bot_width, bot_depth, bot_bottom_y, bot_As, bot_x)
#             bot_ac = 0.42 * bot_x - (bot_depth / 2.0)
#             bot_a1 = bot_bottom_y - (bot_depth / 2.0)
#             bot_a2 = (bot_depth - bot_bottom_y) - (bot_depth / 2.0)
#             bot_mc1 = (bot_cc * bot_ac) + (bot_fs1 * bot_a1) + (bot_fs2 * bot_a2)

#     # -----------------------------
#     # 🔹 Top column
#     # -----------------------------
#     top_col = info.get("top_column_id", 0)
#     design_data_top = design_lookup.get(top_col, {})
#     geom_data_top = geom_lookup.get(top_col, {})

#     top_bar_dia = design_data_top.get("Bar Dia (mm)")
#     top_no_of_bars = design_data_top.get("No of Bars")
#     top_depth = geom_data_top.get("Column Depth (mm)")
#     top_width = geom_data_top.get("Column Width (mm)")
#     top_beta = geom_data_top.get("Column Beta (deg)")

#     top_bottom_y = 40 + 10 + (top_bar_dia / 2.0) if top_bar_dia else None
#     top_A = top_W = top_sigma_plus = None

#     if top_depth and top_width:
#         top_A = top_width * top_depth
#         top_W = top_width * (top_depth ** 2) / 6.0
#         fx_t = info.get("Top Fx")
#         my_t = info.get("Top My")
#         mz_t = info.get("Top Mz")
#         if fx_t is not None:
#             if top_beta == 0:
#                 top_sigma_plus = (fx_t / top_A) + ((mz_t or 0.0) / top_W)
#             elif top_beta == 90:
#                 top_sigma_plus = (fx_t / top_A) + ((my_t or 0.0) / top_W)

#     top_As = None
#     if top_no_of_bars and top_bar_dia:
#         top_As = top_no_of_bars * (math.pi * (top_bar_dia ** 2) / 4.0)

#     top_x = top_cc = top_fs1 = top_fs2 = top_ac = top_a1 = top_a2 = top_mc1 = None
#     fx_tkN = info.get("Top Fx")
#     if fx_tkN is not None and top_width and top_depth and top_bottom_y and top_As:
#         top_x = solve_x_for_equilibrium(fx_tkN, top_width, top_depth, top_bottom_y, top_As)
#         if top_x:
#             top_cc, top_fs1, top_fs2 = compute_cc_fs(fx_tkN, top_width, top_depth, top_bottom_y, top_As, top_x)
#             top_ac = 0.42 * top_x - (top_depth / 2.0)
#             top_a1 = top_bottom_y - (top_depth / 2.0)
#             top_a2 = (top_depth - top_bottom_y) - (top_depth / 2.0)
#             top_mc1 = (top_cc * top_ac) + (top_fs1 * top_a1) + (top_fs2 * top_a2)

#     # Store all results
#     info.update({
#         # bottom
#         "bot_bar_dia": bot_bar_dia, "bot_no_of_bars": bot_no_of_bars, "bot_bottom_y": bot_bottom_y,
#         "bot_depth": bot_depth, "bot_width": bot_width, "bot_beta": bot_beta,
#         "bot_A": bot_A, "bot_W": bot_W, "bot_sigma_plus": bot_sigma_plus, "bot_As": bot_As,
#         "bot_x": bot_x, "bot_cc": bot_cc, "bot_fs1": bot_fs1, "bot_fs2": bot_fs2,
#         "bot_ac": bot_ac, "bot_a1": bot_a1, "bot_a2": bot_a2, "bot_mc1": bot_mc1,
#         # top
#         "top_bar_dia": top_bar_dia, "top_no_of_bars": top_no_of_bars, "top_bottom_y": top_bottom_y,
#         "top_depth": top_depth, "top_width": top_width, "top_beta": top_beta,
#         "top_A": top_A, "top_W": top_W, "top_sigma_plus": top_sigma_plus, "top_As": top_As,
#         "top_x": top_x, "top_cc": top_cc, "top_fs1": top_fs1, "top_fs2": top_fs2,
#         "top_ac": top_ac, "top_a1": top_a1, "top_a2": top_a2, "top_mc1": top_mc1
#     })

# print("✅ Computed top and bottom column equilibrium results for all matching nodes.")

# # ============================================================
# # STEP 5: WRITE TO EXCEL (headers include both top and bottom)
# # ============================================================
# wb = openpyxl.Workbook()
# ws = wb.active
# ws.title = "Column_Node_Map"

# headers = [
#     "Node ID", "Top Column ID", "Bottom Column ID",
#     "Top Col Max Fx (kN)", "Top Col My (kN-m)", "Top Col Mz (kN-m)",
#     "Bottom Col Max Fx (kN)", "Bottom Col My (kN-m)", "Bottom Col Mz (kN-m)",
#     # ---- TOP COLUMN ----
#     "top_bar_dia", "top_no_of_bars", "top_bottom_y", "top_depth", "top_width", "top_beta",
#     "top_A", "top_W", "top_sigma_plus", "top_As", "top_x", "top_cc", "top_fs1", "top_fs2",
#     "top_ac", "top_a1", "top_a2", "top_mc1",
#     # ---- BOTTOM COLUMN ----
#     "bot_bar_dia", "bot_no_of_bars", "bot_bottom_y", "bot_depth", "bot_width", "bot_beta",
#     "bot_A", "bot_W", "bot_sigma_plus", "bot_As", "bot_x", "bot_cc", "bot_fs1", "bot_fs2",
#     "bot_ac", "bot_a1", "bot_a2", "bot_mc1"
# ]
# ws.append(headers)

# for nid, info in node_map.items():
#     ws.append([
#         info["node_id"], info["top_column_id"], info["bottom_column_id"],
#         info["Top Fx"], info["Top My"], info["Top Mz"],
#         info["Bottom Fx"], info["Bottom My"], info["Bottom Mz"],
#         info.get("top_bar_dia"), info.get("top_no_of_bars"), info.get("top_bottom_y"),
#         info.get("top_depth"), info.get("top_width"), info.get("top_beta"),
#         info.get("top_A"), info.get("top_W"), info.get("top_sigma_plus"),
#         info.get("top_As"), info.get("top_x"), info.get("top_cc"), info.get("top_fs1"),
#         info.get("top_fs2"), info.get("top_ac"), info.get("top_a1"), info.get("top_a2"),
#         info.get("top_mc1"),
#         info.get("bot_bar_dia"), info.get("bot_no_of_bars"), info.get("bot_bottom_y"),
#         info.get("bot_depth"), info.get("bot_width"), info.get("bot_beta"),
#         info.get("bot_A"), info.get("bot_W"), info.get("bot_sigma_plus"),
#         info.get("bot_As"), info.get("bot_x"), info.get("bot_cc"), info.get("bot_fs1"),
#         info.get("bot_fs2"), info.get("bot_ac"), info.get("bot_a1"), info.get("bot_a2"),
#         info.get("bot_mc1")
#     ])

# safe_save(wb, OUTPUT_XLSX)
# print(f"✅ Output with computed top and bottom column equilibrium results saved: {OUTPUT_XLSX}")


# ============================================================
# STEP 4B: LOAD COLUMN DESIGN & GEOMETRY DATA AND COMPUTE VALUES
# (updated: includes solving for neutral axis x and computing cc, fs1, fs2, ac, a1, a2, mc1)
# Added: Z-PLANE calculations for both top & bottom (fields suffixed with _z)
# ============================================================
import math

# Material / constants
fck = 30.0        # N/mm^2
es = 2.0e5        # N/mm^2
fy = 550.0        # N/mm^2

# ---- 1️⃣ Read column design data (for Bar Dia & No of Bars)
DESIGN_XLSX = os.path.join(BASE_DIR, "column_design_data_checked.xlsx")
print("\n📘 Reading column design data...")
df_design = pd.read_excel(DESIGN_XLSX)
df_design.columns = [c.strip() for c in df_design.columns]

df_design["Column No"] = pd.to_numeric(df_design["Column No"], errors="coerce")
df_design["Bar Dia (mm)"] = pd.to_numeric(df_design["Bar Dia (mm)"], errors="coerce")
df_design["No of Bars"] = pd.to_numeric(df_design["No of Bars"], errors="coerce")

design_lookup = {
    int(row["Column No"]): {
        "Bar Dia (mm)": row["Bar Dia (mm)"],
        "No of Bars": row["No of Bars"]
    }
    for _, row in df_design.iterrows() if not pd.isna(row["Column No"])
}

# ---- 2️⃣ Read geometry data (Depth, Width, Beta)
GEOM_XLSX = os.path.join(BASE_DIR, "columns_with_beam_clear_height.xlsx")
print("\n📗 Reading column geometry data...")
df_geom = pd.read_excel(GEOM_XLSX)
df_geom.columns = [c.strip() for c in df_geom.columns]

df_geom["Column ID"] = pd.to_numeric(df_geom["Column ID"], errors="coerce")
df_geom["Column Depth (mm)"] = pd.to_numeric(df_geom["Column Depth (mm)"], errors="coerce")
df_geom["Column Width (mm)"] = pd.to_numeric(df_geom["Column Width (mm)"], errors="coerce")
df_geom["Column Beta (deg)"] = pd.to_numeric(df_geom["Column Beta (deg)"], errors="coerce")

geom_lookup = {
    int(row["Column ID"]): {
        "Column Depth (mm)": row["Column Depth (mm)"],
        "Column Width (mm)": row["Column Width (mm)"],
        "Column Beta (deg)": row["Column Beta (deg)"]
    }
    for _, row in df_geom.iterrows() if not pd.isna(row["Column ID"])
}

# helper: compute cc, fs1, fs2 in kN for a given x (mm)
def compute_cc_fs(F_bkN, b_mm, D_mm, bottom_y_mm, As_mm2, x_mm):
    cc_N = 0.36 * fck * b_mm * x_mm
    cc_kN = cc_N / 1000.0

    if x_mm <= 1e-12:
        s1 = s2 = 0.0
    else:
        s1_raw = es * 0.0035 * (1.0 - (bottom_y_mm / x_mm))
        s1 = min(s1_raw, 0.87 * fy)

        s2_raw = es * 0.0035 * (1.0 - ((D_mm - bottom_y_mm) / x_mm))
        s2 = min(s2_raw, 0.87 * fy)

    fs1_kN = (As_mm2 * s1) / 1000.0
    fs2_kN = (As_mm2 * s2) / 1000.0

    return cc_kN, fs1_kN, fs2_kN

# solver using bisection for root of residual(x) = cc + fs1 + fs2 - F_bkN
def solve_x_for_equilibrium(F_bkN, b_mm, D_mm, bottom_y_mm, As_mm2):
    def residual(x):
        cc_kN, fs1_kN, fs2_kN = compute_cc_fs(F_bkN, b_mm, D_mm, bottom_y_mm, As_mm2, x)
        return (cc_kN + fs1_kN + fs2_kN) - F_bkN

    if As_mm2 is None or b_mm is None or D_mm is None or bottom_y_mm is None:
        return None

    low = 1e-6
    high = max(D_mm * 2.0, 1.0)

    rlow = residual(low)
    rhigh = residual(high)

    max_expand = 10
    expand_count = 0
    while rlow * rhigh > 0 and expand_count < max_expand:
        high *= 2.0
        rhigh = residual(high)
        expand_count += 1

    if rlow * rhigh > 0:
        return None

    tol = 1e-6
    max_iter = 100
    for _ in range(max_iter):
        mid = 0.5 * (low + high)
        rmid = residual(mid)
        if abs(rmid) <= tol:
            return mid
        if rlow * rmid <= 0:
            high = mid
            rhigh = rmid
        else:
            low = mid
            rlow = rmid
    return mid


# ---- 3️⃣ Perform computations for each node (both top & bottom)
for nid, info in node_map.items():
    # -----------------------------
    # 🔹 Bottom column (Y-plane)  - existing calculations (unchanged)
    # -----------------------------
    bottom_col = info.get("bottom_column_id", 0)
    design_data_bot = design_lookup.get(bottom_col, {})
    geom_data_bot = geom_lookup.get(bottom_col, {})

    bot_bar_dia = design_data_bot.get("Bar Dia (mm)")
    bot_no_of_bars = design_data_bot.get("No of Bars")
    bot_depth = geom_data_bot.get("Column Depth (mm)")
    bot_width = geom_data_bot.get("Column Width (mm)")
    bot_beta = geom_data_bot.get("Column Beta (deg)")

    bot_bottom_y = 40 + 10 + (bot_bar_dia / 2.0) if bot_bar_dia else None
    bot_A = bot_W = bot_sigma_plus = None

    if bot_depth and bot_width:
        bot_A = bot_width * bot_depth
        bot_W = bot_width * (bot_depth ** 2) / 6.0
        fx_b = info.get("Bottom Fx")
        my_b = info.get("Bottom My")
        mz_b = info.get("Bottom Mz")
        if fx_b is not None:
            if bot_beta == 0:
                bot_sigma_plus = (fx_b / bot_A) + ((mz_b or 0.0) / bot_W)
            elif bot_beta == 90:
                bot_sigma_plus = (fx_b / bot_A) + ((my_b or 0.0) / bot_W)

    bot_As = None
    if bot_no_of_bars and bot_bar_dia:
        bot_As = bot_no_of_bars * (math.pi * (bot_bar_dia ** 2) / 4.0)

    bot_x = bot_cc = bot_fs1 = bot_fs2 = bot_ac = bot_a1 = bot_a2 = bot_mc1 = None
    fx_bkN = info.get("Bottom Fx")
    if fx_bkN is not None and bot_width and bot_depth and bot_bottom_y and bot_As:
        bot_x = solve_x_for_equilibrium(fx_bkN, bot_width, bot_depth, bot_bottom_y, bot_As)
        if bot_x:
            bot_cc, bot_fs1, bot_fs2 = compute_cc_fs(fx_bkN, bot_width, bot_depth, bot_bottom_y, bot_As, bot_x)
            bot_ac = 0.42 * bot_x - (bot_depth / 2.0)
            bot_a1 = bot_bottom_y - (bot_depth / 2.0)
            bot_a2 = (bot_depth - bot_bottom_y) - (bot_depth / 2.0)
            bot_mc1 = (bot_cc * bot_ac) + (bot_fs1 * bot_a1) + (bot_fs2 * bot_a2)

    # -----------------------------
    # 🔹 Top column (Y-plane) - existing calculations (unchanged)
    # -----------------------------
    top_col = info.get("top_column_id", 0)
    design_data_top = design_lookup.get(top_col, {})
    geom_data_top = geom_lookup.get(top_col, {})

    top_bar_dia = design_data_top.get("Bar Dia (mm)")
    top_no_of_bars = design_data_top.get("No of Bars")
    top_depth = geom_data_top.get("Column Depth (mm)")
    top_width = geom_data_top.get("Column Width (mm)")
    top_beta = geom_data_top.get("Column Beta (deg)")

    top_bottom_y = 40 + 10 + (top_bar_dia / 2.0) if top_bar_dia else None
    top_A = top_W = top_sigma_plus = None

    if top_depth and top_width:
        top_A = top_width * top_depth
        top_W = top_width * (top_depth ** 2) / 6.0
        fx_t = info.get("Top Fx")
        my_t = info.get("Top My")
        mz_t = info.get("Top Mz")
        if fx_t is not None:
            if top_beta == 0:
                top_sigma_plus = (fx_t / top_A) + ((mz_t or 0.0) / top_W)
            elif top_beta == 90:
                top_sigma_plus = (fx_t / top_A) + ((my_t or 0.0) / top_W)

    top_As = None
    if top_no_of_bars and top_bar_dia:
        top_As = top_no_of_bars * (math.pi * (top_bar_dia ** 2) / 4.0)

    top_x = top_cc = top_fs1 = top_fs2 = top_ac = top_a1 = top_a2 = top_mc1 = None
    fx_tkN = info.get("Top Fx")
    if fx_tkN is not None and top_width and top_depth and top_bottom_y and top_As:
        top_x = solve_x_for_equilibrium(fx_tkN, top_width, top_depth, top_bottom_y, top_As)
        if top_x:
            top_cc, top_fs1, top_fs2 = compute_cc_fs(fx_tkN, top_width, top_depth, top_bottom_y, top_As, top_x)
            top_ac = 0.42 * top_x - (top_depth / 2.0)
            top_a1 = top_bottom_y - (top_depth / 2.0)
            top_a2 = (top_depth - top_bottom_y) - (top_depth / 2.0)
            top_mc1 = (top_cc * top_ac) + (top_fs1 * top_a1) + (top_fs2 * top_a2)

    # -----------------------------
    # === Z-PLANE CALCULATIONS ===
    # For Z-plane we swap the roles of width/depth when passing to the solver:
    #   - call solve_x_for_equilibrium( F, b_mm=depth, D_mm=width, bottom_y_mm, As )
    # This keeps function compute_cc_fs unchanged and consistently computes
    # cc/fs for the Z-plane orientation.
    # -----------------------------

    # ---- Bottom column Z-plane
    bot_x_z = bot_cc_z = bot_fs1_z = bot_fs2_z = bot_ac_z = bot_a1_z = bot_a2_z = bot_mc1_z = None
    if fx_bkN is not None and bot_depth and bot_width and bot_bottom_y and bot_As:
        # Here we pass b_mm = bot_depth and D_mm = bot_width (swap)
        bot_x_z = solve_x_for_equilibrium(fx_bkN, bot_depth, bot_width, bot_bottom_y, bot_As)
        if bot_x_z:
            bot_cc_z, bot_fs1_z, bot_fs2_z = compute_cc_fs(fx_bkN, bot_depth, bot_width, bot_bottom_y, bot_As, bot_x_z)
            # note: for lever arms we also use the swapped D (= bot_width) consistently
            bot_ac_z = 0.42 * bot_x_z - (bot_width / 2.0)
            bot_a1_z = bot_bottom_y - (bot_width / 2.0)
            bot_a2_z = (bot_width - bot_bottom_y) - (bot_width / 2.0)
            bot_mc1_z = (bot_cc_z * bot_ac_z) + (bot_fs1_z * bot_a1_z) + (bot_fs2_z * bot_a2_z)

    # ---- Top column Z-plane
    top_x_z = top_cc_z = top_fs1_z = top_fs2_z = top_ac_z = top_a1_z = top_a2_z = top_mc1_z = None
    if fx_tkN is not None and top_depth and top_width and top_bottom_y and top_As:
        # swap: b_mm = top_depth, D_mm = top_width
        top_x_z = solve_x_for_equilibrium(fx_tkN, top_depth, top_width, top_bottom_y, top_As)
        if top_x_z:
            top_cc_z, top_fs1_z, top_fs2_z = compute_cc_fs(fx_tkN, top_depth, top_width, top_bottom_y, top_As, top_x_z)
            top_ac_z = 0.42 * top_x_z - (top_width / 2.0)
            top_a1_z = top_bottom_y - (top_width / 2.0)
            top_a2_z = (top_width - top_bottom_y) - (top_width / 2.0)
            top_mc1_z = (top_cc_z * top_ac_z) + (top_fs1_z * top_a1_z) + (top_fs2_z * top_a2_z)

    # Store all results (including Z-plane fields)
    info.update({
        # bottom (Y-plane)
        "bot_bar_dia": bot_bar_dia, "bot_no_of_bars": bot_no_of_bars, "bot_bottom_y": bot_bottom_y,
        "bot_depth": bot_depth, "bot_width": bot_width, "bot_beta": bot_beta,
        "bot_A": bot_A, "bot_W": bot_W, "bot_sigma_plus": bot_sigma_plus, "bot_As": bot_As,
        "bot_x": bot_x, "bot_cc": bot_cc, "bot_fs1": bot_fs1, "bot_fs2": bot_fs2,
        "bot_ac": bot_ac, "bot_a1": bot_a1, "bot_a2": bot_a2, "bot_mc1": bot_mc1,
        # bottom (Z-plane)
        "bot_x_z": bot_x_z, "bot_cc_z": bot_cc_z, "bot_fs1_z": bot_fs1_z, "bot_fs2_z": bot_fs2_z,
        "bot_ac_z": bot_ac_z, "bot_a1_z": bot_a1_z, "bot_a2_z": bot_a2_z, "bot_mc1_z": bot_mc1_z,
        # top (Y-plane)
        "top_bar_dia": top_bar_dia, "top_no_of_bars": top_no_of_bars, "top_bottom_y": top_bottom_y,
        "top_depth": top_depth, "top_width": top_width, "top_beta": top_beta,
        "top_A": top_A, "top_W": top_W, "top_sigma_plus": top_sigma_plus, "top_As": top_As,
        "top_x": top_x, "top_cc": top_cc, "top_fs1": top_fs1, "top_fs2": top_fs2,
        "top_ac": top_ac, "top_a1": top_a1, "top_a2": top_a2, "top_mc1": top_mc1,
        # top (Z-plane)
        "top_x_z": top_x_z, "top_cc_z": top_cc_z, "top_fs1_z": top_fs1_z, "top_fs2_z": top_fs2_z,
        "top_ac_z": top_ac_z, "top_a1_z": top_a1_z, "top_a2_z": top_a2_z, "top_mc1_z": top_mc1_z
    })

print("✅ Computed top and bottom column equilibrium results for all matching nodes (Y-plane + Z-plane).")

# ============================================================
# STEP 5: WRITE TO EXCEL (headers include both top/bottom and Z-plane)
# ============================================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Column_Node_Map"

headers = [
    "Node ID", "Top Column ID", "Bottom Column ID",
    "Top Col Max Fx (kN)", "Top Col My (kN-m)", "Top Col Mz (kN-m)",
    "Bottom Col Max Fx (kN)", "Bottom Col My (kN-m)", "Bottom Col Mz (kN-m)",
    # ---- TOP COLUMN (Y-plane) ----
    "top_bar_dia", "top_no_of_bars", "top_bottom_y", "top_depth", "top_width", "top_beta",
    "top_A", "top_W", "top_sigma_plus", "top_As", "top_x", "top_cc", "top_fs1", "top_fs2",
    "top_ac", "top_a1", "top_a2", "top_mc1",
    # ---- TOP COLUMN (Z-plane) ----
    "top_x_z", "top_cc_z", "top_fs1_z", "top_fs2_z", "top_ac_z", "top_a1_z", "top_a2_z", "top_mc1_z",
    # ---- BOTTOM COLUMN (Y-plane) ----
    "bot_bar_dia", "bot_no_of_bars", "bot_bottom_y", "bot_depth", "bot_width", "bot_beta",
    "bot_A", "bot_W", "bot_sigma_plus", "bot_As", "bot_x", "bot_cc", "bot_fs1", "bot_fs2",
    "bot_ac", "bot_a1", "bot_a2", "bot_mc1",
    # ---- BOTTOM COLUMN (Z-plane) ----
    "bot_x_z", "bot_cc_z", "bot_fs1_z", "bot_fs2_z", "bot_ac_z", "bot_a1_z", "bot_a2_z", "bot_mc1_z"
]
ws.append(headers)

for nid, info in node_map.items():
    ws.append([
        info["node_id"], info["top_column_id"], info["bottom_column_id"],
        info["Top Fx"], info["Top My"], info["Top Mz"],
        info["Bottom Fx"], info["Bottom My"], info["Bottom Mz"],
        # top Y-plane
        info.get("top_bar_dia"), info.get("top_no_of_bars"), info.get("top_bottom_y"),
        info.get("top_depth"), info.get("top_width"), info.get("top_beta"),
        info.get("top_A"), info.get("top_W"), info.get("top_sigma_plus"),
        info.get("top_As"), info.get("top_x"), info.get("top_cc"), info.get("top_fs1"),
        info.get("top_fs2"), info.get("top_ac"), info.get("top_a1"), info.get("top_a2"),
        info.get("top_mc1"),
        # top Z-plane
        info.get("top_x_z"), info.get("top_cc_z"), info.get("top_fs1_z"), info.get("top_fs2_z"),
        info.get("top_ac_z"), info.get("top_a1_z"), info.get("top_a2_z"), info.get("top_mc1_z"),
        # bottom Y-plane
        info.get("bot_bar_dia"), info.get("bot_no_of_bars"), info.get("bot_bottom_y"),
        info.get("bot_depth"), info.get("bot_width"), info.get("bot_beta"),
        info.get("bot_A"), info.get("bot_W"), info.get("bot_sigma_plus"),
        info.get("bot_As"), info.get("bot_x"), info.get("bot_cc"), info.get("bot_fs1"),
        info.get("bot_fs2"), info.get("bot_ac"), info.get("bot_a1"), info.get("bot_a2"),
        info.get("bot_mc1"),
        # bottom Z-plane
        info.get("bot_x_z"), info.get("bot_cc_z"), info.get("bot_fs1_z"), info.get("bot_fs2_z"),
        info.get("bot_ac_z"), info.get("bot_a1_z"), info.get("bot_a2_z"), info.get("bot_mc1_z")
    ])

safe_save(wb, OUTPUT_XLSX)
print(f"✅ Output with computed top and bottom column equilibrium results (Y + Z planes) saved: {OUTPUT_XLSX}")
